import calendar
import csv
import io
from datetime import datetime, timedelta

import json
from flask import jsonify, request, send_file, session

from .app_state import app, db
from .auth import login_required, require_ledger_access, get_current_ledger_id
from .models import Account, Budget, BudgetCategoryItem, Category, Loan, MoneyChangeLog, RecurringRule, Transaction, User
from .support import log_money_change

def get_loan_base_query():
    """构建借贷基础查询（按权限过滤）"""
    user_id = session.get('user_id')
    is_admin = session.get('is_admin')
    self_view = session.get('self_view', False)
    current_ledger_id = get_current_ledger_id()

    query = Loan.query

    if current_ledger_id:
        query = query.filter(Loan.ledger_id == current_ledger_id)
    elif is_admin and not self_view:
        pass  # 管理员后台查看全部
    else:
        query = query.filter(Loan.user_id == user_id)

    return query


def serialize_loan(loan):
    """序列化借贷记录"""
    remaining = float(loan.amount) - float(loan.repaid_amount or 0)
    return {
        "id": loan.id,
        "type": loan.type,
        "counterparty": loan.counterparty,
        "amount": float(loan.amount),
        "repaid_amount": float(loan.repaid_amount or 0),
        "remaining": round(remaining, 2),
        "date": loan.date,
        "due_date": loan.due_date,
        "status": loan.status,
        "remark": loan.remark or '',
        "user_id": loan.user_id,
        "ledger_id": loan.ledger_id,
        "created_at": loan.created_at.isoformat() if loan.created_at else None,
        "updated_at": loan.updated_at.isoformat() if loan.updated_at else None
    }


@app.route('/api/loans', methods=['GET'])
@login_required
def list_loans():
    """获取借贷列表及汇总统计"""
    try:
        loan_type = request.args.get('type')  # 'borrow' or 'lend'
        status_filter = request.args.get('status')
        keyword = request.args.get('keyword')

        query = get_loan_base_query()

        if loan_type in ('borrow', 'lend'):
            query = query.filter(Loan.type == loan_type)
        if status_filter in ('active', 'settled'):
            query = query.filter(Loan.status == status_filter)
        if keyword:
            like_pattern = f'%{keyword}%'
            query = query.filter(
                db.or_(Loan.counterparty.like(like_pattern), db.func.ifnull(Loan.remark, '').like(like_pattern))
            )

        query = query.order_by(Loan.id.desc())
        loans = query.all()

        # 汇总统计
        total_borrow = 0.0      # 总借入（负债）
        total_borrow_remaining = 0.0
        total_lend = 0.0        # 总借出（债权）
        total_lend_remaining = 0.0

        for loan in loans:
            amount = float(loan.amount)
            remaining = amount - float(loan.repaid_amount or 0)
            if loan.type == 'borrow':
                total_borrow += amount
                total_borrow_remaining += remaining
            else:
                total_lend += amount
                total_lend_remaining += remaining

        return jsonify({
            "success": True,
            "loans": [serialize_loan(l) for l in loans],
            "total": len(loans),
            "summary": {
                "total_borrow": round(total_borrow, 2),
                "total_borrow_remaining": round(total_borrow_remaining, 2),
                "total_lend": round(total_lend, 2),
                "total_lend_remaining": round(total_lend_remaining, 2)
            }
        })
    except Exception as e:
        app.logger.error(f"获取借贷列表失败: {e}")
        return jsonify({"success": False, "message": "获取借贷列表失败"}), 500


@app.route('/api/loans', methods=['POST'])
@login_required
def create_loan():
    """创建借贷记录"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')

        # 权限检查
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            has_access, role, error = require_ledger_access(current_ledger_id, 'editor')
            if not has_access:
                return error

        # 字段校验
        loan_type = data.get('type')
        if loan_type not in ('borrow', 'lend'):
            return jsonify({"success": False, "message": "借贷类型必须是 borrow 或 lend"}), 400

        counterparty = (data.get('counterparty') or '').strip()
        if not counterparty:
            return jsonify({"success": False, "message": "对方名称不能为空"}), 400

        try:
            amount = float(data.get('amount', 0))
            if amount <= 0:
                return jsonify({"success": False, "message": "金额必须大于0"}), 400
            if amount > 99999999.99:
                return jsonify({"success": False, "message": "金额超出限制"}), 400
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "金额格式不正确"}), 400

        loan_date = (data.get('date') or '').strip()
        if not loan_date:
            loan_date = datetime.now().strftime('%Y-%m-%d')

        due_date = (data.get('due_date') or '').strip() or None

        remark = (data.get('remark') or '').strip()

        loan = Loan(
            type=loan_type,
            counterparty=counterparty,
            amount=amount,
            repaid_amount=0,
            date=loan_date,
            due_date=due_date,
            status='active',
            remark=remark,
            user_id=user_id,
            ledger_id=current_ledger_id
        )
        db.session.add(loan)
        db.session.flush()

        # 更新账户余额：lend=借出(余额减少), borrow=借入(余额增加)
        account = Account.query.filter_by(ledger_id=current_ledger_id, user_id=user_id).first()
        if not account:
            account = Account.query.filter_by(user_id=user_id).first()
        if account:
            balance_before = float(account.balance or 0)
            change = -amount if loan_type == 'lend' else amount
            account.balance = balance_before + change
            log_money_change(
                user_id=user_id,
                action_type='create',
                entity_type='loan',
                entity_id=loan.id,
                amount_change=change,
                balance_before=balance_before,
                balance_after=float(account.balance),
                account_id=account.id,
                ledger_id=current_ledger_id,
                description=f'创建借贷{"借出" if loan_type == "lend" else "借入"} ￥{amount:.2f} - {counterparty}'
            )

        db.session.commit()

        app.logger.info(f"新增借贷: {loan_type} ￥{amount} - {counterparty}")
        return jsonify({
            "success": True,
            "message": "借贷记录添加成功",
            "loan": serialize_loan(loan)
        }), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建借贷记录失败: {e}")
        return jsonify({"success": False, "message": "创建借贷记录失败"}), 500


@app.route('/api/loans/<int:loan_id>', methods=['PUT'])
@login_required
def update_loan(loan_id):
    """更新借贷记录"""
    try:
        loan = Loan.query.get(loan_id)
        if not loan:
            return jsonify({"success": False, "message": "未找到该借贷记录"}), 404

        # 权限检查
        user_id = session.get('user_id')
        if loan.user_id != user_id:
            if loan.ledger_id:
                has_access, role, error = require_ledger_access(loan.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限修改此记录"}), 403

        data = request.get_json()

        if 'type' in data:
            if data['type'] not in ('borrow', 'lend'):
                return jsonify({"success": False, "message": "借贷类型必须是 borrow 或 lend"}), 400
            loan.type = data['type']

        if 'counterparty' in data:
            counterparty = data['counterparty'].strip()
            if not counterparty:
                return jsonify({"success": False, "message": "对方名称不能为空"}), 400
            loan.counterparty = counterparty

        if 'amount' in data:
            try:
                amount = float(data['amount'])
                if amount <= 0:
                    return jsonify({"success": False, "message": "金额必须大于0"}), 400
                if amount > 99999999.99:
                    return jsonify({"success": False, "message": "金额超出限制"}), 400
                # 检查已还金额不能超过新总额
                if float(loan.repaid_amount or 0) > amount:
                    return jsonify({"success": False, "message": "已还金额不能大于总金额"}), 400
                loan.amount = amount
            except (TypeError, ValueError):
                return jsonify({"success": False, "message": "金额格式不正确"}), 400

        if 'date' in data:
            loan.date = data['date']
        if 'due_date' in data:
            loan.due_date = data['due_date'].strip() or None
        if 'status' in data:
            if data['status'] not in ('active', 'settled'):
                return jsonify({"success": False, "message": "状态无效"}), 400
            loan.status = data['status']
        if 'remark' in data:
            loan.remark = data['remark'].strip()

        loan.updated_at = datetime.now()
        db.session.commit()

        app.logger.info(f"更新借贷: ID={loan_id}")
        return jsonify({
            "success": True,
            "message": "借贷记录已更新",
            "loan": serialize_loan(loan)
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新借贷记录失败: {e}")
        return jsonify({"success": False, "message": "更新借贷记录失败"}), 500


@app.route('/api/loans/<int:loan_id>', methods=['DELETE'])
@login_required
def delete_loan(loan_id):
    """删除借贷记录"""
    try:
        loan = Loan.query.get(loan_id)
        if not loan:
            return jsonify({"success": False, "message": "未找到该借贷记录"}), 404

        # 权限检查
        user_id = session.get('user_id')
        if loan.user_id != user_id:
            if loan.ledger_id:
                has_access, role, error = require_ledger_access(loan.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限删除此记录"}), 403

        db.session.delete(loan)
        db.session.commit()

        app.logger.info(f"删除借贷: ID={loan_id}")
        return jsonify({
            "success": True,
            "message": "借贷记录已删除"
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除借贷记录失败: {e}")
        return jsonify({"success": False, "message": "删除借贷记录失败"}), 500


@app.route('/api/loans/<int:loan_id>/repay', methods=['POST'])
@login_required
def repay_loan(loan_id):
    """记录还款/收款"""
    try:
        loan = Loan.query.get(loan_id)
        if not loan:
            return jsonify({"success": False, "message": "未找到该借贷记录"}), 404

        # 权限检查
        user_id = session.get('user_id')
        if loan.user_id != user_id:
            if loan.ledger_id:
                has_access, role, error = require_ledger_access(loan.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限操作此记录"}), 403

        if loan.status == 'settled':
            return jsonify({"success": False, "message": "该借贷已结清，无法继续还款"}), 400

        data = request.get_json()
        try:
            repay_amount = float(data.get('amount', 0))
            if repay_amount <= 0:
                return jsonify({"success": False, "message": "还款金额必须大于0"}), 400
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "金额格式不正确"}), 400

        current_repaid = float(loan.repaid_amount or 0)
        new_total_repaid = current_repaid + repay_amount
        total_amount = float(loan.amount)

        # 超额还款检查
        if new_total_repaid > total_amount:
            return jsonify({
                "success": False,
                "message": f"超额还款错误：累计还款 ￥{new_total_repaid:.2f} 超出借贷总额 ￥{total_amount:.2f}，多出 ￥{new_total_repaid - total_amount:.2f}"
            }), 400

        loan.repaid_amount = new_total_repaid

        # 如果已还清，自动标记为已结清
        if new_total_repaid >= total_amount:
            loan.status = 'settled'

        # 更新账户余额：借出收回(余额增加), 借入还款(余额减少)
        change = repay_amount if loan.type == 'lend' else -repay_amount
        account = Account.query.filter_by(ledger_id=loan.ledger_id, user_id=user_id).first()
        if not account:
            account = Account.query.filter_by(user_id=user_id).first()
        balance_before = float(account.balance or 0) if account else None
        if account:
            account.balance = balance_before + change

        loan.updated_at = datetime.now()
        db.session.commit()

        action_label = '收款' if loan.type == 'lend' else '还款'
        app.logger.info(f"借贷{action_label}: loan={loan_id}, amount={repay_amount}")

        log_money_change(
            user_id=user_id,
            action_type='repay',
            entity_type='loan',
            entity_id=loan_id,
            amount_change=change,
            balance_before=balance_before,
            balance_after=float(account.balance) if account else None,
            account_id=account.id if account else None,
            ledger_id=loan.ledger_id,
            description=f'借贷{action_label} ￥{repay_amount:.2f} - {loan.counterparty}'
        )

        action_msg = '收款' if loan.type == 'lend' else '还款'
        return jsonify({
            "success": True,
            "message": f"{action_msg}记录成功",
            "loan": serialize_loan(loan)
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"还款操作失败: {e}")
        return jsonify({"success": False, "message": "还款操作失败"}), 500


@app.route('/api/loans/summary', methods=['GET'])
@login_required
def loan_summary():
    """获取借贷汇总（用于资产/负债统计）"""
    try:
        query = get_loan_base_query()
        loans = query.all()

        total_borrow_remaining = 0.0  # 总负债
        total_lend_remaining = 0.0    # 总债权

        for loan in loans:
            remaining = float(loan.amount) - float(loan.repaid_amount or 0)
            if remaining > 0:
                if loan.type == 'borrow':
                    total_borrow_remaining += remaining
                else:
                    total_lend_remaining += remaining

        net_loan_asset = round(total_lend_remaining - total_borrow_remaining, 2)

        return jsonify({
            "success": True,
            "total_borrow_remaining": round(total_borrow_remaining, 2),
            "total_lend_remaining": round(total_lend_remaining, 2),
            "net_loan_asset": net_loan_asset
        })
    except Exception as e:
        app.logger.error(f"获取借贷汇总失败: {e}")
        return jsonify({"success": False, "message": "获取借贷汇总失败"}), 500


# ==================== 周期账单 API ====================

def _compute_next_date(current_date_str, period, interval_value=1):
    """根据周期计算下一次生成日期"""
    from datetime import datetime as dt
    from dateutil.relativedelta import relativedelta
    cur = dt.strptime(current_date_str, '%Y-%m-%d').date()
    if period == 'daily':
        next_d = cur + relativedelta(days=interval_value)
    elif period == 'weekly':
        next_d = cur + relativedelta(weeks=interval_value)
    elif period == 'monthly':
        next_d = cur + relativedelta(months=interval_value)
    elif period == 'yearly':
        next_d = cur + relativedelta(years=interval_value)
    else:
        next_d = cur + relativedelta(months=1)
    return next_d.strftime('%Y-%m-%d')


def _generate_rule_transaction(rule):
    """为一条规则生成一笔交易，更新 next_date。返回生成的 Transaction 或 None"""
    today_str = datetime.now().strftime('%Y-%m-%d')
    if rule.next_date > today_str:
        return None
    if rule.end_date and rule.end_date < today_str:
        return None
    if not rule.is_active:
        return None

    account = Account.query.get(rule.account_id) if rule.account_id else None
    if not account:
        account = Account.query.filter_by(ledger_id=rule.ledger_id).first()
    if not account:
        account = Account.query.filter_by(user_id=rule.user_id).first()
    if not account:
        account = Account(name='默认账户', balance=0, account_type='cash',
                          user_id=rule.user_id, ledger_id=rule.ledger_id)
        db.session.add(account)
        db.session.flush()

    now = datetime.now()
    tx = Transaction(
        type=rule.type,
        amount=float(rule.amount),
        category=rule.category,
        date=rule.next_date,
        time=now.strftime('%H:%M:%S'),
        remark=f'[周期] {rule.name}' + (f' - {rule.remark}' if rule.remark else ''),
        account_id=account.id,
        user_id=rule.user_id,
        ledger_id=rule.ledger_id,
        currency='CNY',
    )
    db.session.add(tx)

    if rule.type == 'income':
        account.balance = float(account.balance) + float(rule.amount)
        amt_change = float(rule.amount)
    else:
        account.balance = float(account.balance) - float(rule.amount)
        amt_change = -float(rule.amount)

    balance_after = float(account.balance)
    log_money_change(
        user_id=rule.user_id,
        action_type='create',
        entity_type='transaction',
        amount_change=amt_change,
        balance_before=balance_after - amt_change,
        balance_after=balance_after,
        account_id=account.id,
        ledger_id=rule.ledger_id,
        description=f'周期规则生成{"收入" if rule.type == "income" else "支出"} ￥{float(rule.amount):.2f} - {rule.name}'
    )

    # 计算下一次生成日期
    old_next = rule.next_date
    rule.next_date = _compute_next_date(old_next, rule.period, rule.interval_value or 1)
    rule.updated_at = now

    # 如果下次日期已过，继续往前递推（补生）
    while rule.next_date < today_str and (not rule.end_date or rule.next_date <= rule.end_date):
        # 生成中间日期的交易
        tx_mid = Transaction(
            type=rule.type,
            amount=float(rule.amount),
            category=rule.category,
            date=rule.next_date,
            time=now.strftime('%H:%M:%S'),
            remark=f'[周期] {rule.name}' + (f' - {rule.remark}' if rule.remark else ''),
            account_id=account.id,
            user_id=rule.user_id,
            ledger_id=rule.ledger_id,
            currency='CNY',
        )
        db.session.add(tx_mid)
        if rule.type == 'income':
            account.balance = float(account.balance) + float(rule.amount)
            catch_amt = float(rule.amount)
        else:
            account.balance = float(account.balance) - float(rule.amount)
            catch_amt = -float(rule.amount)
        catch_after = float(account.balance)
        log_money_change(
            user_id=rule.user_id,
            action_type='create',
            entity_type='transaction',
            amount_change=catch_amt,
            balance_before=catch_after - catch_amt,
            balance_after=catch_after,
            account_id=account.id,
            ledger_id=rule.ledger_id,
            description=f'周期规则补生{"收入" if rule.type == "income" else "支出"} ￥{float(rule.amount):.2f} - {rule.name} ({rule.next_date})'
        )
        rule.next_date = _compute_next_date(rule.next_date, rule.period, rule.interval_value or 1)
        rule.updated_at = now

    # 如果超出结束日期则停用
    if rule.end_date and rule.next_date > rule.end_date:
        rule.is_active = False

    return tx


def serialize_rule(rule):
    """序列化周期规则"""
    return {
        'id': rule.id,
        'name': rule.name,
        'amount': float(rule.amount),
        'category': rule.category,
        'type': rule.type,
        'period': rule.period,
        'interval_value': rule.interval_value or 1,
        'start_date': rule.start_date,
        'end_date': rule.end_date,
        'next_date': rule.next_date,
        'is_active': rule.is_active,
        'remark': rule.remark or '',
        'account_id': rule.account_id,
        'account_name': rule.account.name if rule.account else None,
        'user_id': rule.user_id,
        'ledger_id': rule.ledger_id,
        'created_at': rule.created_at.isoformat() if rule.created_at else None,
        'updated_at': rule.updated_at.isoformat() if rule.updated_at else None,
    }


@app.route('/api/recurring-rules', methods=['GET'])
@login_required
def list_recurring_rules():
    """获取周期账单规则列表"""
    try:
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()

        query = RecurringRule.query
        if current_ledger_id:
            query = query.filter(RecurringRule.ledger_id == current_ledger_id)
        elif is_admin and not self_view:
            pass
        else:
            query = query.filter(RecurringRule.user_id == user_id)

        query = query.order_by(RecurringRule.id.desc())
        rules = query.all()

        return jsonify({
            'success': True,
            'rules': [serialize_rule(r) for r in rules],
            'total': len(rules),
        })
    except Exception as e:
        app.logger.error(f"获取周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '获取周期账单规则失败'}), 500


@app.route('/api/recurring-rules', methods=['POST'])
@login_required
def create_recurring_rule():
    """创建周期账单规则"""
    try:
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({'success': False, 'message': '管理员不能创建规则'}), 403

        user_id = session.get('user_id')
        data = request.get_json()
        current_ledger_id = get_current_ledger_id()

        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': '账单名称不能为空'}), 400

        tx_type = data.get('type')
        if tx_type not in ('income', 'expense'):
            return jsonify({'success': False, 'message': '类型必须是 income 或 expense'}), 400

        try:
            amount = float(data.get('amount', 0))
            if amount <= 0:
                return jsonify({'success': False, 'message': '金额必须大于0'}), 400
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': '金额格式不正确'}), 400

        category = (data.get('category') or '').strip()
        if not category:
            return jsonify({'success': False, 'message': '分类不能为空'}), 400
        cat = Category.query.filter_by(name=category, type=tx_type).first()
        if not cat:
            return jsonify({'success': False, 'message': f'分类「{category}」不存在或类型不匹配'}), 400

        period = data.get('period')
        if period not in ('daily', 'weekly', 'monthly', 'yearly'):
            return jsonify({'success': False, 'message': '周期必须是 daily/weekly/monthly/yearly'}), 400

        interval_value = int(data.get('interval_value', 1))
        if interval_value < 1:
            interval_value = 1

        start_date = (data.get('start_date') or '').strip()
        if not start_date:
            start_date = datetime.now().strftime('%Y-%m-%d')
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': '开始日期格式不正确'}), 400

        end_date = (data.get('end_date') or '').strip()
        if end_date:
            try:
                datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'message': '结束日期格式不正确'}), 400
        else:
            end_date = None

        if end_date and end_date < start_date:
            return jsonify({'success': False, 'message': '结束日期不能早于开始日期'}), 400

        account_id = data.get('account_id')
        if account_id:
            acct = Account.query.get(account_id)
            if not acct:
                return jsonify({'success': False, 'message': '账户不存在'}), 400
        else:
            account_id = None

        remark = (data.get('remark') or '').strip()

        rule = RecurringRule(
            name=name,
            amount=amount,
            category=category,
            type=tx_type,
            period=period,
            interval_value=interval_value,
            start_date=start_date,
            end_date=end_date,
            next_date=start_date,
            is_active=True,
            remark=remark,
            account_id=account_id,
            user_id=user_id,
            ledger_id=current_ledger_id,
        )
        db.session.add(rule)
        db.session.commit()

        app.logger.info(f"创建周期账单规则: {name}, {tx_type} ￥{amount}, {period}")
        return jsonify({'success': True, 'message': '周期账单规则创建成功', 'rule': serialize_rule(rule)}), 201

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '创建失败'}), 500


@app.route('/api/recurring-rules/<int:rule_id>', methods=['PUT'])
@login_required
def update_recurring_rule(rule_id):
    """更新周期账单规则"""
    try:
        rule = RecurringRule.query.get(rule_id)
        if not rule:
            return jsonify({'success': False, 'message': '规则不存在'}), 404

        user_id = session.get('user_id')
        if rule.user_id != user_id:
            if rule.ledger_id:
                has_access, role, error = require_ledger_access(rule.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({'success': False, 'message': '无权限修改'}), 403

        data = request.get_json()

        if 'name' in data:
            name = data['name'].strip()
            if name:
                rule.name = name
        if 'amount' in data:
            try:
                amt = float(data['amount'])
                if amt > 0:
                    rule.amount = amt
            except (TypeError, ValueError):
                pass
        if 'category' in data:
            cat = data['category'].strip()
            if cat:
                rule.category = cat
        if 'type' in data and data['type'] in ('income', 'expense'):
            rule.type = data['type']
        if 'period' in data and data['period'] in ('daily', 'weekly', 'monthly', 'yearly'):
            rule.period = data['period']
        if 'interval_value' in data:
            iv = int(data['interval_value'])
            if iv >= 1:
                rule.interval_value = iv
        if 'start_date' in data:
            sd = data['start_date'].strip()
            if sd:
                rule.start_date = sd
        if 'end_date' in data:
            ed = data['end_date'].strip()
            rule.end_date = ed or None
        if 'is_active' in data:
            rule.is_active = bool(data['is_active'])
        if 'remark' in data:
            rule.remark = data['remark'].strip()
        if 'account_id' in data:
            rule.account_id = data['account_id'] or None

        rule.updated_at = datetime.now()
        db.session.commit()

        return jsonify({'success': True, 'message': '规则已更新', 'rule': serialize_rule(rule)})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '更新失败'}), 500


@app.route('/api/recurring-rules/<int:rule_id>', methods=['DELETE'])
@login_required
def delete_recurring_rule(rule_id):
    """删除周期账单规则"""
    try:
        rule = RecurringRule.query.get(rule_id)
        if not rule:
            return jsonify({'success': False, 'message': '规则不存在'}), 404

        user_id = session.get('user_id')
        if rule.user_id != user_id:
            if rule.ledger_id:
                has_access, role, error = require_ledger_access(rule.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({'success': False, 'message': '无权限删除'}), 403

        db.session.delete(rule)
        db.session.commit()

        return jsonify({'success': True, 'message': '规则已删除'})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '删除失败'}), 500


@app.route('/api/recurring-rules/<int:rule_id>/toggle', methods=['POST'])
@login_required
def toggle_recurring_rule(rule_id):
    """启用/停用周期账单规则"""
    try:
        rule = RecurringRule.query.get(rule_id)
        if not rule:
            return jsonify({'success': False, 'message': '规则不存在'}), 404

        user_id = session.get('user_id')
        if rule.user_id != user_id:
            if rule.ledger_id:
                has_access, role, error = require_ledger_access(rule.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({'success': False, 'message': '无权限操作'}), 403

        rule.is_active = not rule.is_active
        rule.updated_at = datetime.now()
        db.session.commit()

        status = '已启用' if rule.is_active else '已停用'
        return jsonify({'success': True, 'message': f'规则{status}', 'is_active': rule.is_active})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"切换周期账单规则状态失败: {e}")
        return jsonify({'success': False, 'message': '操作失败'}), 500


@app.route('/api/recurring-rules/generate', methods=['POST'])
@login_required
def generate_recurring_bills():
    """手动触发周期账单生成"""
    try:
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()
        today_str = datetime.now().strftime('%Y-%m-%d')

        query = RecurringRule.query.filter(
            RecurringRule.is_active == True,
            RecurringRule.next_date <= today_str,
        )
        if current_ledger_id:
            query = query.filter(RecurringRule.ledger_id == current_ledger_id)
        elif not (is_admin and not self_view):
            query = query.filter(RecurringRule.user_id == user_id)

        # Also filter: end_date >= today or end_date is null
        query = query.filter(
            db.or_(RecurringRule.end_date >= today_str, RecurringRule.end_date.is_(None))
        )

        rules = query.all()
        generated = 0

        for rule in rules:
            tx = _generate_rule_transaction(rule)
            if tx:
                generated += 1

        db.session.commit()

        if generated > 0:
            app.logger.info(f"周期账单生成: 用户={user_id}, 生成={generated}笔")
            return jsonify({
                'success': True,
                'message': f'成功生成 {generated} 笔周期账单',
                'generated_count': generated,
            })
        else:
            return jsonify({
                'success': True,
                'message': '没有待生成的周期账单',
                'generated_count': 0,
            })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"生成周期账单失败: {e}")
        return jsonify({'success': False, 'message': '生成失败'}), 500


@app.route('/api/recurring-rules/generate-check', methods=['GET'])
@login_required
def check_recurring_pending():
    """检查是否有待生成的周期账单"""
    try:
        user_id = session.get('user_id')
        current_ledger_id = get_current_ledger_id()
        today_str = datetime.now().strftime('%Y-%m-%d')

        query = RecurringRule.query.filter(
            RecurringRule.is_active == True,
            RecurringRule.next_date <= today_str,
        )
        if current_ledger_id:
            query = query.filter(RecurringRule.ledger_id == current_ledger_id)
        else:
            query = query.filter(RecurringRule.user_id == user_id)

        query = query.filter(
            db.or_(RecurringRule.end_date >= today_str, RecurringRule.end_date.is_(None))
        )

        count = query.count()
        return jsonify({'success': True, 'pending_count': count})

    except Exception as e:
        app.logger.error(f"检查周期账单待生成失败: {e}")
        return jsonify({'success': False, 'message': '检查失败'}), 500


@app.route('/api/recurring-rules/scheduled-generate', methods=['POST'])
def scheduled_generate():
    """定时任务自动生成（允许未登录调用，但只处理全局规则）"""
    try:
        from datetime import date
        today_str = date.today().strftime('%Y-%m-%d')

        rules = RecurringRule.query.filter(
            RecurringRule.is_active == True,
            RecurringRule.next_date <= today_str,
            db.or_(RecurringRule.end_date >= today_str, RecurringRule.end_date.is_(None))
        ).all()

        generated = 0
        for rule in rules:
            tx = _generate_rule_transaction(rule)
            if tx:
                generated += 1

        db.session.commit()
        app.logger.info(f"[定时] 周期账单生成: {generated}笔")
        return jsonify({'success': True, 'generated_count': generated})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[定时] 周期账单生成失败: {e}")
        return jsonify({'success': False, 'message': '生成失败'}), 500


@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    try:
        cat_type = request.args.get('type', None)
        query = Category.query
        if cat_type and cat_type in ['income', 'expense']:
            query = query.filter_by(type=cat_type)
        categories = query.all()
        if not categories:
            # 如无分类自动写入默认分类
            default_income = ['工资', '奖金', '投资收益', '兼职', '红包', '报销收入', '其他收入']
            default_expense = ['餐饮', '交通', '购物', '娱乐', '医疗', '住房', '教育', '通讯', '其他支出']
            for cname in default_income:
                db.session.add(Category(name=cname, type='income'))
            for cname in default_expense:
                db.session.add(Category(name=cname, type='expense'))
            db.session.commit()
            categories = Category.query.all()
        return jsonify({
            "success": True,
            "categories": [{"id": c.id, "name": c.name, "type": c.type} for c in categories]
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"获取分类失败: {e}")
        abort(500, description="获取分类失败")


# ==================== 预算管理 API ====================
@app.route('/api/budgets/current', methods=['GET'])
@login_required
def get_current_budget():
    """获取当前月份的预算及使用情况"""
    try:
        user_id = session.get('user_id')
        now = datetime.now()
        month = now.strftime('%Y-%m')
        account_id = request.args.get('account_id', None, type=int)

        # 查找当月预算（用 ledger 或 user 过滤）
        current_ledger_id = get_current_ledger_id()
        budget_filter = {'user_id': user_id, 'month': month}
        if current_ledger_id:
            budget_filter['ledger_id'] = current_ledger_id
        if account_id:
            budget_filter['account_id'] = account_id
        else:
            budget_filter['account_id'] = None
        budget = Budget.query.filter_by(**budget_filter).first()

        # 计算当月支出（按分类）
        expense_filter = [
            Transaction.type == 'expense',
            Transaction.date.startswith(month)
        ]
        if current_ledger_id:
            expense_filter.append(Transaction.ledger_id == current_ledger_id)
        else:
            expense_filter.append(Transaction.user_id == user_id)
        if account_id:
            expense_filter.append(Transaction.account_id == account_id)

        expenses = Transaction.query.filter(*expense_filter).all()

        total_expense = float(sum(t.amount for t in expenses))

        # 按分类统计支出
        category_expenses = {}
        for t in expenses:
            cat = t.category
            category_expenses[cat] = category_expenses.get(cat, 0) + float(t.amount)

        # 账户名称
        account_name = None
        if account_id:
            acct = Account.query.get(account_id)
            account_name = acct.name if acct else None

        result = {
            'month': month,
            'account_id': account_id,
            'account_name': account_name or '总账户',
            'total_expense': round(total_expense, 2),
            'category_expenses': {k: round(v, 2) for k, v in category_expenses.items()}
        }

        if budget:
            total_amount = float(budget.total_amount)
            remaining = round(total_amount - total_expense, 2)
            progress = round(min(100, total_expense / total_amount * 100), 1) if total_amount > 0 else 0

            cat_items = BudgetCategoryItem.query.filter_by(budget_id=budget.id).all()
            items_data = []
            for item in cat_items:
                cat_name = item.category.name if item.category else '未知'
                used = round(category_expenses.get(cat_name, 0), 2)
                item_amount = float(item.amount)
                items_data.append({
                    'id': item.id,
                    'category_id': item.category_id,
                    'category_name': cat_name,
                    'amount': item_amount,
                    'used': used,
                    'remaining': round(item_amount - used, 2),
                    'progress': round(min(100, used / item_amount * 100), 1) if item_amount > 0 else 0
                })

            result['budget'] = {
                'id': budget.id,
                'account_id': budget.account_id,
                'total_amount': total_amount,
                'remaining': remaining,
                'progress': progress,
                'is_over': total_expense > total_amount,
                'remark': budget.remark or '',
                'category_items': items_data
            }
        else:
            result['budget'] = None

        return jsonify({'success': True, 'data': result})
    except Exception as e:
        app.logger.error(f"获取预算失败: {e}")
        return jsonify({'success': False, 'message': '获取预算失败'}), 500


@app.route('/api/budgets/list', methods=['GET'])
@login_required
def list_budgets():
    """获取当前用户所有月预算列表（含账户信息）"""
    try:
        user_id = session.get('user_id')
        now = datetime.now()
        month = now.strftime('%Y-%m')
        current_ledger_id = get_current_ledger_id()
        budget_filter = {'user_id': user_id, 'month': month}
        if current_ledger_id:
            budget_filter['ledger_id'] = current_ledger_id
        budgets = Budget.query.filter_by(**budget_filter).all()
        result = []
        for b in budgets:
            acct_name = b.account.name if b.account else '总账户'
            result.append({
                'id': b.id,
                'account_id': b.account_id,
                'account_name': acct_name,
                'total_amount': float(b.total_amount),
                'remark': b.remark or ''
            })
        return jsonify({'success': True, 'budgets': result})
    except Exception as e:
        app.logger.error(f"获取预算列表失败: {e}")
        return jsonify({'success': False, 'message': '获取预算列表失败'}), 500


@app.route('/api/budgets', methods=['POST'])
@login_required
def save_budget():
    """创建或更新当月预算"""
    try:
        # 管理员在 self_view 模式下可以管理预算
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({'success': False, 'message': '管理员只能查看数据，不能管理预算'}), 403

        user_id = session.get('user_id')
        data = request.get_json()

        month = data.get('month', datetime.now().strftime('%Y-%m'))
        total_amount = data.get('total_amount', 0)
        remark = data.get('remark', '').strip()
        account_id = data.get('account_id', None)
        category_items = data.get('category_items', [])
        current_ledger_id = get_current_ledger_id()

        # 校验账户（account_id 为 0 或空时视为总账户）
        if account_id:
            acct = Account.query.filter_by(id=account_id).first()
            if not acct or (current_ledger_id and acct.ledger_id != current_ledger_id):
                return jsonify({'success': False, 'message': '账户不存在'}), 400
        else:
            account_id = None

        try:
            total_amount = float(total_amount)
            if total_amount <= 0:
                return jsonify({'success': False, 'message': '预算金额必须大于0'}), 400
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': '预算金额格式不正确'}), 400

        # 查找是否已有该月该账户的预算
        budget_filter = {'user_id': user_id, 'month': month, 'account_id': account_id}
        if current_ledger_id:
            budget_filter['ledger_id'] = current_ledger_id
        budget = Budget.query.filter_by(**budget_filter).first()
        if budget:
            budget.total_amount = total_amount
            budget.remark = remark
            budget.updated_at = datetime.now()
        else:
            budget = Budget(user_id=user_id, month=month, total_amount=total_amount, remark=remark, account_id=account_id, ledger_id=current_ledger_id)
            db.session.add(budget)

        db.session.flush()

        # 删除旧的分类预算项
        BudgetCategoryItem.query.filter_by(budget_id=budget.id).delete()

        # 添加新的分类预算项
        for item in category_items:
            cat_id = item.get('category_id')
            cat_amount = item.get('amount', 0)
            if not cat_id or not cat_amount:
                continue
            try:
                cat_amount = float(cat_amount)
                if cat_amount <= 0:
                    continue
            except (TypeError, ValueError):
                continue

            cat_item = BudgetCategoryItem(
                budget_id=budget.id,
                category_id=cat_id,
                amount=cat_amount
            )
            db.session.add(cat_item)

        db.session.commit()

        acct_label = account_id or '总账户'
        app.logger.info(f"保存预算: user={user_id}, month={month}, account={acct_label}, total={total_amount}")
        return jsonify({'success': True, 'message': '预算保存成功', 'budget_id': budget.id, 'account_id': account_id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"保存预算失败: {e}")
        return jsonify({'success': False, 'message': '保存预算失败'}), 500


@app.route('/api/budgets/<int:budget_id>', methods=['DELETE'])
@login_required
def delete_budget(budget_id):
    """删除预算"""
    try:
        # 管理员在 self_view 模式下可以删除预算
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({'success': False, 'message': '管理员只能查看数据，不能删除预算'}), 403

        user_id = session.get('user_id')
        budget = Budget.query.filter_by(id=budget_id, user_id=user_id).first()
        if not budget:
            return jsonify({'success': False, 'message': '未找到该预算'}), 404

        # 删除关联的分类预算项
        BudgetCategoryItem.query.filter_by(budget_id=budget.id).delete()
        db.session.delete(budget)
        db.session.commit()

        app.logger.info(f"删除预算: user={user_id}, budget_id={budget_id}")
        return jsonify({'success': True, 'message': '预算已删除'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除预算失败: {e}")
        return jsonify({'success': False, 'message': '删除预算失败'}), 500


# ==================== 账户管理 API ====================
@app.route('/api/accounts', methods=['GET'])
@login_required
def get_accounts():
    try:
        user_id = session.get('user_id')
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            accounts = Account.query.filter_by(ledger_id=current_ledger_id).all()
        else:
            accounts = Account.query.filter_by(user_id=user_id).all()
        if not accounts:
            if current_ledger_id:
                default_account = Account(name='默认账户', balance=0, account_type='cash', user_id=user_id, ledger_id=current_ledger_id)
            else:
                default_account = Account(name='默认账户', balance=0, account_type='cash', user_id=user_id)
            db.session.add(default_account)
            db.session.commit()
            accounts = [default_account]
        return jsonify({
            "success": True,
            "accounts": [
                {
                    "id": a.id,
                    "name": a.name,
                    "balance": float(a.balance),
                    "account_type": a.account_type,
                    "created_at": a.created_at.isoformat() if a.created_at else None
                } for a in accounts
            ]
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"获取账户失败: {e}")
        abort(500, description="获取账户失败")


@app.route('/api/accounts', methods=['POST'])
@login_required
def create_account():
    try:
        # 管理员在 self_view 模式下可以创建账户
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({"success": False, "message": "管理员只能查看数据，不能创建账户"}), 403

        data = request.get_json()
        name = data.get('name', '').strip()
        balance = data.get('balance', 0)
        account_type = data.get('account_type', 'cash')
        user_id = session.get('user_id')
        current_ledger_id = get_current_ledger_id()

        if not name:
            return jsonify({"success": False, "message": "账户名称不能为空"}), 400
        if account_type not in ['cash', 'bank', 'credit', 'other']:
            return jsonify({"success": False, "message": "账户类型无效"}), 400

        account = Account(name=name, balance=balance, account_type=account_type, user_id=user_id, ledger_id=current_ledger_id)
        db.session.add(account)
        db.session.commit()
        return jsonify({
            "success": True,
            "message": "账户创建成功",
            "account": {
                "id": account.id,
                "name": account.name,
                "balance": float(account.balance),
                "account_type": account.account_type
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建账户失败: {e}")
        return jsonify({"success": False, "message": "创建账户失败"}), 500


@app.route('/api/accounts/<int:account_id>', methods=['PUT'])
@login_required
def update_account(account_id):
    try:
        # 管理员在 self_view 模式下可以修改账户
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({"success": False, "message": "管理员只能查看数据，不能修改账户"}), 403

        user_id = session.get('user_id')
        account = Account.query.filter_by(id=account_id, user_id=user_id).first()
        if not account:
            return jsonify({"success": False, "message": "账户不存在"}), 404

        data = request.get_json()
        old_balance = float(account.balance)
        if 'name' in data:
            account.name = data['name'].strip()
        if 'account_type' in data and data['account_type'] in ['cash', 'bank', 'credit', 'other']:
            account.account_type = data['account_type']
        if 'balance' in data:
            account.balance = data['balance']

        db.session.commit()

        new_balance = float(account.balance)
        if 'balance' in data and abs(new_balance - old_balance) > 0.001:
            log_money_change(
                user_id=user_id,
                action_type='adjust',
                entity_type='account',
                entity_id=account_id,
                amount_change=round(new_balance - old_balance, 2),
                balance_before=old_balance,
                balance_after=new_balance,
                account_id=account_id,
                ledger_id=account.ledger_id,
                description=f'调整账户余额: ￥{old_balance:.2f} → ￥{new_balance:.2f}'
            )
        return jsonify({
            "success": True,
            "message": "账户更新成功",
            "account": {
                "id": account.id,
                "name": account.name,
                "balance": float(account.balance),
                "account_type": account.account_type
            }
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新账户失败: {e}")
        return jsonify({"success": False, "message": "更新账户失败"}), 500


@app.route('/api/accounts/<int:account_id>', methods=['DELETE'])
@login_required
def delete_account(account_id):
    try:
        # 管理员在 self_view 模式下可以删除账户
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({"success": False, "message": "管理员只能查看数据，不能删除账户"}), 403

        user_id = session.get('user_id')
        account = Account.query.filter_by(id=account_id, user_id=user_id).first()
        if not account:
            return jsonify({"success": False, "message": "账户不存在"}), 404

        if account.transactions.count() > 0:
            return jsonify({"success": False, "message": "该账户存在交易记录，无法删除"}), 400

        db.session.delete(account)
        db.session.commit()
        return jsonify({"success": True, "message": "账户删除成功"})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除账户失败: {e}")
        return jsonify({"success": False, "message": "删除账户失败"}), 500

@app.route('/api/reports/<period>', methods=['GET'])
@login_required
def get_report(period):
    try:
        transactions = filter_transactions_by_period_orm(period).all()
        income = float(sum(t.amount for t in transactions if t.type == 'income'))
        expense = float(sum(t.amount for t in transactions if t.type == 'expense'))
        return jsonify({
            "success": True,
            "period": period,
            "transactions": [
                {
                    "id": t.id,
                    "type": t.type,
                    "amount": float(t.amount) if t.amount else 0,
                    "category": t.category,
                    "date": t.date,
                    "time": t.time,
                    "remark": t.remark,
                    "account_id": t.account_id,
                    "currency": t.currency or 'CNY',
                    "original_amount": float(t.original_amount) if t.original_amount else None,
                    "exchange_rate": float(t.exchange_rate) if t.exchange_rate else None,
                    "created_at": t.created_at.isoformat() if t.created_at else None,
                    "reimbursement_status": t.reimbursement_status or 'none',
                    "reimbursed_amount": float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                    "write_off_id": t.write_off_id
                } for t in transactions
            ],
            "income": income,
            "expense": expense,
            "net": income - expense,
            "count": len(transactions)
        })
    except Exception as e:
        app.logger.error(f"获取报表失败: {e}")
        abort(500, description="获取报表失败")


@app.route('/api/reports/advanced', methods=['GET'])
@login_required
def get_advanced_report():
    try:
        period = request.args.get('period', 'month')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        now = datetime.now()

        # 构建日期范围
        if period == 'week':
            start = (now - timedelta(days=7)).strftime('%Y-%m-%d')
            end = now.strftime('%Y-%m-%d')
        elif period == 'month':
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')
        elif period == 'quarter':
            current_quarter = (now.month - 1) // 3 + 1
            start_month = (current_quarter - 1) * 3 + 1
            start = f'{now.year}-{start_month:02d}-01'
            if current_quarter == 1:
                end_month = 3
            elif current_quarter == 2:
                end_month = 6
            elif current_quarter == 3:
                end_month = 9
            else:
                end_month = 12
            last_day = calendar.monthrange(now.year, end_month)[1]
            end = f'{now.year}-{end_month:02d}-{last_day}'
        elif period == 'year':
            start = f'{now.year}-01-01'
            end = now.strftime('%Y-%m-%d')
        elif period == 'custom':
            start = start_date if start_date else (now - timedelta(days=30)).strftime('%Y-%m-%d')
            end = end_date if end_date else now.strftime('%Y-%m-%d')
        else:
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')

        # 查询数据（带权限过滤）
        query = Transaction.query.filter(
            Transaction.date >= start,
            Transaction.date <= end
        )

        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            query = query.filter(Transaction.ledger_id == current_ledger_id)
        elif user_id and not is_admin or (is_admin and self_view):
            # 普通用户或管理员 self_view 只看到自己的数据
            query = query.filter(Transaction.user_id == user_id)

        transactions = query.order_by(Transaction.date.desc(), Transaction.id.desc()).all()
        income = float(sum(t.amount for t in transactions if t.type == 'income'))
        expense = float(sum(t.amount for t in transactions if t.type == 'expense'))
        count = len(transactions)

        # 分类占比统计（分别统计收入和支出）
        category_stats = {}
        for t in transactions:
            key = f"{t.type}:{t.category}"
            if key not in category_stats:
                category_stats[key] = {'type': t.type, 'category': t.category, 'amount': 0, 'count': 0}
            category_stats[key]['amount'] += float(t.amount)
            category_stats[key]['count'] += 1

        # 分别提取收入和支出分类统计
        expense_category_stats = [v for v in category_stats.values() if v['type'] == 'expense']
        income_category_stats = [v for v in category_stats.values() if v['type'] == 'income']

        # 查找支出/收入最多的分类
        expense_category_stats_sorted = sorted(expense_category_stats, key=lambda x: x['amount'], reverse=True)
        income_category_stats_sorted = sorted(income_category_stats, key=lambda x: x['amount'], reverse=True)

        # 计算天数
        date_diff = max(1, (datetime.strptime(end, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')).days + 1)

        # 每日趋势（按日期汇总）
        daily_trend = {}
        for t in transactions:
            if t.date not in daily_trend:
                daily_trend[t.date] = {'income': 0, 'expense': 0, 'count': 0}
            if t.type == 'income':
                daily_trend[t.date]['income'] += float(t.amount)
            else:
                daily_trend[t.date]['expense'] += float(t.amount)
            daily_trend[t.date]['count'] += 1

        return jsonify({
            "success": True,
            "period": period,
            "start_date": start,
            "end_date": end,
            "summary": {
                "income": income,
                "expense": expense,
                "net": income - expense,
                "count": count,
                "avg_daily_expense": round(expense / date_diff, 2) if expense > 0 else 0,
                "avg_daily_income": round(income / date_diff, 2) if income > 0 else 0,
                "days": date_diff
            },
            "top_expense_category": expense_category_stats_sorted[0] if expense_category_stats_sorted else None,
            "top_income_category": income_category_stats_sorted[0] if income_category_stats_sorted else None,
            "category_stats": list(category_stats.values()),
            "income_category_stats": income_category_stats,
            "expense_category_stats": expense_category_stats,
            "daily_trend": [
                {"date": d, "income": v['income'], "expense": v['expense'], "count": v['count']}
                for d, v in sorted(daily_trend.items())
            ],
            "transactions": [
                {
                    "id": t.id,
                    "type": t.type,
                    "amount": float(t.amount) if t.amount else 0,
                    "category": t.category,
                    "date": t.date,
                    "time": t.time,
                    "remark": t.remark,
                    "account_id": t.account_id,
                    "currency": t.currency or 'CNY',
                    "original_amount": float(t.original_amount) if t.original_amount else None,
                    "exchange_rate": float(t.exchange_rate) if t.exchange_rate else None,
                    "reimbursement_status": t.reimbursement_status or 'none',
                    "reimbursed_amount": float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                    "write_off_id": t.write_off_id
                } for t in transactions[:50]  # 最多返回50条明细
            ]
        })
    except Exception as e:
        app.logger.error(f"获取高级报表失败: {e}")
        return jsonify({"success": False, "message": "获取报表失败"}), 500


@app.route('/api/reports/download', methods=['GET'])
@login_required
def download_report():
    """从报表页面导出统计汇总数据（非明细流水）"""
    try:
        period = request.args.get('period', 'month')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        file_format = request.args.get('format', 'csv').lower()

        if file_format not in ('csv', 'xlsx'):
            return jsonify({"success": False, "message": "不支持的导出格式"}), 400

        now = datetime.now()

        # 构建日期范围（与 advanced report 一致）
        if period == 'week':
            start = (now - timedelta(days=7)).strftime('%Y-%m-%d')
            end = now.strftime('%Y-%m-%d')
        elif period == 'month':
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')
        elif period == 'quarter':
            current_quarter = (now.month - 1) // 3 + 1
            start_month = (current_quarter - 1) * 3 + 1
            start = f'{now.year}-{start_month:02d}-01'
            last_day = calendar.monthrange(now.year, start_month + 2)[1]
            end = f'{now.year}-{start_month + 2:02d}-{last_day}'
        elif period == 'year':
            start = f'{now.year}-01-01'
            end = now.strftime('%Y-%m-%d')
        elif period == 'custom':
            start = start_date if start_date else (now - timedelta(days=30)).strftime('%Y-%m-%d')
            end = end_date if end_date else now.strftime('%Y-%m-%d')
        else:
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')

        # 查询数据（带权限过滤）
        query = Transaction.query.filter(
            Transaction.date >= start,
            Transaction.date <= end
        )

        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            query = query.filter(Transaction.ledger_id == current_ledger_id)
        elif user_id and not is_admin or (is_admin and self_view):
            query = query.filter(Transaction.user_id == user_id)

        transactions = query.order_by(Transaction.date.asc(), Transaction.id.asc()).all()

        if not transactions:
            return jsonify({"success": False, "message": "该时间段内没有交易记录"}), 404

        # ---- 计算统计汇总数据 ----
        income = float(sum(t.amount for t in transactions if t.type == 'income'))
        expense = float(sum(t.amount for t in transactions if t.type == 'expense'))
        count = len(transactions)
        date_diff = max(1, (datetime.strptime(end, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')).days + 1)

        # 每日趋势
        daily_trend = {}
        for t in transactions:
            if t.date not in daily_trend:
                daily_trend[t.date] = {'income': 0, 'expense': 0, 'count': 0}
            if t.type == 'income':
                daily_trend[t.date]['income'] += float(t.amount)
            else:
                daily_trend[t.date]['expense'] += float(t.amount)
            daily_trend[t.date]['count'] += 1

        daily_rows = [
            {'date': d, 'income': v['income'], 'expense': v['expense'],
             'net': round(v['income'] - v['expense'], 2), 'count': v['count']}
            for d, v in sorted(daily_trend.items())
        ]

        # 分类统计
        category_stats = {}
        for t in transactions:
            key = f"{t.type}:{t.category}"
            if key not in category_stats:
                category_stats[key] = {'category': t.category, 'type': t.type, 'amount': 0, 'count': 0}
            category_stats[key]['amount'] += float(t.amount)
            category_stats[key]['count'] += 1

        category_rows = sorted(category_stats.values(), key=lambda x: x['amount'], reverse=True)

        timestamp = now.strftime('%Y%m%d_%H%M%S')

        if file_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # 第1段：汇总统计
            writer.writerow(['=== 汇总统计 ==='])
            writer.writerow(['指标', '值'])
            writer.writerow(['统计周期', f'{start} ~ {end}'])
            writer.writerow(['统计天数', date_diff])
            writer.writerow(['总收入', income])
            writer.writerow(['总支出', expense])
            writer.writerow(['净收支', round(income - expense, 2)])
            writer.writerow(['交易总笔数', count])
            writer.writerow(['日均收入', round(income / date_diff, 2)])
            writer.writerow(['日均支出', round(expense / date_diff, 2)])
            writer.writerow([])

            # 第2段：每日趋势
            writer.writerow(['=== 每日收支趋势 ==='])
            writer.writerow(['日期', '收入', '支出', '净收支', '笔数'])
            for row in daily_rows:
                writer.writerow([row['date'], row['income'], row['expense'], row['net'], row['count']])
            writer.writerow([])

            # 第3段：分类统计
            writer.writerow(['=== 分类统计 ==='])
            writer.writerow(['分类', '类型', '金额', '笔数'])
            for row in category_rows:
                type_label = '收入' if row['type'] == 'income' else '支出'
                writer.writerow([row['category'], type_label, round(row['amount'], 2), row['count']])

            csv_bytes = output.getvalue().encode('utf-8-sig')
            return send_file(
                io.BytesIO(csv_bytes),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'报表汇总_{start}_{end}.csv'
            )
        else:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                return jsonify({"success": False, "message": "导出 XLSX 需要安装 openpyxl"}), 500

            wb = openpyxl.Workbook()
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            label_font = Font(bold=True, size=11)

            def style_header(ws, headers):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

            def auto_width(ws):
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

            # Sheet 1: 汇总
            ws1 = wb.active
            ws1.title = '汇总'
            ws1.cell(row=1, column=1, value='指标').font = label_font
            ws1.cell(row=1, column=2, value='值').font = label_font
            ws1.cell(row=1, column=1).border = thin_border
            ws1.cell(row=1, column=2).border = thin_border
            summary_data = [
                ('统计周期', f'{start} ~ {end}'),
                ('统计天数', date_diff),
                ('总收入', income),
                ('总支出', expense),
                ('净收支', round(income - expense, 2)),
                ('交易总笔数', count),
                ('日均收入', round(income / date_diff, 2)),
                ('日均支出', round(expense / date_diff, 2)),
            ]
            for i, (k, v) in enumerate(summary_data, 2):
                ws1.cell(row=i, column=1, value=k).border = thin_border
                ws1.cell(row=i, column=2, value=v).border = thin_border
            ws1.column_dimensions['A'].width = 14
            ws1.column_dimensions['B'].width = 30

            # Sheet 2: 每日趋势
            ws2 = wb.create_sheet('每日趋势')
            style_header(ws2, ['日期', '收入', '支出', '净收支', '笔数'])
            for i, row in enumerate(daily_rows, 2):
                for j, val in enumerate([row['date'], row['income'], row['expense'], row['net'], row['count']], 1):
                    cell = ws2.cell(row=i, column=j, value=val)
                    cell.border = thin_border
            auto_width(ws2)
            ws2.freeze_panes = 'A2'

            # Sheet 3: 分类统计
            ws3 = wb.create_sheet('分类统计')
            style_header(ws3, ['分类', '类型', '金额', '笔数'])
            for i, row in enumerate(category_rows, 2):
                type_label = '收入' if row['type'] == 'income' else '支出'
                for j, val in enumerate([row['category'], type_label, round(row['amount'], 2), row['count']], 1):
                    cell = ws3.cell(row=i, column=j, value=val)
                    cell.border = thin_border
            auto_width(ws3)
            ws3.freeze_panes = 'A2'

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(tmp.name)
            tmp.close()

            return send_file(
                tmp.name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'报表汇总_{start}_{end}.xlsx'
            )
    except Exception as e:
        app.logger.error(f"导出报表失败: {e}")
        return jsonify({"success": False, "message": "导出失败"}), 500


# ==================== 资金变动记录 API ====================
@app.route('/api/money-change-logs', methods=['GET'])
@login_required
def get_money_change_logs():
    """获取资金变动记录"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        action_type = request.args.get('action_type')
        entity_type = request.args.get('entity_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        keyword = request.args.get('keyword')

        query = MoneyChangeLog.query.options(
            db.joinedload(MoneyChangeLog.user),
            db.joinedload(MoneyChangeLog.account)
        )

        # 权限过滤
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()

        if is_admin and not self_view:
            pass  # 管理员后台查看全部
        elif current_ledger_id:
            query = query.filter(MoneyChangeLog.ledger_id == current_ledger_id)
        else:
            query = query.filter(MoneyChangeLog.user_id == user_id)

        if action_type:
            query = query.filter(MoneyChangeLog.action_type == action_type)
        if entity_type:
            query = query.filter(MoneyChangeLog.entity_type == entity_type)
        if start_date:
            query = query.filter(MoneyChangeLog.created_at >= start_date)
        if end_date:
            query = query.filter(MoneyChangeLog.created_at <= end_date + ' 23:59:59')
        if keyword:
            query = query.filter(MoneyChangeLog.description.like(f'%{keyword}%'))

        query = query.order_by(MoneyChangeLog.created_at.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        logs = pagination.items

        # 为 transaction 类型日志关联当前交易数据
        logs_data = []
        for log in logs:
            entry = {
                "id": log.id,
                "user_id": log.user_id,
                "username": log.user.username if log.user else None,
                "ledger_id": log.ledger_id,
                "account_id": log.account_id,
                "account_name": log.account.name if log.account else None,
                "action_type": log.action_type,
                "entity_type": log.entity_type,
                "entity_id": log.entity_id,
                "amount_change": float(log.amount_change) if log.amount_change else 0,
                "balance_before": float(log.balance_before) if log.balance_before else None,
                "balance_after": float(log.balance_after) if log.balance_after else None,
                "description": log.description,
                "created_at": log.created_at.isoformat() if log.created_at else None,
                "transaction_data": None,
                "transaction_exists": False
            }
            if log.entity_type == 'transaction' and log.entity_id:
                tx = Transaction.query.get(log.entity_id)
                if tx:
                    entry['transaction_data'] = {
                        "type": tx.type,
                        "amount": float(tx.amount),
                        "category": tx.category,
                        "date": tx.date,
                        "time": tx.time,
                        "remark": tx.remark,
                        "account_name": tx.account.name if tx.account else None,
                        "currency": tx.currency or 'CNY',
                        "original_amount": float(tx.original_amount) if tx.original_amount else None,
                        "reimbursement_status": tx.reimbursement_status or 'none',
                        "reimbursed_amount": float(tx.reimbursed_amount) if tx.reimbursed_amount else 0,
                        "write_off_id": tx.write_off_id,
                        "split_details": json.loads(tx.split_details) if tx.split_details else []
                    }
                    entry['transaction_exists'] = True
            logs_data.append(entry)

        return jsonify({
            "success": True,
            "logs": logs_data,
            "total": pagination.total,
            "page": page,
            "per_page": per_page,
            "total_pages": pagination.pages
        })
    except Exception as e:
        app.logger.error(f"获取资金变动记录失败: {e}")
        return jsonify({"success": False, "message": "获取记录失败"}), 500


# ==================== 账本管理 API ====================
import secrets



