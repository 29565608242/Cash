import csv
import io
import json
import os
import re
import threading
import traceback
import uuid
from datetime import datetime, timedelta
from functools import wraps

import chardet
from flask import (Blueprint, current_app, g, jsonify, redirect, request,
                   render_template, send_file, session, url_for)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import get_config

config = get_config()


def _import_models():
    """延迟导入避免循环导入"""
    import app as _app
    return _app.db, _app.Transaction, _app.Category, _app.Account, _app.ExportTask, _app.FileUpload


def _login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': '登录已过期，请刷新页面'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def _smtp_configured():
    """检查 SMTP 是否已配置"""
    return bool(current_app.config.get('SMTP_SERVER', config.SMTP_SERVER)
                and current_app.config.get('SMTP_USER', config.SMTP_USER))

# 蓝图定义
import_export_bp = Blueprint('import_export', __name__,
                             template_folder='../templates',
                             static_folder='../static')


# ==================== 字段映射常量 ====================

FIELD_ALIASES = {
    'type': ['type', '类型', '类别', '交易类型', '收/支', '收支', '收/支类型'],
    'amount': ['amount', '金额', '交易金额', '金额(元)', 'money', '数额', '收入金额', '支出金额'],
    'category': ['category', '分类', '类别名称', '类目', '科目', 'tag', '类别'],
    'date': ['date', '日期', '交易日期', '记账日期', 'day', '时间'],
    'time': ['time', '时间', '交易时间', 'hour', '时分'],
    'remark': ['remark', '备注', '说明', '描述', '摘要', '注释', 'note', 'description', '备注说明'],
    'currency': ['currency', '币种', '货币', '货币类型', '外币类型'],
    'original_amount': ['original_amount', '原币金额', '外币金额', '原始金额', '原金额'],
}

REQUIRED_FIELDS = ['type', 'amount', 'category', 'date']

TYPE_MAPPING = {
    'income': ['income', '收入', '收', '+', '入账', '流入'],
    'expense': ['expense', '支出', '支', '-', '出账', '流出', '消费', '花费'],
}


# ==================== 页面路由 ====================

@import_export_bp.route('/data')
@_login_required
def data_tools_page():
    """数据工具页面（导入/导出）"""
    from app import Account  # 避免循环导入
    accounts = Account.query.filter_by(user_id=g.user.id).all()
    return render_template('data_tools.html', accounts=accounts)


# ==================== 导入 API ====================

@import_export_bp.route('/api/import/upload', methods=['POST'])
@_login_required
def upload_import_file():
    """上传并解析导入文件"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '文件名为空'}), 400

    # 校验文件扩展名
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.csv', '.xlsx', '.xls'):
        return jsonify({'success': False, 'message': '不支持的文件格式，请上传 CSV 或 Excel 文件（.csv/.xlsx/.xls）'}), 400

    # 保存上传文件
    upload_id = uuid.uuid4().hex[:12]
    upload_dir = current_app.config.get('UPLOAD_DIR', config.UPLOAD_DIR)
    os.makedirs(upload_dir, exist_ok=True)
    save_path = os.path.join(upload_dir, f'{upload_id}{ext}')
    file.save(save_path)

    try:
        # 解析文件
        if ext == '.csv':
            rows, columns = _parse_csv(save_path)
        else:
            rows, columns = _parse_excel(save_path)

        if not rows:
            return jsonify({'success': False, 'message': '文件为空或无法解析'}), 400

        # 自动映射列名
        auto_mapping = _auto_map_columns(columns)

        # 保存到数据库
        preview_rows = rows[:5]
        total_rows = len(rows)

        file_format = ext.lstrip('.')
        upload_record = FileUpload(
            upload_id=upload_id,
            user_id=g.user.id,
            original_filename=file.filename,
            file_path=save_path,
            file_format=file_format,
            total_rows=total_rows,
            columns=json.dumps(columns, ensure_ascii=False),
            preview_data=json.dumps(preview_rows, ensure_ascii=False),
            status='uploaded'
        )
        db.session.add(upload_record)
        db.session.commit()

        return jsonify({
            'success': True,
            'upload_id': upload_id,
            'total_rows': total_rows,
            'columns': columns,
            'sample_rows': preview_rows,
            'auto_mapping': auto_mapping
        })

    except Exception as e:
        app = current_app._get_current_object()
        app.logger.error(f"文件解析失败: {e}\n{traceback.format_exc()}")
        if os.path.exists(save_path):
            os.remove(save_path)
        return jsonify({'success': False, 'message': f'文件解析失败: {str(e)}'}), 400


@import_export_bp.route('/api/import/confirm', methods=['POST'])
@_login_required
def confirm_import():
    """确认导入"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    data = request.get_json()
    upload_id = data.get('upload_id')
    mapping = data.get('mapping', {})  # {field: column_index_or_name}
    skip_errors = data.get('skip_errors', True)
    default_type = data.get('default_type', None)

    upload_record = FileUpload.query.filter_by(upload_id=upload_id, user_id=g.user.id).first()
    if not upload_record:
        return jsonify({'success': False, 'message': '上传记录不存在'}), 404

    if upload_record.status != 'uploaded':
        return jsonify({'success': False, 'message': '该文件已处理，不要重复导入'}), 400

    # 重新读取文件
    ext = os.path.splitext(upload_record.file_path)[1].lower()
    try:
        if ext == '.csv':
            all_rows, _ = _parse_csv(upload_record.file_path)
        else:
            all_rows, _ = _parse_excel(upload_record.file_path)
    except Exception as e:
        return jsonify({'success': False, 'message': f'重新读取文件失败: {str(e)}'}), 400

    # 列名到索引映射
    columns = json.loads(upload_record.columns)
    col_to_idx = {col: idx for idx, col in enumerate(columns)}

    # 解析映射配置
    field_col_map = {}  # {field: column_index}
    for field, col_name in mapping.items():
        if col_name and col_name in col_to_idx:
            field_col_map[field] = col_to_idx[col_name]

    # 检查必需字段
    missing = [f for f in REQUIRED_FIELDS if f not in field_col_map]
    if missing:
        field_names_zh = {'type': '类型', 'amount': '金额', 'category': '分类', 'date': '日期'}
        missing_zh = [field_names_zh.get(f, f) for f in missing]
        return jsonify({
            'success': False,
            'message': f'缺少必需字段映射: {", ".join(missing_zh)}'
        }), 400

    # 批量导入
    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(all_rows):
        try:
            tx_data = {}
            for field, col_idx in field_col_map.items():
                if col_idx < len(row):
                    tx_data[field] = row[col_idx].strip() if isinstance(row[col_idx], str) else row[col_idx]
                else:
                    tx_data[field] = ''

            # 处理类型字段
            raw_type = str(tx_data.get('type', '')).strip().lower()
            mapped_type = _map_type(raw_type)
            if not mapped_type:
                if default_type:
                    mapped_type = default_type
                else:
                    raise ValueError(f"无法识别交易类型: '{raw_type}'，请使用 收入/支出 或 income/expense")
            tx_type = mapped_type

            # 处理金额
            raw_amount = tx_data.get('amount', 0)
            amount = _parse_amount(raw_amount)
            if amount is None or amount <= 0:
                raise ValueError(f"无效的金额: '{raw_amount}'")
            if amount > 99999999.99:
                raise ValueError("金额超出限制")

            # 处理分类
            category_name = str(tx_data.get('category', '')).strip()
            if not category_name:
                raise ValueError("分类不能为空")

            # 查找或创建分类
            category = Category.query.filter_by(name=category_name, type=tx_type).first()
            if not category:
                category = Category(name=category_name, type=tx_type)
                db.session.add(category)
                db.session.flush()

            # 处理日期
            tx_date = _parse_date(tx_data.get('date', ''))
            if not tx_date:
                tx_date = datetime.now().strftime('%Y-%m-%d')

            # 处理时间
            tx_time = str(tx_data.get('time', '')).strip()
            if not tx_time:
                tx_time = datetime.now().strftime('%H:%M:%S')

            # 处理币种
            currency = str(tx_data.get('currency', 'CNY')).strip().upper()
            if not currency:
                currency = 'CNY'

            # 处理原币金额
            original_amount = None
            exchange_rate = None
            if currency != 'CNY':
                raw_original = tx_data.get('original_amount', None)
                if raw_original:
                    original_amount = _parse_amount(raw_original)
                    exchange_rate = round(amount / original_amount, 6) if original_amount else None

            # 创建交易记录
            from app import get_current_ledger_id
            from cash_app.support import log_money_change
            current_ledger_id = get_current_ledger_id()
            # 获取或创建账户
            account = Account.query.filter_by(ledger_id=current_ledger_id).first()
            if not account:
                account = Account.query.filter_by(user_id=g.user.id).first()
            if not account:
                account = Account(name='默认账户', balance=0, account_type='cash', user_id=g.user.id, ledger_id=current_ledger_id)
                db.session.add(account)
                db.session.flush()
            transaction = Transaction(
                type=tx_type,
                amount=amount,
                category=category_name,
                date=tx_date,
                time=tx_time,
                remark=str(tx_data.get('remark', '')).strip(),
                user_id=g.user.id,
                ledger_id=current_ledger_id,
                account_id=account.id,
                currency=currency,
                original_amount=original_amount,
                exchange_rate=exchange_rate
            )
            db.session.add(transaction)
            # 更新账户余额
            if tx_type == 'income':
                account.balance = float(account.balance) + amount
                imp_amt = amount
            else:
                account.balance = float(account.balance) - amount
                imp_amt = -amount
            log_money_change(
                user_id=g.user.id,
                action_type='import',
                entity_type='transaction',
                amount_change=imp_amt,
                balance_before=float(account.balance) - imp_amt,
                balance_after=float(account.balance),
                account_id=account.id,
                ledger_id=current_ledger_id,
                description=f'导入创建{"收入" if tx_type == "income" else "支出"} ￥{amount:.2f} - {category_name}'
            )
            imported += 1

        except Exception as e:
            if skip_errors:
                skipped += 1
                errors.append({'row': row_idx + 2, 'message': str(e)})
                continue
            else:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'message': f'第 {row_idx + 2} 行导入失败: {str(e)}',
                    'imported': imported,
                    'errors': errors + [{'row': row_idx + 2, 'message': str(e)}]
                }), 422

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'数据库保存失败: {str(e)}'}), 500

    # 标记上传记录已处理
    upload_record.status = 'confirmed'
    db.session.commit()

    # 删除上传文件
    try:
        if os.path.exists(upload_record.file_path):
            os.remove(upload_record.file_path)
    except Exception:
        pass

    return jsonify({
        'success': True,
        'imported': imported,
        'skipped': skipped,
        'errors': errors[:50],  # 最多返回前50个错误
        'total': imported + skipped
    })


# ==================== 导出 API ====================

@import_export_bp.route('/api/export/create', methods=['POST'])
@_login_required
def create_export():
    """创建导出任务"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    data = request.get_json() or {}
    start_date = data.get('start_date', '')
    end_date = data.get('end_date', '')
    raw_account_id = data.get('account_id')
    account_id = int(raw_account_id) if raw_account_id is not None else None
    file_format = data.get('format', 'xlsx')
    email_to = (data.get('email_to') or '').strip()

    if file_format not in ('csv', 'xlsx'):
        return jsonify({'success': False, 'message': '不支持的导出格式'}), 400

    if email_to and not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email_to):
        return jsonify({'success': False, 'message': '邮箱格式不正确'}), 400

    if email_to and not _smtp_configured():
        return jsonify({'success': False, 'message': '系统未配置 SMTP，无法发送邮件'}), 400

    # 构建查询（按当前账本过滤）
    from app import get_current_ledger_id
    current_ledger_id = get_current_ledger_id()
    query = Transaction.query.filter_by(user_id=g.user.id)
    if current_ledger_id:
        query = query.filter(Transaction.ledger_id == current_ledger_id)
    if start_date:
        query = query.filter(Transaction.date >= start_date)
    if end_date:
        query = query.filter(Transaction.date <= end_date)
    if account_id:
        query = query.filter_by(account_id=account_id)
    query = query.order_by(Transaction.date.desc(), Transaction.id.desc())

    # 计数
    total = query.count()
    if total == 0:
        return jsonify({'success': False, 'message': '没有符合条件的交易记录'}), 404

    # 保存筛选条件
    filters = json.dumps({
        'start_date': start_date,
        'end_date': end_date,
        'account_id': account_id,
    }, ensure_ascii=False)

    # 创建任务记录
    task = ExportTask(
        user_id=g.user.id,
        status='pending',
        progress=0,
        file_format=file_format,
        filters=filters,
        total_records=total,
        email_to=email_to if email_to else None,
    )
    db.session.add(task)
    db.session.commit()

    # 判断是否异步
    threshold = current_app.config.get('EXPORT_ASYNC_THRESHOLD', config.EXPORT_ASYNC_THRESHOLD)
    is_async = total >= threshold or bool(email_to)

    if is_async:
        # 异步：后台线程处理
        app = current_app._get_current_object()
        thread = threading.Thread(target=_run_export_background, args=(task.id, app))
        thread.daemon = True
        thread.start()

        return jsonify({
            'success': True,
            'task_id': task.id,
            'total_records': total,
            'status': 'processing',
            'sync': False,
            'message': f'正在导出 {total} 条记录，请稍候...'
        })
    else:
        # 同步：直接生成
        try:
            transactions = query.all()
            filepath = _generate_export_file(task.id, transactions, file_format)
            filesize = os.path.getsize(filepath) if os.path.exists(filepath) else 0

            task.status = 'completed'
            task.progress = 100
            task.file_path = filepath
            task.file_size = filesize
            task.completed_at = datetime.now()
            db.session.commit()

            return jsonify({
                'success': True,
                'task_id': task.id,
                'total_records': total,
                'status': 'completed',
                'sync': True,
                'download_url': url_for('import_export.download_export', task_id=task.id)
            })
        except Exception as e:
            task.status = 'failed'
            task.error_message = str(e)
            db.session.commit()
            return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500


@import_export_bp.route('/api/export/status/<int:task_id>')
@_login_required
def get_export_status(task_id):
    """查询导出任务状态"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    task = ExportTask.query.filter_by(id=task_id, user_id=g.user.id).first()
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404

    result = {
        'success': True,
        'task_id': task.id,
        'status': task.status,
        'progress': task.progress,
        'total_records': task.total_records,
        'file_format': task.file_format,
        'file_size': task.file_size,
        'created_at': task.created_at.isoformat() if task.created_at else None,
        'completed_at': task.completed_at.isoformat() if task.completed_at else None,
        'error_message': task.error_message,
    }

    if task.status == 'completed' and task.file_path:
        result['download_url'] = url_for('import_export.download_export', task_id=task.id)

    return jsonify(result)


@import_export_bp.route('/api/export/download/<int:task_id>')
@_login_required
def download_export(task_id):
    """下载导出文件"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    task = ExportTask.query.filter_by(id=task_id, user_id=g.user.id).first()
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404
    if task.status != 'completed':
        return jsonify({'success': False, 'message': '导出未完成'}), 400
    if not task.file_path or not os.path.exists(task.file_path):
        return jsonify({'success': False, 'message': '文件不存在或已过期'}), 404

    timestamp = task.created_at.strftime('%Y%m%d_%H%M%S')
    ext = task.file_format
    filename = f'交易记录_{timestamp}.{ext}'

    mimetypes = {
        'csv': 'text/csv',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    }

    return send_file(
        task.file_path,
        as_attachment=True,
        download_name=filename,
        mimetype=mimetypes.get(ext, 'application/octet-stream')
    )


@import_export_bp.route('/api/export/list')
@_login_required
def list_exports():
    """获取用户的导出任务列表（分页）"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 8, type=int)
    per_page = min(per_page, 50)

    pagination = ExportTask.query.filter_by(user_id=g.user.id) \
        .order_by(ExportTask.id.desc()) \
        .paginate(page=page, per_page=per_page, error_out=False)

    tasks = pagination.items

    return jsonify({
        'success': True,
        'tasks': [{
            'id': t.id,
            'status': t.status,
            'progress': t.progress,
            'file_format': t.file_format,
            'total_records': t.total_records,
            'file_size': t.file_size,
            'error_message': t.error_message,
            'created_at': t.created_at.isoformat() if t.created_at else None,
            'completed_at': t.completed_at.isoformat() if t.completed_at else None,
            'email_to': t.email_to,
            'email_sent': t.email_sent,
            'download_url': url_for('import_export.download_export', task_id=t.id)
                if t.status == 'completed' and t.file_path else None,
        } for t in tasks],
        'total': pagination.total,
        'page': pagination.page,
        'per_page': pagination.per_page,
        'pages': pagination.pages,
    })


@import_export_bp.route('/api/export/delete/<int:task_id>', methods=['POST'])
@_login_required
def delete_export(task_id):
    """删除导出任务记录"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    task = ExportTask.query.filter_by(id=task_id, user_id=g.user.id).first()
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404

    # 删除关联的文件
    if task.file_path and os.path.exists(task.file_path):
        try:
            os.remove(task.file_path)
        except Exception:
            pass

    db.session.delete(task)
    db.session.commit()
    return jsonify({'success': True, 'message': '已删除'})


@import_export_bp.route('/api/export/email', methods=['POST'])
@_login_required
def send_export_email():
    """通过邮件发送导出文件"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    data = request.get_json() or {}
    task_id = data.get('task_id')
    email_to = (data.get('email_to') or '').strip()

    if not email_to or not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email_to):
        return jsonify({'success': False, 'message': '邮箱格式不正确'}), 400

    if not _smtp_configured():
        return jsonify({'success': False, 'message': '系统未配置 SMTP，无法发送邮件'}), 400

    task = ExportTask.query.filter_by(id=task_id, user_id=g.user.id).first()
    if not task:
        return jsonify({'success': False, 'message': '任务不存在'}), 404

    if task.status == 'completed':
        # 立即发送
        try:
            _send_email(task, email_to)
            return jsonify({'success': True, 'message': '邮件发送成功'})
        except Exception as e:
            return jsonify({'success': False, 'message': f'邮件发送失败: {str(e)}'}), 500
    elif task.status in ('pending', 'processing'):
        # 标记为完成后发送
        task.email_to = email_to
        db.session.commit()
        return jsonify({'success': True, 'message': '导出完成后将自动发送邮件'})
    else:
        return jsonify({'success': False, 'message': f'任务状态异常: {task.status}'}), 400


# ==================== 文件解析函数 ====================

def _parse_csv(filepath):
    """解析 CSV 文件，返回 (rows, columns)"""
    # 检测编码
    with open(filepath, 'rb') as f:
        raw = f.read()
    detected = chardet.detect(raw)
    encoding = detected.get('encoding', 'utf-8') or 'utf-8'
    # 常见中文编码兜底
    if encoding.lower() in ('ascii', 'iso-8859-1'):
        encoding = 'utf-8'

    # 解码并去除 BOM
    text = raw.decode(encoding, errors='replace')
    if text.startswith('﻿'):
        text = text[1:]

    # 嗅探分隔符
    try:
        dialect = csv.Sniffer().sniff(text[:4096])
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ','  # 默认逗号分隔

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return [], []

    # 第一行为表头
    headers = rows[0]
    data_rows = rows[1:]

    return data_rows, headers


def _parse_excel(filepath):
    """解析 Excel 文件，返回 (rows, columns)"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append([str(v).strip() if v is not None else '' for v in row])

    return data_rows, headers


def _auto_map_columns(columns):
    """自动映射列名到系统字段"""
    mapping = {}
    for field, aliases in FIELD_ALIASES.items():
        for col in columns:
            col_clean = col.strip().lower().replace(' ', '').replace('　', '')
            for alias in aliases:
                alias_clean = alias.lower().replace(' ', '').replace('　', '')
                if col_clean == alias_clean or alias_clean in col_clean or col_clean in alias_clean:
                    mapping[field] = col
                    break
            if field in mapping:
                break
    return mapping


# ==================== 导出文件生成函数 ====================

def _generate_export_file(task_id, transactions, file_format):
    """生成导出文件，返回文件路径"""
    export_dir = current_app.config.get('EXPORT_DIR', config.EXPORT_DIR)
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, f'export_{task_id}.{file_format}')

    if file_format == 'csv':
        _generate_csv(transactions, filepath)
    else:
        _generate_xlsx(transactions, filepath)

    return filepath


def _generate_csv(transactions, filepath):
    """生成 CSV 导出文件（UTF-8 BOM）"""
    headers = ['ID', '类型', '金额', '分类', '日期', '时间', '备注',
               '账户', '币种', '原币金额', '汇率', '创建时间',
               '报销状态', '已报销金额', '核销ID']

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for t in transactions:
            status_map = {'none': '', 'pending': '待报销', 'partial': '部分报销', 'reimbursed': '已报销'}
            writer.writerow([
                t.id,
                '收入' if t.type == 'income' else '支出',
                float(t.amount) if t.amount else 0,
                t.category,
                t.date,
                t.time,
                t.remark or '',
                t.account.name if t.account else '',
                t.currency or 'CNY',
                float(t.original_amount) if t.original_amount else '',
                float(t.exchange_rate) if t.exchange_rate else '',
                t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else '',
                status_map.get(t.reimbursement_status or 'none', ''),
                float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                t.write_off_id or '',
            ])


def _generate_xlsx(transactions, filepath):
    """生成 Excel 导出文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '交易记录'

    headers = ['ID', '类型', '金额', '分类', '日期', '时间', '备注',
               '账户', '币种', '原币金额', '汇率', '创建时间',
               '报销状态', '已报销金额', '核销ID']

    # 表头样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 数据行
    for row_idx, t in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=t.id)
        ws.cell(row=row_idx, column=2, value='收入' if t.type == 'income' else '支出')
        ws.cell(row=row_idx, column=3, value=float(t.amount) if t.amount else 0)
        ws.cell(row=row_idx, column=4, value=t.category)
        ws.cell(row=row_idx, column=5, value=t.date)
        ws.cell(row=row_idx, column=6, value=t.time)
        ws.cell(row=row_idx, column=7, value=t.remark or '')
        ws.cell(row=row_idx, column=8, value=t.account.name if t.account else '')
        ws.cell(row=row_idx, column=9, value=t.currency or 'CNY')
        ws.cell(row=row_idx, column=10, value=float(t.original_amount) if t.original_amount else '')
        ws.cell(row=row_idx, column=11, value=float(t.exchange_rate) if t.exchange_rate else '')
        ws.cell(row=row_idx, column=12, value=t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else '')
        status_map = {'none': '', 'pending': '待报销', 'partial': '部分报销', 'reimbursed': '已报销'}
        ws.cell(row=row_idx, column=13, value=status_map.get(t.reimbursement_status or 'none', ''))
        ws.cell(row=row_idx, column=14, value=float(t.reimbursed_amount) if t.reimbursed_amount else 0)
        ws.cell(row=row_idx, column=15, value=t.write_off_id or '')

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(2, min(len(transactions) + 2, 100)):  # 取前100行估算
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
            max_length = max(max_length, len(cell_value.encode('utf-8')))
        adjusted_width = min(max_length + 4, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = 'A2'

    wb.save(filepath)


# ==================== 异步导出 ====================

def _run_export_background(task_id, app):
    """后台线程执行导出"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    with app.app_context():
        task = db.session.get(ExportTask, task_id)
        if not task:
            return

        try:
            task.status = 'processing'
            db.session.commit()

            filters = json.loads(task.filters)
            query = Transaction.query.filter_by(user_id=task.user_id)
            if filters.get('start_date'):
                query = query.filter(Transaction.date >= filters['start_date'])
            if filters.get('end_date'):
                query = query.filter(Transaction.date <= filters['end_date'])
            if filters.get('account_id'):
                query = query.filter_by(account_id=filters['account_id'])
            query = query.order_by(Transaction.date.desc(), Transaction.id.desc())

            transactions = query.all()
            task.total_records = len(transactions)
            task.progress = 50
            db.session.commit()

            filepath = _generate_export_file(task_id, transactions, task.file_format)
            filesize = os.path.getsize(filepath) if os.path.exists(filepath) else 0

            task.status = 'completed'
            task.progress = 100
            task.file_path = filepath
            task.file_size = filesize
            task.completed_at = datetime.now()
            db.session.commit()

            # 发送邮件
            if task.email_to:
                try:
                    _send_email(task, task.email_to)
                except Exception as e:
                    app.logger.error(f"导出邮件发送失败 (task {task_id}): {e}")

        except Exception as e:
            task.status = 'failed'
            task.error_message = str(e)
            db.session.commit()
            app.logger.error(f"导出任务 {task_id} 失败: {e}\n{traceback.format_exc()}")


# ==================== 邮件发送 ====================

def _send_email(task, email_to):
    """通过 SMTP 发送导出文件"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    smtp_server = current_app.config.get('SMTP_SERVER', config.SMTP_SERVER)
    smtp_port = current_app.config.get('SMTP_PORT', config.SMTP_PORT)
    smtp_user = current_app.config.get('SMTP_USER', config.SMTP_USER)
    smtp_password = current_app.config.get('SMTP_PASSWORD', config.SMTP_PASSWORD)
    smtp_from = current_app.config.get('SMTP_FROM', config.SMTP_FROM)

    if not smtp_server or not smtp_user:
        raise ValueError('SMTP 未配置，无法发送邮件')

    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    import smtplib

    msg = MIMEMultipart()
    msg['Subject'] = f'交易记录导出 - {task.created_at.strftime("%Y-%m-%d")}'
    msg['From'] = smtp_from or smtp_user
    msg['To'] = email_to

    body = MIMEText(
        f'您好，\n\n您导出的交易记录已生成，请查收附件。\n'
        f'共 {task.total_records} 条记录。\n'
        f'文件格式: {task.file_format.upper()}\n\n'
        f'—— 线上记账系统',
        'plain', 'utf-8'
    )
    msg.attach(body)

    if task.file_path and os.path.exists(task.file_path):
        filename = f'transactions_{task.created_at.strftime("%Y%m%d")}.{task.file_format}'
        with open(task.file_path, 'rb') as f:
            attachment = MIMEBase('application', 'octet-stream')
            attachment.set_payload(f.read())
            encoders.encode_base64(attachment)
            attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(attachment)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        if current_app.config.get('SMTP_USE_TLS', config.SMTP_USE_TLS):
            server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)

    task.email_sent = True
    task.email_to = email_to
    db.session.commit()


# ==================== 数据清洗与转换 ====================

def _map_type(raw_type):
    """将类型文本映射为 income/expense"""
    raw = raw_type.strip().lower()
    for mapped_type, aliases in TYPE_MAPPING.items():
        if raw in aliases:
            return mapped_type
    # 尝试部分匹配
    if '收' in raw or 'in' in raw or '+' in raw:
        return 'income'
    if '支' in raw or 'ex' in raw or '-' in raw or '消' in raw or '花' in raw:
        return 'expense'
    return None


def _parse_amount(raw):
    """解析金额值"""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        # 去除货币符号、空格、逗号
        cleaned = re.sub(r'[¥￥$€£\s,，]', '', raw).strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_date(raw):
    """解析日期字符串为 YYYY-MM-DD 格式"""
    if not raw:
        return None
    if isinstance(raw, (datetime,)):
        return raw.strftime('%Y-%m-%d')
    if isinstance(raw, str):
        raw = raw.strip()
        # 尝试多种格式
        formats = [
            '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d',
            '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S',
            '%Y年%m月%d日',
            '%m/%d/%Y', '%m-%d-%Y',
            '%d/%m/%Y', '%d-%m-%Y',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


# ==================== 文件清理 ====================

def cleanup_old_files():
    """清理过期文件（24小时前）"""
    db, Transaction, Category, Account, ExportTask, FileUpload = _import_models()
    cutoff = datetime.now() - timedelta(hours=24)

    # 清理导出文件
    old_exports = ExportTask.query.filter(
        ExportTask.created_at < cutoff,
        ExportTask.status.in_(['completed', 'failed'])
    ).all()

    for task in old_exports:
        if task.file_path and os.path.exists(task.file_path):
            try:
                os.remove(task.file_path)
            except Exception:
                pass
        db.session.delete(task)

    # 清理已处理或过期的上传文件
    old_uploads = FileUpload.query.filter(
        FileUpload.created_at < cutoff
    ).all()

    for upload in old_uploads:
        if os.path.exists(upload.file_path):
            try:
                os.remove(upload.file_path)
            except Exception:
                pass
        db.session.delete(upload)

    if old_exports or old_uploads:
        db.session.commit()


# 启动时清理一次
try:
    db, _, _, _, _, _ = _import_models()
    if db.session:
        cleanup_old_files()
except Exception:
    pass
