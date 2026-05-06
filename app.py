from flask import Flask, jsonify, request, abort, render_template, send_from_directory, send_file, redirect, url_for, session, g
from datetime import datetime, timedelta
import json
import os
import logging
import calendar
import time
import sys
import csv
import io
import tempfile
from logging.handlers import RotatingFileHandler
from werkzeug.exceptions import HTTPException
from config import get_config
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import requests

# 修复循环导入：直接运行 app.py 时使 import app 返回 __main__ 模块
if __name__ == '__main__':
    sys.modules['app'] = sys.modules['__main__']

# 初始化应用
app = Flask(__name__, static_folder='static', template_folder='templates')
config = get_config()
config.init_app()

# 配置应用
app.config.from_object(config)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.jinja_env.auto_reload = True
DATA_FILE = config.DATA_FILE_PATH
app.secret_key = config.SECRET_KEY

# 初始化数据库
db = SQLAlchemy(app)

class Transaction(db.Model):
    __tablename__ = 'transactions'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    type = db.Column(db.String(10), nullable=False)
    amount = db.Column(db.Numeric(10, 2), nullable=False)
    category = db.Column(db.String(50), nullable=False)
    date = db.Column(db.String(20), nullable=False)
    time = db.Column(db.String(10), nullable=False)
    remark = db.Column(db.String(255))
    account_id = db.Column(db.Integer, db.ForeignKey('accounts.id'), nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    currency = db.Column(db.String(10), default='CNY')
    original_amount = db.Column(db.Numeric(10, 2), nullable=True)
    exchange_rate = db.Column(db.Numeric(10, 6), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, onupdate=datetime.now)

    reimbursement_status = db.Column(db.String(20), default='none')  # none, pending, partial, reimbursed
    reimbursed_amount = db.Column(db.Numeric(10, 2), default=0)
    write_off_id = db.Column(db.Integer, db.ForeignKey('transactions.id'), nullable=True)

    ledger_id = db.Column(db.Integer, db.ForeignKey('ledgers.id'), nullable=True)

    account = db.relationship('Account', backref=db.backref('transactions', lazy='dynamic'))
    user = db.relationship('User', backref=db.backref('transactions', lazy='dynamic'))
    ledger = db.relationship('Ledger', backref=db.backref('transactions', lazy='dynamic'))

class Category(db.Model):
    __tablename__ = 'categories'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(50), nullable=False)
    type = db.Column(db.String(10), nullable=False, default='expense')

class Account(db.Model):
    __tablename__ = 'accounts'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(50), nullable=False)
    balance = db.Column(db.Numeric(10, 2), default=0)
    account_type = db.Column(db.String(20), default='cash')
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    ledger_id = db.Column(db.Integer, db.ForeignKey('ledgers.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, onupdate=datetime.now)

    user = db.relationship('User', backref=db.backref('accounts', lazy='dynamic'))
    ledger = db.relationship('Ledger', backref=db.backref('accounts', lazy='dynamic'))

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username = db.Column(db.String(64), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    avatar = db.Column(db.String(255), nullable=True, default='default_avatar.png')
    email = db.Column(db.String(120), nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    nickname = db.Column(db.String(64), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime, nullable=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Budget(db.Model):
    __tablename__ = 'budgets'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    ledger_id = db.Column(db.Integer, db.ForeignKey('ledgers.id'), nullable=True)
    account_id = db.Column(db.Integer, db.ForeignKey('accounts.id'), nullable=True)
    month = db.Column(db.String(7), nullable=False)
    total_amount = db.Column(db.Numeric(10, 2), nullable=False)
    remark = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, onupdate=datetime.now)

    user = db.relationship('User', backref=db.backref('budgets', lazy='dynamic'))
    ledger = db.relationship('Ledger', backref=db.backref('budgets', lazy='dynamic'))
    account = db.relationship('Account', backref=db.backref('budgets', lazy='dynamic'))


class BudgetCategoryItem(db.Model):
    __tablename__ = 'budget_category_items'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    budget_id = db.Column(db.Integer, db.ForeignKey('budgets.id'), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('categories.id'), nullable=False)
    amount = db.Column(db.Numeric(10, 2), nullable=False)

    budget = db.relationship('Budget', backref=db.backref('category_items', lazy='dynamic'))
    category = db.relationship('Category')


class ExportTask(db.Model):
    __tablename__ = 'export_tasks'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending, processing, completed, failed
    progress = db.Column(db.Integer, default=0)
    file_format = db.Column(db.String(10), nullable=False)  # csv, xlsx
    file_path = db.Column(db.String(500))
    file_size = db.Column(db.Integer)
    filters = db.Column(db.Text)  # JSON string
    total_records = db.Column(db.Integer, default=0)
    error_message = db.Column(db.Text)
    email_to = db.Column(db.String(200))
    email_sent = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    completed_at = db.Column(db.DateTime)

    user = db.relationship('User', backref=db.backref('export_tasks', lazy='dynamic'))


class FileUpload(db.Model):
    __tablename__ = 'file_uploads'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    upload_id = db.Column(db.String(36), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_format = db.Column(db.String(10), nullable=False)
    total_rows = db.Column(db.Integer, default=0)
    columns = db.Column(db.Text)  # JSON
    preview_data = db.Column(db.Text)  # JSON
    status = db.Column(db.String(20), default='uploaded')  # uploaded, confirmed, cancelled
    created_at = db.Column(db.DateTime, default=datetime.now)

    user = db.relationship('User', backref=db.backref('file_uploads', lazy='dynamic'))


class Ledger(db.Model):
    __tablename__ = 'ledgers'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255))
    owner_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    currency = db.Column(db.String(10), default='CNY')
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, onupdate=datetime.now)

    owner = db.relationship('User', backref=db.backref('owned_ledgers', lazy='dynamic'))


class LedgerMember(db.Model):
    __tablename__ = 'ledger_members'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ledger_id = db.Column(db.Integer, db.ForeignKey('ledgers.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    role = db.Column(db.String(20), default='viewer')  # viewer, editor, manager
    joined_at = db.Column(db.DateTime, default=datetime.now)

    ledger = db.relationship('Ledger', backref=db.backref('members', lazy='dynamic'))
    user = db.relationship('User', backref=db.backref('ledger_memberships', lazy='dynamic'))

    __table_args__ = (db.UniqueConstraint('ledger_id', 'user_id', name='uq_ledger_user'),)


class InviteCode(db.Model):
    __tablename__ = 'invite_codes'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    ledger_id = db.Column(db.Integer, db.ForeignKey('ledgers.id'), nullable=False)
    code = db.Column(db.String(64), unique=True, nullable=False, index=True)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    max_uses = db.Column(db.Integer, default=0)  # 0 = unlimited
    used_count = db.Column(db.Integer, default=0)
    expires_at = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    ledger = db.relationship('Ledger', backref=db.backref('invite_codes', lazy='dynamic'))
    creator = db.relationship('User', backref=db.backref('created_invite_codes', lazy='dynamic'))


class Loan(db.Model):
    __tablename__ = 'loans'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    ledger_id = db.Column(db.Integer, db.ForeignKey('ledgers.id'), nullable=True)
    type = db.Column(db.String(10), nullable=False)  # 'borrow' 借入(负债), 'lend' 借出(债权)
    counterparty = db.Column(db.String(100), nullable=False)  # 对方名称
    amount = db.Column(db.Numeric(10, 2), nullable=False)  # 总金额
    repaid_amount = db.Column(db.Numeric(10, 2), default=0)  # 已还/已收金额
    date = db.Column(db.String(20), nullable=False)  # 借贷日期
    due_date = db.Column(db.String(20), nullable=True)  # 到期日
    status = db.Column(db.String(20), default='active')  # 'active', 'settled'
    remark = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, onupdate=datetime.now)

    user = db.relationship('User', backref=db.backref('loans', lazy='dynamic'))
    ledger = db.relationship('Ledger', backref=db.backref('loans', lazy='dynamic'))


class RecurringRule(db.Model):
    """周期账单规则"""
    __tablename__ = 'recurring_rules'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(100), nullable=False)  # 账单名称
    amount = db.Column(db.Numeric(10, 2), nullable=False)  # 金额
    category = db.Column(db.String(50), nullable=False)  # 分类
    type = db.Column(db.String(10), nullable=False)  # income / expense
    period = db.Column(db.String(20), nullable=False)  # daily / weekly / monthly / yearly
    interval_value = db.Column(db.Integer, default=1)  # 间隔值，如每2周、每3个月
    start_date = db.Column(db.String(20), nullable=False)  # 开始日期 YYYY-MM-DD
    end_date = db.Column(db.String(20), nullable=True)  # 结束日期 YYYY-MM-DD
    next_date = db.Column(db.String(20), nullable=False)  # 下次生成日期 YYYY-MM-DD
    is_active = db.Column(db.Boolean, default=True)
    remark = db.Column(db.String(255))
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    ledger_id = db.Column(db.Integer, db.ForeignKey('ledgers.id'), nullable=True)
    account_id = db.Column(db.Integer, db.ForeignKey('accounts.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, onupdate=datetime.now)

    user = db.relationship('User', backref=db.backref('recurring_rules', lazy='dynamic'))
    ledger = db.relationship('Ledger', backref=db.backref('recurring_rules', lazy='dynamic'))
    account = db.relationship('Account', backref=db.backref('recurring_rules', lazy='dynamic'))


# ==================== 日志配置 ====================
def setup_logging():
    """配置日志系统"""
    log_level = getattr(logging, config.LOG_LEVEL.upper(), logging.INFO)
    
    # 创建日志格式
    formatter = logging.Formatter(
        '[%(asctime)s] %(levelname)s in %(module)s: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # 文件处理器（带日志轮转）
    file_handler = RotatingFileHandler(
        config.LOG_FILE,
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5,
        encoding='utf-8'
    )
    file_handler.setFormatter(formatter)
    file_handler.setLevel(log_level)
    
    # 控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    console_handler.setLevel(log_level)
    
    # 配置 Flask 日志
    app.logger.addHandler(file_handler)
    app.logger.addHandler(console_handler)
    app.logger.setLevel(log_level)
    
    app.logger.info(f"{config.APP_NAME} 启动成功")
    app.logger.info(f"数据文件: {DATA_FILE}")
    app.logger.info(f"日志级别: {config.LOG_LEVEL}")


setup_logging()

# 启动时尝试生成已到期的周期账单
try:
    with app.app_context():
        from datetime import date as dt_date
        today_str = dt_date.today().strftime('%Y-%m-%d')
        rules = RecurringRule.query.filter(
            RecurringRule.is_active == True,
            RecurringRule.next_date <= today_str,
            db.or_(RecurringRule.end_date >= today_str, RecurringRule.end_date.is_(None))
        ).all()
        gen_count = 0
        for rule in rules:
            tx = _generate_rule_transaction(rule)
            if tx:
                gen_count += 1
        if gen_count:
            db.session.commit()
            app.logger.info(f"启动时生成周期账单: {gen_count}笔")
except Exception as e:
    app.logger.warning(f"启动时生成周期账单跳过: {e}")


# ==================== 数据管理 ====================
def init_data_file():
    """初始化数据文件"""
    if not os.path.exists(DATA_FILE):
        default_data = {
            "transactions": [],
            "balance": 0,
            "categories": config.DEFAULT_CATEGORIES
        }
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_data, f, ensure_ascii=False, indent=2)
        app.logger.info(f"已创建新的数据文件: {DATA_FILE}")


def load_data():
    """读取数据"""
    try:
        init_data_file()
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        app.logger.error(f"JSON 解析错误: {e}")
        # 备份损坏的文件
        backup_file = f"{DATA_FILE}.error.{datetime.now().strftime('%Y%m%d%H%M%S')}"
        os.rename(DATA_FILE, backup_file)
        app.logger.warning(f"已备份损坏文件到: {backup_file}")
        init_data_file()
        return load_data()
    except Exception as e:
        app.logger.error(f"读取数据失败: {e}")
        raise


def save_data(data):
    """保存数据（带备份）"""
    try:
        # 先保存到临时文件
        temp_file = f"{DATA_FILE}.tmp"
        with open(temp_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        # 验证临时文件可以正常读取
        with open(temp_file, 'r', encoding='utf-8') as f:
            json.load(f)
        
        # 替换原文件
        if os.path.exists(DATA_FILE):
            os.replace(temp_file, DATA_FILE)
        else:
            os.rename(temp_file, DATA_FILE)
        
        app.logger.debug("数据保存成功")
    except Exception as e:
        app.logger.error(f"保存数据失败: {e}")
        if os.path.exists(temp_file):
            os.remove(temp_file)
        raise


# ==================== 汇率服务 ====================
# 默认汇率（API 不可用时兜底）
DEFAULT_EXCHANGE_RATES = {
    'USD': 7.24, 'EUR': 7.87, 'GBP': 9.15, 'JPY': 0.048,
    'HKD': 0.93, 'KRW': 0.0054, 'SGD': 5.38, 'THB': 0.20,
    'TWD': 0.22, 'AUD': 4.82, 'CAD': 5.28, 'MYR': 1.55,
}

# 币种名称映射
CURRENCY_NAMES = {
    'CNY': '人民币', 'USD': '美元', 'EUR': '欧元', 'GBP': '英镑',
    'JPY': '日元', 'HKD': '港币', 'KRW': '韩元', 'SGD': '新加坡元',
    'THB': '泰铢', 'TWD': '台币', 'AUD': '澳元', 'CAD': '加元', 'MYR': '马币',
}

# 币种符号映射
CURRENCY_SYMBOLS = {
    'CNY': '¥', 'USD': '$', 'EUR': '€', 'GBP': '£',
    'JPY': '¥', 'HKD': 'HK$', 'KRW': '₩', 'SGD': 'S$',
    'THB': '฿', 'TWD': 'NT$', 'AUD': 'A$', 'CAD': 'C$', 'MYR': 'RM',
}

# 汇率缓存 {cache_key: (rate, timestamp)}
_exchange_rate_cache = {}

def get_exchange_rate(from_currency, to_currency='CNY'):
    """获取实时汇率，优先缓存，其次 API,最后默认汇率"""
    if from_currency == to_currency:
        return 1.0

    cache_key = f"{from_currency}_{to_currency}"

    # 检查缓存（5分钟有效）
    if cache_key in _exchange_rate_cache:
        rate, ts = _exchange_rate_cache[cache_key]
        if time.time() - ts < 300:
            return rate

    # 调用免费汇率 API
    try:
        url = f"https://open.er-api.com/v6/latest/{from_currency}"
        resp = requests.get(url, timeout=5)
        data = resp.json()
        if data.get('result') == 'success':
            rate = data['rates'].get(to_currency)
            if rate:
                _exchange_rate_cache[cache_key] = (rate, time.time())
                app.logger.info(f"获取实时汇率: 1 {from_currency} = {rate} {to_currency}")
                return rate
    except Exception as e:
        app.logger.warning(f"汇率 API 请求失败: {e}")

    # API 失败，使用默认汇率
    rate = DEFAULT_EXCHANGE_RATES.get(from_currency)
    if rate:
        app.logger.info(f"使用默认汇率: 1 {from_currency} = {rate} {to_currency}")
        return rate

    app.logger.error(f"无法获取汇率: {from_currency} -> {to_currency}")
    return None


# ==================== 错误处理 ====================
@app.errorhandler(Exception)
def handle_exception(e):
    """全局异常处理"""
    # 处理 HTTP 异常
    if isinstance(e, HTTPException):
        response = {
            "success": False,
            "error": e.name,
            "message": e.description,
            "status_code": e.code
        }
        return jsonify(response), e.code
    
    # 处理其他异常
    app.logger.error(f"未处理的异常: {str(e)}", exc_info=True)
    response = {
        "success": False,
        "error": "Internal Server Error",
        "message": "服务器内部错误，请稍后重试",
        "status_code": 500
    }
    return jsonify(response), 500


@app.errorhandler(404)
def not_found(e):
    """404 错误处理"""
    return jsonify({
        "success": False,
        "error": "Not Found",
        "message": "请求的资源不存在",
        "status_code": 404
    }), 404


@app.errorhandler(400)
def bad_request(e):
    """400 错误处理"""
    return jsonify({
        "success": False,
        "error": "Bad Request",
        "message": str(e.description) if e.description else "请求参数错误",
        "status_code": 400
    }), 400


# ==================== 健康检查 ====================
@app.route('/health', methods=['GET'])
def health_check():
    """健康检查端点"""
    try:
        # 检查数据文件是否可读
        data = load_data()
        
        return jsonify({
            "status": "healthy",
            "app_name": config.APP_NAME,
            "timestamp": datetime.now().isoformat(),
            "data_file": os.path.exists(DATA_FILE),
            "transactions_count": len(data.get('transactions', [])),
            "balance": data.get('balance', 0)
        }), 200
    except Exception as e:
        app.logger.error(f"健康检查失败: {e}")
        return jsonify({
            "status": "unhealthy",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }), 503


@app.route('/api/info', methods=['GET'])
def api_info():
    """API 信息端点"""
    return jsonify({
        "app_name": config.APP_NAME,
        "version": "2.0",
        "endpoints": {
            "health": "/health",
            "index": "/",
            "admin": "/admin",
            "transactions": "/api/transactions",
            "categories": "/api/categories",
            "accounts": "/api/accounts",
            "reports": "/api/reports/<period>",
            "reports_advanced": "/api/reports/advanced"
        }
    })


# ==================== 页面路由 ====================
# 登录路由
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role', 'user')
        is_admin = role == 'admin'
        user = User.query.filter_by(username=username, is_admin=is_admin).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['is_admin'] = user.is_admin
            if is_admin:
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('index'))
        error = '管理员账号或密码错误' if is_admin else '用户名或密码错误'
        return render_template('login.html', error=error)
    reset_success = request.args.get('reset_success')
    return render_template('login.html', reset_success=reset_success)

@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    """保留旧路由兼容，重定向到统一登录页"""
    return redirect(url_for('login'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# 注册路由
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        password2 = request.form.get('password2')
        if not username or not password:
            return render_template('register.html', error='用户名和密码必填')
        if password != password2:
            return render_template('register.html', error='两次输入的密码不一致')
        if User.query.filter_by(username=username).first():
            return render_template('register.html', error='用户名已被注册')
        email = request.form.get('email')
        if email and User.query.filter_by(email=email).first():
            return render_template('register.html', error='邮箱已被注册')
        user = User(username=username, email=email, is_admin=False)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        # 注册成功后提示并2秒后自动跳转
        return render_template('register.html', success='注册成功，2秒后自动跳转登录页面~', redirect_login=True)
    return render_template('register.html')

# 登录校验装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# 忘记密码路由
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')

        if not username or not email:
            return render_template('forgot_password.html', error='用户名和邮箱必须填写')

        user = User.query.filter_by(username=username).first()
        if not user:
            return render_template('forgot_password.html', error='用户不存在')

        if not user.email or user.email != email:
            return render_template('forgot_password.html', error='邮箱不匹配，请使用注册时填写的邮箱')

        session['reset_allowed'] = True
        session['reset_username'] = username

        return render_template('reset_password.html', username=username)

    return render_template('forgot_password.html')


# 密码重置路由
@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        username = request.form.get('username')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        # 验证是否通过 forgot-password 页面允许
        if not session.get('reset_allowed') or session.get('reset_username') != username:
            return redirect(url_for('forgot_password'))

        if not username or not new_password or not confirm_password:
            return render_template('reset_password.html', error='所有字段都必须填写')

        if new_password != confirm_password:
            return render_template('reset_password.html', error='新密码和确认密码不一致')

        if len(new_password) < 6:
            return render_template('reset_password.html', error='密码长度不能少于6位')

        user = User.query.filter_by(username=username).first()
        if not user:
            return render_template('reset_password.html', error='用户不存在')

        # 更新密码
        user.set_password(new_password)
        db.session.commit()

        # 清除重置标识
        session.pop('reset_allowed', None)
        session.pop('reset_username', None)

        # 跳转到登录页并提示成功
        return redirect(url_for('login', reset_success=1))

    username = request.args.get('username', '')
    # 没有 session 标识则不让直接访问重置页
    if not session.get('reset_allowed') or session.get('reset_username') != username:
        return redirect(url_for('forgot_password'))
    return render_template('reset_password.html', username=username)

# 用户信息路由
@app.route('/user-profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    user = User.query.get(session['user_id'])

    if request.method == 'POST':
        # 处理头像上传（立即提交，确保头像独立保存）
        avatar_file = request.files.get('avatar')
        if avatar_file and avatar_file.filename:
            import os
            import uuid
            from werkzeug.utils import secure_filename

            # 验证文件类型
            allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
            ext = avatar_file.filename.rsplit('.', 1)[-1].lower() if '.' in avatar_file.filename else ''
            if ext not in allowed_extensions:
                return render_template('user_profile.html', error='仅支持 JPG、PNG、GIF、SVG、WebP 格式的图片', user=user)

            # 确保avatars目录存在
            avatar_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'avatars')
            os.makedirs(avatar_dir, exist_ok=True)

            # 生成唯一文件名，防止冲突
            unique_name = f"{uuid.uuid4().hex}.{ext}"
            file_path = os.path.join(avatar_dir, unique_name)

            # 删除旧头像（如果不是默认头像）
            if user.avatar and user.avatar not in ('default_avatar.svg', 'default_avatar.png'):
                old_path = os.path.join(avatar_dir, user.avatar)
                if os.path.exists(old_path):
                    os.remove(old_path)

            # 保存文件
            avatar_file.save(file_path)
            user.avatar = unique_name
            db.session.commit()  # 立即提交头像变更

        # 更新用户信息
        new_email = request.form.get('email')
        new_phone = request.form.get('phone')
        new_nickname = request.form.get('nickname')

        user.email = new_email
        user.phone = new_phone
        user.nickname = new_nickname

        db.session.commit()

        # 重新查询用户以获取最新数据
        user = User.query.get(session['user_id'])

        return render_template('user_profile.html', success='个人信息更新成功', user=user)

    return render_template('user_profile.html', user=user)

# 密码修改路由（在用户信息模块中）
@app.route('/user-profile/change-password', methods=['GET', 'POST'])
@login_required
def user_profile_change_password():
    user = User.query.get(session['user_id'])

    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        # 验证输入
        if not current_password or not new_password or not confirm_password:
            return render_template('user_profile.html', error='所有字段都必须填写', user=user)

        if new_password != confirm_password:
            return render_template('user_profile.html', error='新密码和确认密码不一致', user=user)

        # 验证当前密码
        if not user.check_password(current_password):
            return render_template('user_profile.html', error='当前密码错误', user=user)

        # 更新密码
        user.set_password(new_password)
        db.session.commit()

        # 清除会话并要求重新登录
        session.clear()
        return render_template('user_profile.html', success='密码修改成功，请重新登录', user=user)

    return render_template('user_profile.html', user=user)
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': '登录已过期，请刷新页面'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_admin(*args, **kwargs):
        if not session.get('user_id') or not session.get('is_admin'):
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': '登录已过期，请刷新页面'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_admin

@app.before_request
def load_logged_in_user():
    user_id = session.get('user_id')
    if user_id is None:
        g.user = None
    else:
        g.user = User.query.get(user_id)


# ==================== 账本权限工具函数 ====================
def get_user_ledger_role(ledger_id, user_id):
    """获取用户在账本中的角色，返回 None 表示非成员"""
    member = LedgerMember.query.filter_by(ledger_id=ledger_id, user_id=user_id).first()
    if member:
        return member.role
    ledger = Ledger.query.get(ledger_id)
    if ledger and ledger.owner_id == user_id:
        return 'manager'
    return None


def require_ledger_access(ledger_id, min_role='viewer'):
    """检查当前用户是否有至少 min_role 的权限
    返回 (has_access, role, error_response)"""
    user_id = session.get('user_id')
    if not user_id:
        return False, None, (jsonify({'success': False, 'message': '未登录'}), 401)
    # 管理员在后台模式(self_view=False)下拥有全部访问权限
    # 但在前台查看共享账本时需遵守账本权限
    if session.get('is_admin') and not session.get('self_view', True):
        return True, 'manager', None
    role = get_user_ledger_role(ledger_id, user_id)
    if not role:
        return False, None, (jsonify({'success': False, 'message': '无权限访问该账本'}), 403)
    role_hierarchy = {'viewer': 0, 'editor': 1, 'manager': 2}
    if role_hierarchy.get(role, -1) < role_hierarchy.get(min_role, 0):
        return False, role, (jsonify({'success': False, 'message': '权限不足'}), 403)
    return True, role, None


def get_current_ledger_id():
    """获取当前活动账本 ID，从 session 读取，回退到用户第一个账本"""
    ledger_id = session.get('active_ledger_id')
    user_id = session.get('user_id')
    if not user_id:
        return None
    if ledger_id:
        member = LedgerMember.query.filter_by(ledger_id=ledger_id, user_id=user_id).first()
        if member:
            return ledger_id
        owned = Ledger.query.filter_by(id=ledger_id, owner_id=user_id).first()
        if owned:
            return ledger_id
    membership = LedgerMember.query.filter_by(user_id=user_id).first()
    if membership:
        session['active_ledger_id'] = membership.ledger_id
        return membership.ledger_id
    owned = Ledger.query.filter_by(owner_id=user_id).first()
    if owned:
        session['active_ledger_id'] = owned.id
        return owned.id
    return None


# 首页、admin页路由权限修改（index和admin_dashboard分别加装饰器）
@app.route('/')
@login_required
def index():
    """首页 - 管理员在此页面仅看自己的数据"""
    session['self_view'] = True
    username = g.user.username if g.user else None
    is_admin = g.user.is_admin if g.user else False
    return render_template('index.html', username=username, is_admin=is_admin, self_view=is_admin)

@app.route('/admin')
@admin_required
def admin_dashboard():
    """后台管理页面 - 管理员查看所有用户的数据"""
    session['self_view'] = False
    try:
        # 所有用户
        users = User.query.all()
        user_list = [{'id': u.id, 'username': u.username, 'is_admin': u.is_admin} for u in users]

        # 所有交易统计
        total_income = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='income').scalar() or 0
        total_expense = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='expense').scalar() or 0
        total_balance = float(total_income - total_expense)

        # 今日数据
        today = datetime.now().strftime('%Y-%m-%d')
        today_transactions = Transaction.query.filter(Transaction.date == today).all()
        today_income = sum(float(t.amount) for t in today_transactions if t.type == 'income')
        today_expense = sum(float(t.amount) for t in today_transactions if t.type == 'expense')

        # 每个用户的统计数据
        user_stats = []
        for u in users:
            u_income = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='income', user_id=u.id).scalar() or 0
            u_expense = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='expense', user_id=u.id).scalar() or 0
            u_count = Transaction.query.filter_by(user_id=u.id).count()
            user_stats.append({
                'username': u.username,
                'income': float(u_income),
                'expense': float(u_expense),
                'balance': float(u_income - u_expense),
                'count': u_count
            })

        return render_template('admin.html',
            balance=total_balance,
            today_income=float(today_income),
            today_expense=float(today_expense),
            total_income=float(total_income),
            total_expense=float(total_expense),
            user_stats=user_stats,
            users=user_list
        )

@app.route('/create-admin', methods=['GET', 'POST'])
def create_admin():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            return render_template('create_admin.html', error='用户名和密码不能为空')

        if User.query.filter_by(username=username).first():
            return render_template('create_admin.html', error='用户名已存在')

        admin = User(username=username, password_hash=generate_password_hash(password), is_admin=True)
        db.session.add(admin)
        db.session.commit()

        return render_template('create_admin.html', success='管理员创建成功！')

    return render_template('create_admin.html')
    except Exception as e:
        app.logger.error(f"加载后台页面失败: {e}")
        abort(500, description="加载后台页面失败")


# 报表页面
@app.route('/reports')
@login_required
def reports_page():
    """财务报表页面"""
    return render_template('reports.html')

@app.route('/static/<path:filename>')
def static_files(filename):
    """静态文件路由"""
    return send_from_directory(app.static_folder, filename)


# ==================== API 路由 ====================
@app.route('/api/transactions', methods=['GET', 'POST'])
@login_required
def handle_transactions():
    # 管理员只能查看，不能创建交易（self_view 模式下允许记账）
    if request.method == 'POST' and session.get('is_admin') and not session.get('self_view'):
        return jsonify({"success": False, "message": "管理员只能查看数据，不能记账"}), 403

    if request.method == 'GET':
        limit = request.args.get('limit', config.MAX_TRANSACTIONS_DISPLAY, type=int)
        period = request.args.get('period', None)
        date = request.args.get('date', None)
        account_id = request.args.get('account_id', None, type=int)
        page = request.args.get('page', None, type=int)
        per_page = request.args.get('per_page', 8, type=int)
        # 搜索筛选参数
        keyword = request.args.get('keyword', None)
        category = request.args.get('category', None)
        tx_type = request.args.get('type', None)
        amount_min = request.args.get('amount_min', None, type=float)
        amount_max = request.args.get('amount_max', None, type=float)
        start_date = request.args.get('start_date', None)
        end_date = request.args.get('end_date', None)
        transactions_query = filter_transactions_by_period_orm(period, date)
        if account_id:
            transactions_query = transactions_query.filter_by(account_id=account_id)
        if keyword:
            like_pattern = f'%{keyword}%'
            transactions_query = transactions_query.filter(
                db.or_(Transaction.remark.like(like_pattern), Transaction.category.like(like_pattern))
            )
        if category:
            transactions_query = transactions_query.filter(Transaction.category == category)
        if tx_type in ('income', 'expense'):
            transactions_query = transactions_query.filter(Transaction.type == tx_type)
        if amount_min is not None:
            transactions_query = transactions_query.filter(Transaction.amount >= amount_min)
        if amount_max is not None:
            transactions_query = transactions_query.filter(Transaction.amount <= amount_max)
        if start_date:
            transactions_query = transactions_query.filter(Transaction.date >= start_date)
        if end_date:
            transactions_query = transactions_query.filter(Transaction.date <= end_date)

        if page is not None:
            pagination = transactions_query.paginate(page=page, per_page=per_page, error_out=False)
            transactions = pagination.items
            # 确保总数统计也遵循权限过滤
            current_ledger_id = get_current_ledger_id()
            if current_ledger_id:
                total = Transaction.query.filter(Transaction.ledger_id == current_ledger_id).count()
            else:
                user_id = session.get('user_id')
                is_admin = session.get('is_admin')
                self_view = session.get('self_view', False)
                if user_id and is_admin and not self_view:
                    total = Transaction.query.count()
                elif user_id:
                    total = Transaction.query.filter(Transaction.user_id == user_id).count()
                else:
                    total = 0
            total_pages = pagination.pages
        else:
            transactions = transactions_query.limit(limit).all()
            # 总数也要遵循权限过滤
            current_ledger_id = get_current_ledger_id()
            if current_ledger_id:
                total = Transaction.query.filter(Transaction.ledger_id == current_ledger_id).count()
            else:
                user_id = session.get('user_id')
                is_admin = session.get('is_admin')
                self_view = session.get('self_view', False)
                if user_id and is_admin and not self_view:
                    total = Transaction.query.count()
                elif user_id:
                    total = Transaction.query.filter(Transaction.user_id == user_id).count()
                else:
                    total = 0
            total_pages = 1

        balance = get_balance()
        return jsonify({
            "success": True,
            "transactions": [
                {
                    "id": t.id,
                    "type": t.type,
                    "amount": float(t.amount) if t.amount else 0,
                    "category": t.category,
                    "date": t.date,
                    "time": t.time,
                    "remark": t.remark,
                    "account_id": t.account_id,
                    "account_name": t.account.name if t.account else None,
                    "currency": t.currency or 'CNY',
                    "original_amount": float(t.original_amount) if t.original_amount else None,
                    "exchange_rate": float(t.exchange_rate) if t.exchange_rate else None,
                    "created_at": t.created_at.isoformat() if t.created_at else None,
                    "updated_at": t.updated_at.isoformat() if t.updated_at else None,
                    "reimbursement_status": t.reimbursement_status or 'none',
                    "reimbursed_amount": float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                    "write_off_id": t.write_off_id,
                    "user_id": t.user_id,
                    "username": t.user.username if t.user else None
                } for t in transactions
            ],
            "balance": balance,
            "total": total,
            "page": page or 1,
            "per_page": per_page,
            "total_pages": total_pages
        })
    elif request.method == 'POST':
        try:
            transaction = request.get_json()
            user_id = session.get('user_id')

            # 检查用户在当前账本中是否有编辑权限
            current_ledger_id = get_current_ledger_id()
            if current_ledger_id:
                has_access, role, error = require_ledger_access(current_ledger_id, 'editor')
                if not has_access:
                    return error

            # 字段校验
            tx_type = transaction.get('type')
            if tx_type not in ['income', 'expense']:
                return jsonify({"success": False, "message": "交易类型必须是 income 或 expense"}), 400

            # 金额校验
            try:
                original_amount = float(transaction.get('amount', 0))
                if original_amount <= 0:
                    return jsonify({"success": False, "message": "金额必须大于0"}), 400
                if original_amount > 99999999.99:
                    return jsonify({"success": False, "message": "金额超出限制"}), 400
            except (TypeError, ValueError):
                return jsonify({"success": False, "message": "金额格式不正确"}), 400

            # 币种处理
            currency = transaction.get('currency', 'CNY').upper()
            if currency not in CURRENCY_NAMES and currency != 'CNY':
                return jsonify({"success": False, "message": f"不支持的币种: {currency}"}), 400

            exchange_rate = None
            tx_amount = original_amount
            tx_original_amount = None

            if currency == 'CNY':
                tx_amount = original_amount
            else:
                tx_original_amount = original_amount
                rate = get_exchange_rate(currency, 'CNY')
                if rate is None:
                    return jsonify({
                        "success": False,
                        "message": f"获取 {currency} 汇率失败，请稍后重试或使用人民币记账"
                    }), 502
                exchange_rate = round(rate, 6)
                tx_amount = round(original_amount * rate, 2)
                if tx_amount > 99999999.99:
                    return jsonify({"success": False, "message": "换算后金额超出限制"}), 400

            # 分类校验
            category_name = transaction.get('category', '').strip()
            if not category_name:
                return jsonify({"success": False, "message": "分类不能为空"}), 400
            category = Category.query.filter_by(name=category_name, type=tx_type).first()
            if not category:
                return jsonify({"success": False, "message": f"分类 '{category_name}' 不存在或类型不匹配"}), 400

            # 日期校验
            tx_date = transaction.get('date', '').strip()
            now = datetime.now()
            if not tx_date:
                tx_date = now.strftime('%Y-%m-%d')
            else:
                try:
                    parsed_date = datetime.strptime(tx_date, '%Y-%m-%d')
                    if parsed_date.date() > now.date():
                        return jsonify({"success": False, "message": "日期不能晚于今天"}), 400
                except ValueError:
                    return jsonify({"success": False, "message": "日期格式不正确，应为 YYYY-MM-DD"}), 400

            # 时间校验
            tx_time = transaction.get('time', '').strip()
            if not tx_time:
                tx_time = now.strftime('%H:%M:%S')
            else:
                try:
                    datetime.strptime(tx_time, '%H:%M:%S')
                except ValueError:
                    try:
                        datetime.strptime(tx_time, '%H:%M')
                        tx_time = tx_time + ':00'
                    except ValueError:
                        tx_time = now.strftime('%H:%M:%S')

            # 账户校验
            account_id = transaction.get('account_id')
            account = None
            current_ledger_id = get_current_ledger_id()
            if account_id:
                account = Account.query.filter_by(id=account_id).first()
                if not account or (current_ledger_id and account.ledger_id != current_ledger_id):
                    return jsonify({"success": False, "message": "账户不存在或不属于当前账本"}), 400
            else:
                if current_ledger_id:
                    account = Account.query.filter_by(ledger_id=current_ledger_id).first()
                else:
                    account = Account.query.filter_by(user_id=user_id).first()
                if not account:
                    account = Account(name='默认账户', balance=0, account_type='cash', user_id=user_id, ledger_id=current_ledger_id)
                    db.session.add(account)
                    db.session.flush()
                account_id = account.id

            # 创建交易记录
            tx = Transaction(
                type=tx_type,
                amount=tx_amount,
                category=category_name,
                date=tx_date,
                time=tx_time,
                remark=transaction.get('remark', '').strip(),
                account_id=account_id,
                user_id=user_id,
                ledger_id=current_ledger_id,
                currency=currency,
                original_amount=tx_original_amount,
                exchange_rate=exchange_rate,
                reimbursement_status=transaction.get('reimbursement_status', 'none')
            )
            db.session.add(tx)

            # 更新账户余额
            if tx_type == 'income':
                account.balance = float(account.balance) + tx_amount
            else:
                account.balance = float(account.balance) - tx_amount

            db.session.commit()

            balance = get_balance()
            app.logger.info(f"新增交易: {tx_type} ¥{tx_amount} - {category_name}, 账户余额: {account.balance}")

            # 支出时检查预算
            budget_warnings = []
            if tx_type == 'expense':
                try:
                    month = tx_date[:7]
                    # 优先查找账户级预算，再查总账户预算
                    budget_filter = {'user_id': user_id, 'month': month}
                    if current_ledger_id:
                        budget_filter['ledger_id'] = current_ledger_id
                    if account_id:
                        budget_filter['account_id'] = account_id
                        budget = Budget.query.filter_by(**budget_filter).first()
                    if not budget:
                        budget_filter.pop('account_id', None)
                        budget_filter['account_id'] = None
                        budget = Budget.query.filter_by(**budget_filter).first()
                    else:
                        budget = Budget.query.filter_by(**budget_filter).first()

                    if budget:
                        total_amount = float(budget.total_amount)
                        expense_filter = [
                            Transaction.type == 'expense',
                            Transaction.date.startswith(month)
                        ]
                        if current_ledger_id:
                            expense_filter.append(Transaction.ledger_id == current_ledger_id)
                        else:
                            expense_filter.append(Transaction.user_id == user_id)
                        if budget.account_id:
                            expense_filter.append(Transaction.account_id == budget.account_id)
                        month_expenses = db.session.query(db.func.sum(Transaction.amount)).filter(
                            *expense_filter
                        ).scalar() or 0
                        month_expenses = float(month_expenses)

                        # 检查总预算
                        if month_expenses > total_amount:
                            budget_warnings.append({
                                'type': 'total',
                                'message': f'本月支出 ¥{month_expenses:.2f}，已超出预算 ¥{total_amount:.2f}，超出 ¥{month_expenses - total_amount:.2f}'
                            })
                        elif month_expenses >= total_amount * 0.9:
                            budget_warnings.append({
                                'type': 'total_warn',
                                'message': f'本月支出 ¥{month_expenses:.2f}，已达预算 ¥{total_amount:.2f} 的 {month_expenses / total_amount * 100:.1f}%'
                            })

                        # 检查分类预算
                        cat_items = BudgetCategoryItem.query.filter_by(budget_id=budget.id).all()
                        for item in cat_items:
                            cat_name = item.category.name if item.category else None
                            if not cat_name or cat_name != category_name:
                                continue
                            cat_amount = float(item.amount)
                            cat_exp_filter = [
                                Transaction.type == 'expense',
                                Transaction.category == cat_name,
                                Transaction.date.startswith(month)
                            ]
                            if current_ledger_id:
                                cat_exp_filter.append(Transaction.ledger_id == current_ledger_id)
                            else:
                                cat_exp_filter.append(Transaction.user_id == user_id)
                            if budget.account_id:
                                cat_exp_filter.append(Transaction.account_id == budget.account_id)
                            cat_expense = db.session.query(db.func.sum(Transaction.amount)).filter(
                                *cat_exp_filter
                            ).scalar() or 0
                            cat_expense = float(cat_expense)
                            if cat_expense > cat_amount:
                                budget_warnings.append({
                                    'type': 'category',
                                    'category': cat_name,
                                    'message': f'分类「{cat_name}」本月支出 ¥{cat_expense:.2f}，已超出预算 ¥{cat_amount:.2f}'
                                })
                            elif cat_expense >= cat_amount * 0.9:
                                budget_warnings.append({
                                    'type': 'category_warn',
                                    'category': cat_name,
                                    'message': f'分类「{cat_name}」支出 ¥{cat_expense:.2f}，已达预算 ¥{cat_amount:.2f} 的 {cat_expense / cat_amount * 100:.1f}%'
                                })
                except Exception as e:
                    app.logger.warning(f"预算检查失败: {e}")

            # 报销收入关联：手动添加"报销收入"时可关联原始支出
            write_off_warning = None
            if tx_type == 'income' and category_name == '报销收入':
                linked_expense_id = transaction.get('write_off_id')
                if linked_expense_id:
                    try:
                        expense_tx = Transaction.query.get(linked_expense_id)
                        if expense_tx and expense_tx.type == 'expense' and expense_tx.reimbursement_status != 'none':
                            current_reimbursed = float(expense_tx.reimbursed_amount or 0)
                            new_total = current_reimbursed + tx_amount
                            orig_amount = float(expense_tx.amount)
                            if new_total > orig_amount:
                                write_off_warning = f'核销金额累计 {new_total:.2f} 超出原支出金额 {orig_amount:.2f}'
                            else:
                                expense_tx.reimbursed_amount = new_total
                                expense_tx.write_off_id = tx.id
                                expense_tx.reimbursement_status = 'reimbursed' if new_total >= orig_amount else 'partial'
                                expense_tx.updated_at = datetime.now()
                                db.session.commit()
                                app.logger.info(f"报销收入关联支出: income={tx.id}, expense={expense_tx.id}")
                    except Exception as e:
                        app.logger.warning(f"报销收入关联失败: {e}")
                else:
                    write_off_warning = '未关联原始支出，请前往报销管理进行核销'

            response_data = {
                "success": True,
                "message": "交易添加成功",
                "balance": balance,
                "account_balance": float(account.balance),
                "transaction_id": tx.id
            }
            if budget_warnings:
                response_data['budget_warnings'] = budget_warnings
                response_data['message'] = '交易添加成功，但存在预算超支警告'
            if write_off_warning:
                response_data['write_off_warning'] = write_off_warning
                if not response_data.get('budget_warnings'):
                    response_data['message'] = '交易添加成功，' + write_off_warning
            return jsonify(response_data), 201
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"添加交易失败: {e}")
            return jsonify({"success": False, "message": "添加交易失败"}), 500

@app.route('/api/transactions/<int:transaction_id>', methods=['DELETE'])
@login_required
def delete_transaction(transaction_id):
    try:
        tx = Transaction.query.get(transaction_id)
        if not tx:
            return jsonify({"success": False, "message": "未找到对应交易记录"}), 404

        # 检查权限：管理员在 self_view 模式下可以删除自己的交易，否则只能查看
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        if is_admin and not self_view:
            return jsonify({"success": False, "message": "管理员只能查看数据，不能删除交易"}), 403
        if tx.user_id != user_id:
            # Check if user has editor role in the transaction's ledger
            if tx.ledger_id:
                has_access, role, error = require_ledger_access(tx.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限删除此记录"}), 403

        # 回滚账户余额
        account = Account.query.get(tx.account_id) if tx.account_id else None
        if account:
            if tx.type == 'income':
                account.balance = float(account.balance) - float(tx.amount)
            elif tx.type == 'expense':
                account.balance = float(account.balance) + float(tx.amount)

        # 处理核销关联：删除已报销的支出时同步删除关联的收入
        if tx.type == 'expense' and tx.write_off_id:
            linked_income = Transaction.query.get(tx.write_off_id)
            if linked_income:
                # 回滚关联收入的账户余额
                linked_account = Account.query.get(linked_income.account_id) if linked_income.account_id else None
                if linked_account:
                    linked_account.balance = float(linked_account.balance) - float(linked_income.amount)
                db.session.delete(linked_income)
        # 处理核销关联：删除报销收入时重置对应支出的报销状态
        elif tx.type == 'income':
            linked_expenses = Transaction.query.filter_by(write_off_id=transaction_id).all()
            for exp in linked_expenses:
                exp.reimbursement_status = 'none'
                exp.reimbursed_amount = 0
                exp.write_off_id = None

        db.session.delete(tx)
        db.session.commit()
        balance = get_balance()
        app.logger.info(f"删除交易: ID={transaction_id}, {tx.type} ¥{tx.amount}")
        return jsonify({
            "success": True,
            "message": "记录已删除",
            "new_balance": balance
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除交易失败: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/transactions/<int:transaction_id>', methods=['PUT'])
@login_required
def update_transaction(transaction_id):
    try:
        update_data = request.get_json()
        tx = Transaction.query.get(transaction_id)
        if not tx:
            return jsonify({"success": False, "message": "未找到对应交易记录"}), 404

        # 检查权限：管理员在 self_view 模式下可以修改自己的交易，否则只能查看
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        if is_admin and not self_view:
            return jsonify({"success": False, "message": "管理员只能查看数据，不能修改交易"}), 403
        if tx.user_id != user_id:
            if tx.ledger_id:
                has_access, role, error = require_ledger_access(tx.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限修改此记录"}), 403

        # 先回滚旧的账户余额影响
        old_type = tx.type
        old_amount = float(tx.amount)
        account = Account.query.get(tx.account_id) if tx.account_id else None
        if account:
            if old_type == 'income':
                account.balance = float(account.balance) - old_amount
            else:
                account.balance = float(account.balance) + old_amount

        # 更新字段
        new_type = update_data.get('type', old_type)
        new_amount = old_amount
        if new_type not in ['income', 'expense']:
            return jsonify({"success": False, "message": "交易类型必须是 income 或 expense"}), 400
        tx.type = new_type

        if 'amount' in update_data:
            try:
                raw_amount = float(update_data['amount'])
                if raw_amount <= 0:
                    return jsonify({"success": False, "message": "金额必须大于0"}), 400
                if raw_amount > 99999999.99:
                    return jsonify({"success": False, "message": "金额超出限制"}), 400
            except (TypeError, ValueError):
                return jsonify({"success": False, "message": "金额格式不正确"}), 400

            # 币种处理
            currency = update_data.get('currency', tx.currency or 'CNY').upper()
            if currency not in CURRENCY_NAMES and currency != 'CNY':
                return jsonify({"success": False, "message": f"不支持的币种: {currency}"}), 400

            if currency == 'CNY':
                new_amount = raw_amount
                tx.currency = 'CNY'
                tx.original_amount = None
                tx.exchange_rate = None
            else:
                rate = get_exchange_rate(currency, 'CNY')
                if rate is None:
                    return jsonify({
                        "success": False,
                        "message": f"获取 {currency} 汇率失败，请稍后重试或使用人民币记账"
                    }), 502
                new_amount = round(raw_amount * rate, 2)
                if new_amount > 99999999.99:
                    return jsonify({"success": False, "message": "换算后金额超出限制"}), 400
                tx.currency = currency
                tx.original_amount = raw_amount
                tx.exchange_rate = round(rate, 6)

            tx.amount = new_amount

        if 'category' in update_data and update_data['category']:
            category_name = update_data['category'].strip()
            category = Category.query.filter_by(name=category_name, type=new_type).first()
            if not category:
                return jsonify({"success": False, "message": f"分类 '{category_name}' 不存在或类型不匹配"}), 400
            tx.category = category_name

        if 'remark' in update_data:
            tx.remark = update_data['remark'].strip() if update_data['remark'] else ''
        if 'account_id' in update_data:
            new_account = Account.query.get(update_data['account_id'])
            if new_account:
                tx.account_id = update_data['account_id']
                account = new_account

        tx.updated_at = datetime.now()

        # 应用新余额影响
        if account:
            if new_type == 'income':
                account.balance = float(account.balance) + new_amount
            else:
                account.balance = float(account.balance) - new_amount

        db.session.commit()
        balance = get_balance()
        app.logger.info(f"更新交易: ID={transaction_id}, {new_type} ¥{new_amount}")
        return jsonify({
            "success": True,
            "message": "记录已更新",
            "transaction": {
                "id": tx.id,
                "type": tx.type,
                "amount": float(tx.amount),
                "category": tx.category,
                "date": tx.date,
                "time": tx.time,
                "remark": tx.remark,
                "account_id": tx.account_id,
                "currency": tx.currency or 'CNY',
                "original_amount": float(tx.original_amount) if tx.original_amount else None,
                "exchange_rate": float(tx.exchange_rate) if tx.exchange_rate else None,
                "created_at": tx.created_at.isoformat() if tx.created_at else None,
                "updated_at": tx.updated_at.isoformat() if tx.updated_at else None,
                "reimbursement_status": tx.reimbursement_status or 'none',
                "reimbursed_amount": float(tx.reimbursed_amount) if tx.reimbursed_amount else 0,
                "write_off_id": tx.write_off_id
            },
            "new_balance": balance
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新交易失败: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# ==================== 报销管理 API ====================

@app.route('/api/transactions/<int:transaction_id>/reimbursement', methods=['PUT'])
@login_required
def update_reimbursement(transaction_id):
    """Update reimbursement status of a transaction"""
    try:
        tx = Transaction.query.get(transaction_id)
        if not tx:
            return jsonify({"success": False, "message": "未找到对应交易记录"}), 404

        # 检查权限：管理员在 self_view 模式下可以操作报销
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        if is_admin and not self_view:
            return jsonify({"success": False, "message": "管理员只能查看数据，不能操作报销"}), 403
        if tx.user_id != user_id:
            return jsonify({"success": False, "message": "无权限修改此记录"}), 403
        if tx.type != 'expense':
            return jsonify({"success": False, "message": "只有支出记录可以标记报销"}), 400

        data = request.get_json()
        new_status = data.get('reimbursement_status')
        if new_status not in ('none', 'pending', 'partial', 'reimbursed'):
            return jsonify({"success": False, "message": "无效的报销状态"}), 400

        tx.reimbursement_status = new_status
        tx.updated_at = datetime.now()
        db.session.commit()

        app.logger.info(f"更新报销状态: transaction={transaction_id}, status={new_status}")
        return jsonify({
            "success": True,
            "message": "报销状态已更新",
            "transaction": {
                "id": tx.id,
                "reimbursement_status": tx.reimbursement_status,
                "reimbursed_amount": float(tx.reimbursed_amount) if tx.reimbursed_amount else 0,
                "write_off_id": tx.write_off_id
            }
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新报销状态失败: {e}")
        return jsonify({"success": False, "message": "更新报销状态失败"}), 500


@app.route('/api/transactions/<int:transaction_id>/write-off', methods=['POST'])
@login_required
def create_write_off(transaction_id):
    """Create reimbursement income and link as write-off to the original expense"""
    try:
        tx = Transaction.query.get(transaction_id)
        if not tx:
            return jsonify({"success": False, "message": "未找到对应交易记录"}), 404

        # 检查权限：管理员在 self_view 模式下可以核销
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        if is_admin and not self_view:
            return jsonify({"success": False, "message": "管理员只能查看数据，不能核销报销"}), 403
        if tx.user_id != user_id:
            return jsonify({"success": False, "message": "无权限修改此记录"}), 403
        if tx.type != 'expense':
            return jsonify({"success": False, "message": "只有支出记录可以核销"}), 400
        if tx.reimbursement_status == 'none':
            return jsonify({"success": False, "message": "该记录未标记为报销，请先标记"}), 400

        data = request.get_json()
        write_off_amount = float(data.get('amount', 0))
        if write_off_amount <= 0:
            return jsonify({"success": False, "message": "核销金额必须大于0"}), 400

        original_amount = float(tx.amount)
        current_reimbursed = float(tx.reimbursed_amount or 0)
        new_total_reimbursed = current_reimbursed + write_off_amount

        if new_total_reimbursed > original_amount:
            return jsonify({
                "success": False,
                "message": f"核销金额累计 {new_total_reimbursed:.2f} 超出原支出金额 {original_amount:.2f}"
            }), 400

        # Create the reimbursement income transaction
        write_off_tx = Transaction(
            type='income',
            amount=write_off_amount,
            category='报销收入',
            date=data.get('date', datetime.now().strftime('%Y-%m-%d')),
            time=data.get('time', datetime.now().strftime('%H:%M:%S')),
            remark=data.get('remark', f'核销报销: {tx.remark or tx.category}')[:255],
            account_id=tx.account_id,
            user_id=session.get('user_id'),
            currency='CNY',
            ledger_id=tx.ledger_id
        )
        db.session.add(write_off_tx)
        db.session.flush()

        # Update the original expense
        tx.reimbursed_amount = new_total_reimbursed
        tx.write_off_id = write_off_tx.id
        tx.reimbursement_status = 'reimbursed' if new_total_reimbursed >= original_amount else 'partial'
        tx.updated_at = datetime.now()

        # Update account balance for the reimbursement income
        account = Account.query.get(tx.account_id) if tx.account_id else None
        if account:
            account.balance = float(account.balance) + write_off_amount

        db.session.commit()

        app.logger.info(f"核销报销: expense={transaction_id}, amount={write_off_amount}, income={write_off_tx.id}")
        return jsonify({
            "success": True,
            "message": "核销成功",
            "write_off_transaction": {
                "id": write_off_tx.id,
                "amount": write_off_amount,
                "date": write_off_tx.date
            },
            "updated_expense": {
                "id": tx.id,
                "reimbursement_status": tx.reimbursement_status,
                "reimbursed_amount": float(tx.reimbursed_amount),
                "write_off_id": tx.write_off_id
            },
            "account_balance": float(account.balance) if account else None
        }), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"核销报销失败: {e}")
        return jsonify({"success": False, "message": "核销失败"}), 500


@app.route('/api/reimbursements', methods=['GET'])
@login_required
def list_reimbursements():
    """List expenses with reimbursement activity, with optional filtering"""
    try:
        status_filter = request.args.get('status', None)
        account_id = request.args.get('account_id', None, type=int)
        date_from = request.args.get('date_from', None)
        date_to = request.args.get('date_to', None)

        query = Transaction.query.filter(
            Transaction.type == 'expense',
            Transaction.reimbursement_status != 'none'
        )

        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            query = query.filter(Transaction.ledger_id == current_ledger_id)
        elif user_id and not is_admin or (is_admin and self_view):
            # 普通用户或管理员 self_view 只能看到自己的数据
            query = query.filter(Transaction.user_id == user_id)

        if status_filter and status_filter != 'all':
            query = query.filter(Transaction.reimbursement_status == status_filter)
        if account_id:
            query = query.filter_by(account_id=account_id)
        if date_from:
            query = query.filter(Transaction.date >= date_from)
        if date_to:
            query = query.filter(Transaction.date <= date_to)

        query = query.order_by(Transaction.id.desc())
        transactions = query.all()

        # Build response with optional write-off info
        result = []
        for t in transactions:
            item = {
                "id": t.id,
                "amount": float(t.amount),
                "category": t.category,
                "date": t.date,
                "remark": t.remark,
                "account_id": t.account_id,
                "account_name": t.account.name if t.account else None,
                "reimbursement_status": t.reimbursement_status,
                "reimbursed_amount": float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                "write_off_id": t.write_off_id,
                "created_at": t.created_at.isoformat() if t.created_at else None,
                "user_id": t.user_id,
                "username": t.user.username if t.user else None
            }
            # Include linked write-off income info
            if t.write_off_id:
                if is_admin and not self_view:
                    # 管理员可以看到所有关联数据
                    wot = Transaction.query.get(t.write_off_id)
                    if wot:
                        item["write_off_transaction"] = {
                            "id": wot.id,
                            "amount": float(wot.amount),
                            "date": wot.date,
                            "remark": wot.remark
                        }
                else:
                    # 普通用户或管理员 self_view 只能看到自己的关联数据
                    if t.user_id == user_id:
                        wot = Transaction.query.get(t.write_off_id)
                        if wot:
                            item["write_off_transaction"] = {
                                "id": wot.id,
                                "amount": float(wot.amount),
                                "date": wot.date,
                                "remark": wot.remark
                            }
            result.append(item)

        # 统计计数（根据权限）
        def _count_with_ledger(base_filter):
            if current_ledger_id:
                return Transaction.query.filter(*base_filter, Transaction.ledger_id == current_ledger_id).count()
            if is_admin and not self_view:
                return Transaction.query.filter(*base_filter).count()
            return Transaction.query.filter(*base_filter, Transaction.user_id == user_id).count()

        pending_count = _count_with_ledger([
            Transaction.type == 'expense',
            Transaction.reimbursement_status == 'pending'
        ])
        partial_count = _count_with_ledger([
            Transaction.type == 'expense',
            Transaction.reimbursement_status == 'partial'
        ])
        reimbursed_count = _count_with_ledger([
            Transaction.type == 'expense',
            Transaction.reimbursement_status == 'reimbursed'
        ])

        return jsonify({
            "success": True,
            "reimbursements": result,
            "total": len(result),
            "summary": {
                "total_pending": pending_count,
                "total_partial": partial_count,
                "total_reimbursed": reimbursed_count
            }
        })
    except Exception as e:
        app.logger.error(f"获取报销列表失败: {e}")
        return jsonify({"success": False, "message": "获取报销列表失败"}), 500


# ==================== 借贷管理 API ====================

def get_loan_base_query():
    """构建借贷基础查询（按权限过滤）"""
    user_id = session.get('user_id')
    is_admin = session.get('is_admin')
    self_view = session.get('self_view', False)
    current_ledger_id = get_current_ledger_id()

    query = Loan.query

    if current_ledger_id:
        query = query.filter(Loan.ledger_id == current_ledger_id)
    elif is_admin and not self_view:
        pass  # 管理员后台查看全部
    else:
        query = query.filter(Loan.user_id == user_id)

    return query


def serialize_loan(loan):
    """序列化借贷记录"""
    remaining = float(loan.amount) - float(loan.repaid_amount or 0)
    return {
        "id": loan.id,
        "type": loan.type,
        "counterparty": loan.counterparty,
        "amount": float(loan.amount),
        "repaid_amount": float(loan.repaid_amount or 0),
        "remaining": round(remaining, 2),
        "date": loan.date,
        "due_date": loan.due_date,
        "status": loan.status,
        "remark": loan.remark or '',
        "user_id": loan.user_id,
        "ledger_id": loan.ledger_id,
        "created_at": loan.created_at.isoformat() if loan.created_at else None,
        "updated_at": loan.updated_at.isoformat() if loan.updated_at else None
    }


@app.route('/api/loans', methods=['GET'])
@login_required
def list_loans():
    """获取借贷列表及汇总统计"""
    try:
        loan_type = request.args.get('type')  # 'borrow' or 'lend'
        status_filter = request.args.get('status')
        keyword = request.args.get('keyword')

        query = get_loan_base_query()

        if loan_type in ('borrow', 'lend'):
            query = query.filter(Loan.type == loan_type)
        if status_filter in ('active', 'settled'):
            query = query.filter(Loan.status == status_filter)
        if keyword:
            like_pattern = f'%{keyword}%'
            query = query.filter(
                db.or_(Loan.counterparty.like(like_pattern), db.func.ifnull(Loan.remark, '').like(like_pattern))
            )

        query = query.order_by(Loan.id.desc())
        loans = query.all()

        # 汇总统计
        total_borrow = 0.0      # 总借入（负债）
        total_borrow_remaining = 0.0
        total_lend = 0.0        # 总借出（债权）
        total_lend_remaining = 0.0

        for loan in loans:
            amount = float(loan.amount)
            remaining = amount - float(loan.repaid_amount or 0)
            if loan.type == 'borrow':
                total_borrow += amount
                total_borrow_remaining += remaining
            else:
                total_lend += amount
                total_lend_remaining += remaining

        return jsonify({
            "success": True,
            "loans": [serialize_loan(l) for l in loans],
            "total": len(loans),
            "summary": {
                "total_borrow": round(total_borrow, 2),
                "total_borrow_remaining": round(total_borrow_remaining, 2),
                "total_lend": round(total_lend, 2),
                "total_lend_remaining": round(total_lend_remaining, 2)
            }
        })
    except Exception as e:
        app.logger.error(f"获取借贷列表失败: {e}")
        return jsonify({"success": False, "message": "获取借贷列表失败"}), 500


@app.route('/api/loans', methods=['POST'])
@login_required
def create_loan():
    """创建借贷记录"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')

        # 权限检查
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            has_access, role, error = require_ledger_access(current_ledger_id, 'editor')
            if not has_access:
                return error

        # 字段校验
        loan_type = data.get('type')
        if loan_type not in ('borrow', 'lend'):
            return jsonify({"success": False, "message": "借贷类型必须是 borrow 或 lend"}), 400

        counterparty = (data.get('counterparty') or '').strip()
        if not counterparty:
            return jsonify({"success": False, "message": "对方名称不能为空"}), 400

        try:
            amount = float(data.get('amount', 0))
            if amount <= 0:
                return jsonify({"success": False, "message": "金额必须大于0"}), 400
            if amount > 99999999.99:
                return jsonify({"success": False, "message": "金额超出限制"}), 400
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "金额格式不正确"}), 400

        loan_date = (data.get('date') or '').strip()
        if not loan_date:
            loan_date = datetime.now().strftime('%Y-%m-%d')

        due_date = (data.get('due_date') or '').strip() or None

        remark = (data.get('remark') or '').strip()

        loan = Loan(
            type=loan_type,
            counterparty=counterparty,
            amount=amount,
            repaid_amount=0,
            date=loan_date,
            due_date=due_date,
            status='active',
            remark=remark,
            user_id=user_id,
            ledger_id=current_ledger_id
        )
        db.session.add(loan)
        db.session.commit()

        app.logger.info(f"新增借贷: {loan_type} ¥{amount} - {counterparty}")
        return jsonify({
            "success": True,
            "message": "借贷记录添加成功",
            "loan": serialize_loan(loan)
        }), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建借贷记录失败: {e}")
        return jsonify({"success": False, "message": "创建借贷记录失败"}), 500


@app.route('/api/loans/<int:loan_id>', methods=['PUT'])
@login_required
def update_loan(loan_id):
    """更新借贷记录"""
    try:
        loan = Loan.query.get(loan_id)
        if not loan:
            return jsonify({"success": False, "message": "未找到该借贷记录"}), 404

        # 权限检查
        user_id = session.get('user_id')
        if loan.user_id != user_id:
            if loan.ledger_id:
                has_access, role, error = require_ledger_access(loan.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限修改此记录"}), 403

        data = request.get_json()

        if 'type' in data:
            if data['type'] not in ('borrow', 'lend'):
                return jsonify({"success": False, "message": "借贷类型必须是 borrow 或 lend"}), 400
            loan.type = data['type']

        if 'counterparty' in data:
            counterparty = data['counterparty'].strip()
            if not counterparty:
                return jsonify({"success": False, "message": "对方名称不能为空"}), 400
            loan.counterparty = counterparty

        if 'amount' in data:
            try:
                amount = float(data['amount'])
                if amount <= 0:
                    return jsonify({"success": False, "message": "金额必须大于0"}), 400
                if amount > 99999999.99:
                    return jsonify({"success": False, "message": "金额超出限制"}), 400
                # 检查已还金额不能超过新总额
                if float(loan.repaid_amount or 0) > amount:
                    return jsonify({"success": False, "message": "已还金额不能大于总金额"}), 400
                loan.amount = amount
            except (TypeError, ValueError):
                return jsonify({"success": False, "message": "金额格式不正确"}), 400

        if 'date' in data:
            loan.date = data['date']
        if 'due_date' in data:
            loan.due_date = data['due_date'].strip() or None
        if 'status' in data:
            if data['status'] not in ('active', 'settled'):
                return jsonify({"success": False, "message": "状态无效"}), 400
            loan.status = data['status']
        if 'remark' in data:
            loan.remark = data['remark'].strip()

        loan.updated_at = datetime.now()
        db.session.commit()

        app.logger.info(f"更新借贷: ID={loan_id}")
        return jsonify({
            "success": True,
            "message": "借贷记录已更新",
            "loan": serialize_loan(loan)
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新借贷记录失败: {e}")
        return jsonify({"success": False, "message": "更新借贷记录失败"}), 500


@app.route('/api/loans/<int:loan_id>', methods=['DELETE'])
@login_required
def delete_loan(loan_id):
    """删除借贷记录"""
    try:
        loan = Loan.query.get(loan_id)
        if not loan:
            return jsonify({"success": False, "message": "未找到该借贷记录"}), 404

        # 权限检查
        user_id = session.get('user_id')
        if loan.user_id != user_id:
            if loan.ledger_id:
                has_access, role, error = require_ledger_access(loan.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限删除此记录"}), 403

        db.session.delete(loan)
        db.session.commit()

        app.logger.info(f"删除借贷: ID={loan_id}")
        return jsonify({
            "success": True,
            "message": "借贷记录已删除"
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除借贷记录失败: {e}")
        return jsonify({"success": False, "message": "删除借贷记录失败"}), 500


@app.route('/api/loans/<int:loan_id>/repay', methods=['POST'])
@login_required
def repay_loan(loan_id):
    """记录还款/收款"""
    try:
        loan = Loan.query.get(loan_id)
        if not loan:
            return jsonify({"success": False, "message": "未找到该借贷记录"}), 404

        # 权限检查
        user_id = session.get('user_id')
        if loan.user_id != user_id:
            if loan.ledger_id:
                has_access, role, error = require_ledger_access(loan.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({"success": False, "message": "无权限操作此记录"}), 403

        if loan.status == 'settled':
            return jsonify({"success": False, "message": "该借贷已结清，无法继续还款"}), 400

        data = request.get_json()
        try:
            repay_amount = float(data.get('amount', 0))
            if repay_amount <= 0:
                return jsonify({"success": False, "message": "还款金额必须大于0"}), 400
        except (TypeError, ValueError):
            return jsonify({"success": False, "message": "金额格式不正确"}), 400

        current_repaid = float(loan.repaid_amount or 0)
        new_total_repaid = current_repaid + repay_amount
        total_amount = float(loan.amount)

        # 超额还款检查
        if new_total_repaid > total_amount:
            return jsonify({
                "success": False,
                "message": f"超额还款错误：累计还款 ¥{new_total_repaid:.2f} 超出借贷总额 ¥{total_amount:.2f}，多出 ¥{new_total_repaid - total_amount:.2f}"
            }), 400

        loan.repaid_amount = new_total_repaid

        # 如果已还清，自动标记为已结清
        if new_total_repaid >= total_amount:
            loan.status = 'settled'

        loan.updated_at = datetime.now()
        db.session.commit()

        action = '收款' if loan.type == 'lend' else '还款'
        app.logger.info(f"借贷{action}: loan={loan_id}, amount={repay_amount}")

        action_msg = '收款' if loan.type == 'lend' else '还款'
        return jsonify({
            "success": True,
            "message": f"{action_msg}记录成功",
            "loan": serialize_loan(loan)
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"还款操作失败: {e}")
        return jsonify({"success": False, "message": "还款操作失败"}), 500


@app.route('/api/loans/summary', methods=['GET'])
@login_required
def loan_summary():
    """获取借贷汇总（用于资产/负债统计）"""
    try:
        query = get_loan_base_query()
        loans = query.all()

        total_borrow_remaining = 0.0  # 总负债
        total_lend_remaining = 0.0    # 总债权

        for loan in loans:
            remaining = float(loan.amount) - float(loan.repaid_amount or 0)
            if remaining > 0:
                if loan.type == 'borrow':
                    total_borrow_remaining += remaining
                else:
                    total_lend_remaining += remaining

        net_loan_asset = round(total_lend_remaining - total_borrow_remaining, 2)

        return jsonify({
            "success": True,
            "total_borrow_remaining": round(total_borrow_remaining, 2),
            "total_lend_remaining": round(total_lend_remaining, 2),
            "net_loan_asset": net_loan_asset
        })
    except Exception as e:
        app.logger.error(f"获取借贷汇总失败: {e}")
        return jsonify({"success": False, "message": "获取借贷汇总失败"}), 500


# ==================== 周期账单 API ====================

def _compute_next_date(current_date_str, period, interval_value=1):
    """根据周期计算下一次生成日期"""
    from datetime import datetime as dt
    from dateutil.relativedelta import relativedelta
    cur = dt.strptime(current_date_str, '%Y-%m-%d').date()
    if period == 'daily':
        next_d = cur + relativedelta(days=interval_value)
    elif period == 'weekly':
        next_d = cur + relativedelta(weeks=interval_value)
    elif period == 'monthly':
        next_d = cur + relativedelta(months=interval_value)
    elif period == 'yearly':
        next_d = cur + relativedelta(years=interval_value)
    else:
        next_d = cur + relativedelta(months=1)
    return next_d.strftime('%Y-%m-%d')


def _generate_rule_transaction(rule):
    """为一条规则生成一笔交易，更新 next_date。返回生成的 Transaction 或 None"""
    today_str = datetime.now().strftime('%Y-%m-%d')
    if rule.next_date > today_str:
        return None
    if rule.end_date and rule.end_date < today_str:
        return None
    if not rule.is_active:
        return None

    account = Account.query.get(rule.account_id) if rule.account_id else None
    if not account:
        account = Account.query.filter_by(ledger_id=rule.ledger_id).first()
    if not account:
        account = Account.query.filter_by(user_id=rule.user_id).first()
    if not account:
        account = Account(name='默认账户', balance=0, account_type='cash',
                          user_id=rule.user_id, ledger_id=rule.ledger_id)
        db.session.add(account)
        db.session.flush()

    now = datetime.now()
    tx = Transaction(
        type=rule.type,
        amount=float(rule.amount),
        category=rule.category,
        date=rule.next_date,
        time=now.strftime('%H:%M:%S'),
        remark=f'[周期] {rule.name}' + (f' - {rule.remark}' if rule.remark else ''),
        account_id=account.id,
        user_id=rule.user_id,
        ledger_id=rule.ledger_id,
        currency='CNY',
    )
    db.session.add(tx)

    if rule.type == 'income':
        account.balance = float(account.balance) + float(rule.amount)
    else:
        account.balance = float(account.balance) - float(rule.amount)

    # 计算下一次生成日期
    old_next = rule.next_date
    rule.next_date = _compute_next_date(old_next, rule.period, rule.interval_value or 1)
    rule.updated_at = now

    # 如果下次日期已过，继续往前递推（补生）
    while rule.next_date < today_str and (not rule.end_date or rule.next_date <= rule.end_date):
        # 生成中间日期的交易
        tx_mid = Transaction(
            type=rule.type,
            amount=float(rule.amount),
            category=rule.category,
            date=rule.next_date,
            time=now.strftime('%H:%M:%S'),
            remark=f'[周期] {rule.name}' + (f' - {rule.remark}' if rule.remark else ''),
            account_id=account.id,
            user_id=rule.user_id,
            ledger_id=rule.ledger_id,
            currency='CNY',
        )
        db.session.add(tx_mid)
        if rule.type == 'income':
            account.balance = float(account.balance) + float(rule.amount)
        else:
            account.balance = float(account.balance) - float(rule.amount)
        rule.next_date = _compute_next_date(rule.next_date, rule.period, rule.interval_value or 1)
        rule.updated_at = now

    # 如果超出结束日期则停用
    if rule.end_date and rule.next_date > rule.end_date:
        rule.is_active = False

    return tx


def serialize_rule(rule):
    """序列化周期规则"""
    return {
        'id': rule.id,
        'name': rule.name,
        'amount': float(rule.amount),
        'category': rule.category,
        'type': rule.type,
        'period': rule.period,
        'interval_value': rule.interval_value or 1,
        'start_date': rule.start_date,
        'end_date': rule.end_date,
        'next_date': rule.next_date,
        'is_active': rule.is_active,
        'remark': rule.remark or '',
        'account_id': rule.account_id,
        'account_name': rule.account.name if rule.account else None,
        'user_id': rule.user_id,
        'ledger_id': rule.ledger_id,
        'created_at': rule.created_at.isoformat() if rule.created_at else None,
        'updated_at': rule.updated_at.isoformat() if rule.updated_at else None,
    }


@app.route('/api/recurring-rules', methods=['GET'])
@login_required
def list_recurring_rules():
    """获取周期账单规则列表"""
    try:
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()

        query = RecurringRule.query
        if current_ledger_id:
            query = query.filter(RecurringRule.ledger_id == current_ledger_id)
        elif is_admin and not self_view:
            pass
        else:
            query = query.filter(RecurringRule.user_id == user_id)

        query = query.order_by(RecurringRule.id.desc())
        rules = query.all()

        return jsonify({
            'success': True,
            'rules': [serialize_rule(r) for r in rules],
            'total': len(rules),
        })
    except Exception as e:
        app.logger.error(f"获取周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '获取周期账单规则失败'}), 500


@app.route('/api/recurring-rules', methods=['POST'])
@login_required
def create_recurring_rule():
    """创建周期账单规则"""
    try:
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({'success': False, 'message': '管理员不能创建规则'}), 403

        user_id = session.get('user_id')
        data = request.get_json()
        current_ledger_id = get_current_ledger_id()

        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'success': False, 'message': '账单名称不能为空'}), 400

        tx_type = data.get('type')
        if tx_type not in ('income', 'expense'):
            return jsonify({'success': False, 'message': '类型必须是 income 或 expense'}), 400

        try:
            amount = float(data.get('amount', 0))
            if amount <= 0:
                return jsonify({'success': False, 'message': '金额必须大于0'}), 400
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': '金额格式不正确'}), 400

        category = (data.get('category') or '').strip()
        if not category:
            return jsonify({'success': False, 'message': '分类不能为空'}), 400
        cat = Category.query.filter_by(name=category, type=tx_type).first()
        if not cat:
            return jsonify({'success': False, 'message': f'分类「{category}」不存在或类型不匹配'}), 400

        period = data.get('period')
        if period not in ('daily', 'weekly', 'monthly', 'yearly'):
            return jsonify({'success': False, 'message': '周期必须是 daily/weekly/monthly/yearly'}), 400

        interval_value = int(data.get('interval_value', 1))
        if interval_value < 1:
            interval_value = 1

        start_date = (data.get('start_date') or '').strip()
        if not start_date:
            start_date = datetime.now().strftime('%Y-%m-%d')
        try:
            datetime.strptime(start_date, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': '开始日期格式不正确'}), 400

        end_date = (data.get('end_date') or '').strip()
        if end_date:
            try:
                datetime.strptime(end_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'message': '结束日期格式不正确'}), 400
        else:
            end_date = None

        if end_date and end_date < start_date:
            return jsonify({'success': False, 'message': '结束日期不能早于开始日期'}), 400

        account_id = data.get('account_id')
        if account_id:
            acct = Account.query.get(account_id)
            if not acct:
                return jsonify({'success': False, 'message': '账户不存在'}), 400
        else:
            account_id = None

        remark = (data.get('remark') or '').strip()

        rule = RecurringRule(
            name=name,
            amount=amount,
            category=category,
            type=tx_type,
            period=period,
            interval_value=interval_value,
            start_date=start_date,
            end_date=end_date,
            next_date=start_date,
            is_active=True,
            remark=remark,
            account_id=account_id,
            user_id=user_id,
            ledger_id=current_ledger_id,
        )
        db.session.add(rule)
        db.session.commit()

        app.logger.info(f"创建周期账单规则: {name}, {tx_type} ¥{amount}, {period}")
        return jsonify({'success': True, 'message': '周期账单规则创建成功', 'rule': serialize_rule(rule)}), 201

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '创建失败'}), 500


@app.route('/api/recurring-rules/<int:rule_id>', methods=['PUT'])
@login_required
def update_recurring_rule(rule_id):
    """更新周期账单规则"""
    try:
        rule = RecurringRule.query.get(rule_id)
        if not rule:
            return jsonify({'success': False, 'message': '规则不存在'}), 404

        user_id = session.get('user_id')
        if rule.user_id != user_id:
            if rule.ledger_id:
                has_access, role, error = require_ledger_access(rule.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({'success': False, 'message': '无权限修改'}), 403

        data = request.get_json()

        if 'name' in data:
            name = data['name'].strip()
            if name:
                rule.name = name
        if 'amount' in data:
            try:
                amt = float(data['amount'])
                if amt > 0:
                    rule.amount = amt
            except (TypeError, ValueError):
                pass
        if 'category' in data:
            cat = data['category'].strip()
            if cat:
                rule.category = cat
        if 'type' in data and data['type'] in ('income', 'expense'):
            rule.type = data['type']
        if 'period' in data and data['period'] in ('daily', 'weekly', 'monthly', 'yearly'):
            rule.period = data['period']
        if 'interval_value' in data:
            iv = int(data['interval_value'])
            if iv >= 1:
                rule.interval_value = iv
        if 'start_date' in data:
            sd = data['start_date'].strip()
            if sd:
                rule.start_date = sd
        if 'end_date' in data:
            ed = data['end_date'].strip()
            rule.end_date = ed or None
        if 'is_active' in data:
            rule.is_active = bool(data['is_active'])
        if 'remark' in data:
            rule.remark = data['remark'].strip()
        if 'account_id' in data:
            rule.account_id = data['account_id'] or None

        rule.updated_at = datetime.now()
        db.session.commit()

        return jsonify({'success': True, 'message': '规则已更新', 'rule': serialize_rule(rule)})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '更新失败'}), 500


@app.route('/api/recurring-rules/<int:rule_id>', methods=['DELETE'])
@login_required
def delete_recurring_rule(rule_id):
    """删除周期账单规则"""
    try:
        rule = RecurringRule.query.get(rule_id)
        if not rule:
            return jsonify({'success': False, 'message': '规则不存在'}), 404

        user_id = session.get('user_id')
        if rule.user_id != user_id:
            if rule.ledger_id:
                has_access, role, error = require_ledger_access(rule.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({'success': False, 'message': '无权限删除'}), 403

        db.session.delete(rule)
        db.session.commit()

        return jsonify({'success': True, 'message': '规则已删除'})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除周期账单规则失败: {e}")
        return jsonify({'success': False, 'message': '删除失败'}), 500


@app.route('/api/recurring-rules/<int:rule_id>/toggle', methods=['POST'])
@login_required
def toggle_recurring_rule(rule_id):
    """启用/停用周期账单规则"""
    try:
        rule = RecurringRule.query.get(rule_id)
        if not rule:
            return jsonify({'success': False, 'message': '规则不存在'}), 404

        user_id = session.get('user_id')
        if rule.user_id != user_id:
            if rule.ledger_id:
                has_access, role, error = require_ledger_access(rule.ledger_id, 'editor')
                if not has_access:
                    return error
            else:
                return jsonify({'success': False, 'message': '无权限操作'}), 403

        rule.is_active = not rule.is_active
        rule.updated_at = datetime.now()
        db.session.commit()

        status = '已启用' if rule.is_active else '已停用'
        return jsonify({'success': True, 'message': f'规则{status}', 'is_active': rule.is_active})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"切换周期账单规则状态失败: {e}")
        return jsonify({'success': False, 'message': '操作失败'}), 500


@app.route('/api/recurring-rules/generate', methods=['POST'])
@login_required
def generate_recurring_bills():
    """手动触发周期账单生成"""
    try:
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()
        today_str = datetime.now().strftime('%Y-%m-%d')

        query = RecurringRule.query.filter(
            RecurringRule.is_active == True,
            RecurringRule.next_date <= today_str,
        )
        if current_ledger_id:
            query = query.filter(RecurringRule.ledger_id == current_ledger_id)
        elif not (is_admin and not self_view):
            query = query.filter(RecurringRule.user_id == user_id)

        # Also filter: end_date >= today or end_date is null
        query = query.filter(
            db.or_(RecurringRule.end_date >= today_str, RecurringRule.end_date.is_(None))
        )

        rules = query.all()
        generated = 0

        for rule in rules:
            tx = _generate_rule_transaction(rule)
            if tx:
                generated += 1

        db.session.commit()

        if generated > 0:
            app.logger.info(f"周期账单生成: 用户={user_id}, 生成={generated}笔")
            return jsonify({
                'success': True,
                'message': f'成功生成 {generated} 笔周期账单',
                'generated_count': generated,
            })
        else:
            return jsonify({
                'success': True,
                'message': '没有待生成的周期账单',
                'generated_count': 0,
            })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"生成周期账单失败: {e}")
        return jsonify({'success': False, 'message': '生成失败'}), 500


@app.route('/api/recurring-rules/generate-check', methods=['GET'])
@login_required
def check_recurring_pending():
    """检查是否有待生成的周期账单"""
    try:
        user_id = session.get('user_id')
        current_ledger_id = get_current_ledger_id()
        today_str = datetime.now().strftime('%Y-%m-%d')

        query = RecurringRule.query.filter(
            RecurringRule.is_active == True,
            RecurringRule.next_date <= today_str,
        )
        if current_ledger_id:
            query = query.filter(RecurringRule.ledger_id == current_ledger_id)
        else:
            query = query.filter(RecurringRule.user_id == user_id)

        query = query.filter(
            db.or_(RecurringRule.end_date >= today_str, RecurringRule.end_date.is_(None))
        )

        count = query.count()
        return jsonify({'success': True, 'pending_count': count})

    except Exception as e:
        app.logger.error(f"检查周期账单待生成失败: {e}")
        return jsonify({'success': False, 'message': '检查失败'}), 500


@app.route('/api/recurring-rules/scheduled-generate', methods=['POST'])
def scheduled_generate():
    """定时任务自动生成（允许未登录调用，但只处理全局规则）"""
    try:
        from datetime import date
        today_str = date.today().strftime('%Y-%m-%d')

        rules = RecurringRule.query.filter(
            RecurringRule.is_active == True,
            RecurringRule.next_date <= today_str,
            db.or_(RecurringRule.end_date >= today_str, RecurringRule.end_date.is_(None))
        ).all()

        generated = 0
        for rule in rules:
            tx = _generate_rule_transaction(rule)
            if tx:
                generated += 1

        db.session.commit()
        app.logger.info(f"[定时] 周期账单生成: {generated}笔")
        return jsonify({'success': True, 'generated_count': generated})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"[定时] 周期账单生成失败: {e}")
        return jsonify({'success': False, 'message': '生成失败'}), 500


@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    try:
        cat_type = request.args.get('type', None)
        query = Category.query
        if cat_type and cat_type in ['income', 'expense']:
            query = query.filter_by(type=cat_type)
        categories = query.all()
        if not categories:
            # 如无分类自动写入默认分类
            default_income = ['工资', '奖金', '投资收益', '兼职', '红包', '报销收入', '其他收入']
            default_expense = ['餐饮', '交通', '购物', '娱乐', '医疗', '住房', '教育', '通讯', '其他支出']
            for cname in default_income:
                db.session.add(Category(name=cname, type='income'))
            for cname in default_expense:
                db.session.add(Category(name=cname, type='expense'))
            db.session.commit()
            categories = Category.query.all()
        return jsonify({
            "success": True,
            "categories": [{"id": c.id, "name": c.name, "type": c.type} for c in categories]
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"获取分类失败: {e}")
        abort(500, description="获取分类失败")


# ==================== 预算管理 API ====================
@app.route('/api/budgets/current', methods=['GET'])
@login_required
def get_current_budget():
    """获取当前月份的预算及使用情况"""
    try:
        user_id = session.get('user_id')
        now = datetime.now()
        month = now.strftime('%Y-%m')
        account_id = request.args.get('account_id', None, type=int)

        # 查找当月预算（用 ledger 或 user 过滤）
        current_ledger_id = get_current_ledger_id()
        budget_filter = {'user_id': user_id, 'month': month}
        if current_ledger_id:
            budget_filter['ledger_id'] = current_ledger_id
        if account_id:
            budget_filter['account_id'] = account_id
        else:
            budget_filter['account_id'] = None
        budget = Budget.query.filter_by(**budget_filter).first()

        # 计算当月支出（按分类）
        expense_filter = [
            Transaction.type == 'expense',
            Transaction.date.startswith(month)
        ]
        if current_ledger_id:
            expense_filter.append(Transaction.ledger_id == current_ledger_id)
        else:
            expense_filter.append(Transaction.user_id == user_id)
        if account_id:
            expense_filter.append(Transaction.account_id == account_id)

        expenses = Transaction.query.filter(*expense_filter).all()

        total_expense = float(sum(t.amount for t in expenses))

        # 按分类统计支出
        category_expenses = {}
        for t in expenses:
            cat = t.category
            category_expenses[cat] = category_expenses.get(cat, 0) + float(t.amount)

        # 账户名称
        account_name = None
        if account_id:
            acct = Account.query.get(account_id)
            account_name = acct.name if acct else None

        result = {
            'month': month,
            'account_id': account_id,
            'account_name': account_name or '总账户',
            'total_expense': round(total_expense, 2),
            'category_expenses': {k: round(v, 2) for k, v in category_expenses.items()}
        }

        if budget:
            total_amount = float(budget.total_amount)
            remaining = round(total_amount - total_expense, 2)
            progress = round(min(100, total_expense / total_amount * 100), 1) if total_amount > 0 else 0

            cat_items = BudgetCategoryItem.query.filter_by(budget_id=budget.id).all()
            items_data = []
            for item in cat_items:
                cat_name = item.category.name if item.category else '未知'
                used = round(category_expenses.get(cat_name, 0), 2)
                item_amount = float(item.amount)
                items_data.append({
                    'id': item.id,
                    'category_id': item.category_id,
                    'category_name': cat_name,
                    'amount': item_amount,
                    'used': used,
                    'remaining': round(item_amount - used, 2),
                    'progress': round(min(100, used / item_amount * 100), 1) if item_amount > 0 else 0
                })

            result['budget'] = {
                'id': budget.id,
                'account_id': budget.account_id,
                'total_amount': total_amount,
                'remaining': remaining,
                'progress': progress,
                'is_over': total_expense > total_amount,
                'remark': budget.remark or '',
                'category_items': items_data
            }
        else:
            result['budget'] = None

        return jsonify({'success': True, 'data': result})
    except Exception as e:
        app.logger.error(f"获取预算失败: {e}")
        return jsonify({'success': False, 'message': '获取预算失败'}), 500


@app.route('/api/budgets/list', methods=['GET'])
@login_required
def list_budgets():
    """获取当前用户所有月预算列表（含账户信息）"""
    try:
        user_id = session.get('user_id')
        now = datetime.now()
        month = now.strftime('%Y-%m')
        current_ledger_id = get_current_ledger_id()
        budget_filter = {'user_id': user_id, 'month': month}
        if current_ledger_id:
            budget_filter['ledger_id'] = current_ledger_id
        budgets = Budget.query.filter_by(**budget_filter).all()
        result = []
        for b in budgets:
            acct_name = b.account.name if b.account else '总账户'
            result.append({
                'id': b.id,
                'account_id': b.account_id,
                'account_name': acct_name,
                'total_amount': float(b.total_amount),
                'remark': b.remark or ''
            })
        return jsonify({'success': True, 'budgets': result})
    except Exception as e:
        app.logger.error(f"获取预算列表失败: {e}")
        return jsonify({'success': False, 'message': '获取预算列表失败'}), 500


@app.route('/api/budgets', methods=['POST'])
@login_required
def save_budget():
    """创建或更新当月预算"""
    try:
        # 管理员在 self_view 模式下可以管理预算
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({'success': False, 'message': '管理员只能查看数据，不能管理预算'}), 403

        user_id = session.get('user_id')
        data = request.get_json()

        month = data.get('month', datetime.now().strftime('%Y-%m'))
        total_amount = data.get('total_amount', 0)
        remark = data.get('remark', '').strip()
        account_id = data.get('account_id', None)
        category_items = data.get('category_items', [])
        current_ledger_id = get_current_ledger_id()

        # 校验账户（account_id 为 0 或空时视为总账户）
        if account_id:
            acct = Account.query.filter_by(id=account_id).first()
            if not acct or (current_ledger_id and acct.ledger_id != current_ledger_id):
                return jsonify({'success': False, 'message': '账户不存在'}), 400
        else:
            account_id = None

        try:
            total_amount = float(total_amount)
            if total_amount <= 0:
                return jsonify({'success': False, 'message': '预算金额必须大于0'}), 400
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': '预算金额格式不正确'}), 400

        # 查找是否已有该月该账户的预算
        budget_filter = {'user_id': user_id, 'month': month, 'account_id': account_id}
        if current_ledger_id:
            budget_filter['ledger_id'] = current_ledger_id
        budget = Budget.query.filter_by(**budget_filter).first()
        if budget:
            budget.total_amount = total_amount
            budget.remark = remark
            budget.updated_at = datetime.now()
        else:
            budget = Budget(user_id=user_id, month=month, total_amount=total_amount, remark=remark, account_id=account_id, ledger_id=current_ledger_id)
            db.session.add(budget)

        db.session.flush()

        # 删除旧的分类预算项
        BudgetCategoryItem.query.filter_by(budget_id=budget.id).delete()

        # 添加新的分类预算项
        for item in category_items:
            cat_id = item.get('category_id')
            cat_amount = item.get('amount', 0)
            if not cat_id or not cat_amount:
                continue
            try:
                cat_amount = float(cat_amount)
                if cat_amount <= 0:
                    continue
            except (TypeError, ValueError):
                continue

            cat_item = BudgetCategoryItem(
                budget_id=budget.id,
                category_id=cat_id,
                amount=cat_amount
            )
            db.session.add(cat_item)

        db.session.commit()

        acct_label = account_id or '总账户'
        app.logger.info(f"保存预算: user={user_id}, month={month}, account={acct_label}, total={total_amount}")
        return jsonify({'success': True, 'message': '预算保存成功', 'budget_id': budget.id, 'account_id': account_id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"保存预算失败: {e}")
        return jsonify({'success': False, 'message': '保存预算失败'}), 500


@app.route('/api/budgets/<int:budget_id>', methods=['DELETE'])
@login_required
def delete_budget(budget_id):
    """删除预算"""
    try:
        # 管理员在 self_view 模式下可以删除预算
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({'success': False, 'message': '管理员只能查看数据，不能删除预算'}), 403

        user_id = session.get('user_id')
        budget = Budget.query.filter_by(id=budget_id, user_id=user_id).first()
        if not budget:
            return jsonify({'success': False, 'message': '未找到该预算'}), 404

        # 删除关联的分类预算项
        BudgetCategoryItem.query.filter_by(budget_id=budget.id).delete()
        db.session.delete(budget)
        db.session.commit()

        app.logger.info(f"删除预算: user={user_id}, budget_id={budget_id}")
        return jsonify({'success': True, 'message': '预算已删除'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除预算失败: {e}")
        return jsonify({'success': False, 'message': '删除预算失败'}), 500


# ==================== 账户管理 API ====================
@app.route('/api/accounts', methods=['GET'])
@login_required
def get_accounts():
    try:
        user_id = session.get('user_id')
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            accounts = Account.query.filter_by(ledger_id=current_ledger_id).all()
        else:
            accounts = Account.query.filter_by(user_id=user_id).all()
        if not accounts:
            if current_ledger_id:
                default_account = Account(name='默认账户', balance=0, account_type='cash', user_id=user_id, ledger_id=current_ledger_id)
            else:
                default_account = Account(name='默认账户', balance=0, account_type='cash', user_id=user_id)
            db.session.add(default_account)
            db.session.commit()
            accounts = [default_account]
        return jsonify({
            "success": True,
            "accounts": [
                {
                    "id": a.id,
                    "name": a.name,
                    "balance": float(a.balance),
                    "account_type": a.account_type,
                    "created_at": a.created_at.isoformat() if a.created_at else None
                } for a in accounts
            ]
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"获取账户失败: {e}")
        abort(500, description="获取账户失败")


@app.route('/api/accounts', methods=['POST'])
@login_required
def create_account():
    try:
        # 管理员在 self_view 模式下可以创建账户
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({"success": False, "message": "管理员只能查看数据，不能创建账户"}), 403

        data = request.get_json()
        name = data.get('name', '').strip()
        balance = data.get('balance', 0)
        account_type = data.get('account_type', 'cash')
        user_id = session.get('user_id')
        current_ledger_id = get_current_ledger_id()

        if not name:
            return jsonify({"success": False, "message": "账户名称不能为空"}), 400
        if account_type not in ['cash', 'bank', 'credit', 'other']:
            return jsonify({"success": False, "message": "账户类型无效"}), 400

        account = Account(name=name, balance=balance, account_type=account_type, user_id=user_id, ledger_id=current_ledger_id)
        db.session.add(account)
        db.session.commit()
        return jsonify({
            "success": True,
            "message": "账户创建成功",
            "account": {
                "id": account.id,
                "name": account.name,
                "balance": float(account.balance),
                "account_type": account.account_type
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建账户失败: {e}")
        return jsonify({"success": False, "message": "创建账户失败"}), 500


@app.route('/api/accounts/<int:account_id>', methods=['PUT'])
@login_required
def update_account(account_id):
    try:
        # 管理员在 self_view 模式下可以修改账户
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({"success": False, "message": "管理员只能查看数据，不能修改账户"}), 403

        user_id = session.get('user_id')
        account = Account.query.filter_by(id=account_id, user_id=user_id).first()
        if not account:
            return jsonify({"success": False, "message": "账户不存在"}), 404

        data = request.get_json()
        if 'name' in data:
            account.name = data['name'].strip()
        if 'account_type' in data and data['account_type'] in ['cash', 'bank', 'credit', 'other']:
            account.account_type = data['account_type']
        if 'balance' in data:
            account.balance = data['balance']

        db.session.commit()
        return jsonify({
            "success": True,
            "message": "账户更新成功",
            "account": {
                "id": account.id,
                "name": account.name,
                "balance": float(account.balance),
                "account_type": account.account_type
            }
        })
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新账户失败: {e}")
        return jsonify({"success": False, "message": "更新账户失败"}), 500


@app.route('/api/accounts/<int:account_id>', methods=['DELETE'])
@login_required
def delete_account(account_id):
    try:
        # 管理员在 self_view 模式下可以删除账户
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({"success": False, "message": "管理员只能查看数据，不能删除账户"}), 403

        user_id = session.get('user_id')
        account = Account.query.filter_by(id=account_id, user_id=user_id).first()
        if not account:
            return jsonify({"success": False, "message": "账户不存在"}), 404

        if account.transactions.count() > 0:
            return jsonify({"success": False, "message": "该账户存在交易记录，无法删除"}), 400

        db.session.delete(account)
        db.session.commit()
        return jsonify({"success": True, "message": "账户删除成功"})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除账户失败: {e}")
        return jsonify({"success": False, "message": "删除账户失败"}), 500

@app.route('/api/reports/<period>', methods=['GET'])
@login_required
def get_report(period):
    try:
        transactions = filter_transactions_by_period_orm(period).all()
        income = float(sum(t.amount for t in transactions if t.type == 'income'))
        expense = float(sum(t.amount for t in transactions if t.type == 'expense'))
        return jsonify({
            "success": True,
            "period": period,
            "transactions": [
                {
                    "id": t.id,
                    "type": t.type,
                    "amount": float(t.amount) if t.amount else 0,
                    "category": t.category,
                    "date": t.date,
                    "time": t.time,
                    "remark": t.remark,
                    "account_id": t.account_id,
                    "currency": t.currency or 'CNY',
                    "original_amount": float(t.original_amount) if t.original_amount else None,
                    "exchange_rate": float(t.exchange_rate) if t.exchange_rate else None,
                    "created_at": t.created_at.isoformat() if t.created_at else None,
                    "reimbursement_status": t.reimbursement_status or 'none',
                    "reimbursed_amount": float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                    "write_off_id": t.write_off_id
                } for t in transactions
            ],
            "income": income,
            "expense": expense,
            "net": income - expense,
            "count": len(transactions)
        })
    except Exception as e:
        app.logger.error(f"获取报表失败: {e}")
        abort(500, description="获取报表失败")


@app.route('/api/reports/advanced', methods=['GET'])
@login_required
def get_advanced_report():
    try:
        period = request.args.get('period', 'month')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        now = datetime.now()

        # 构建日期范围
        if period == 'week':
            start = (now - timedelta(days=7)).strftime('%Y-%m-%d')
            end = now.strftime('%Y-%m-%d')
        elif period == 'month':
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')
        elif period == 'quarter':
            current_quarter = (now.month - 1) // 3 + 1
            start_month = (current_quarter - 1) * 3 + 1
            start = f'{now.year}-{start_month:02d}-01'
            if current_quarter == 1:
                end_month = 3
            elif current_quarter == 2:
                end_month = 6
            elif current_quarter == 3:
                end_month = 9
            else:
                end_month = 12
            last_day = calendar.monthrange(now.year, end_month)[1]
            end = f'{now.year}-{end_month:02d}-{last_day}'
        elif period == 'year':
            start = f'{now.year}-01-01'
            end = now.strftime('%Y-%m-%d')
        elif period == 'custom':
            start = start_date if start_date else (now - timedelta(days=30)).strftime('%Y-%m-%d')
            end = end_date if end_date else now.strftime('%Y-%m-%d')
        else:
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')

        # 查询数据（带权限过滤）
        query = Transaction.query.filter(
            Transaction.date >= start,
            Transaction.date <= end
        )

        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            query = query.filter(Transaction.ledger_id == current_ledger_id)
        elif user_id and not is_admin or (is_admin and self_view):
            # 普通用户或管理员 self_view 只看到自己的数据
            query = query.filter(Transaction.user_id == user_id)

        transactions = query.order_by(Transaction.date.desc(), Transaction.id.desc()).all()
        income = float(sum(t.amount for t in transactions if t.type == 'income'))
        expense = float(sum(t.amount for t in transactions if t.type == 'expense'))
        count = len(transactions)

        # 分类占比统计（分别统计收入和支出）
        category_stats = {}
        for t in transactions:
            key = f"{t.type}:{t.category}"
            if key not in category_stats:
                category_stats[key] = {'type': t.type, 'category': t.category, 'amount': 0, 'count': 0}
            category_stats[key]['amount'] += float(t.amount)
            category_stats[key]['count'] += 1

        # 分别提取收入和支出分类统计
        expense_category_stats = [v for v in category_stats.values() if v['type'] == 'expense']
        income_category_stats = [v for v in category_stats.values() if v['type'] == 'income']

        # 查找支出/收入最多的分类
        expense_category_stats_sorted = sorted(expense_category_stats, key=lambda x: x['amount'], reverse=True)
        income_category_stats_sorted = sorted(income_category_stats, key=lambda x: x['amount'], reverse=True)

        # 计算天数
        date_diff = max(1, (datetime.strptime(end, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')).days + 1)

        # 每日趋势（按日期汇总）
        daily_trend = {}
        for t in transactions:
            if t.date not in daily_trend:
                daily_trend[t.date] = {'income': 0, 'expense': 0, 'count': 0}
            if t.type == 'income':
                daily_trend[t.date]['income'] += float(t.amount)
            else:
                daily_trend[t.date]['expense'] += float(t.amount)
            daily_trend[t.date]['count'] += 1

        return jsonify({
            "success": True,
            "period": period,
            "start_date": start,
            "end_date": end,
            "summary": {
                "income": income,
                "expense": expense,
                "net": income - expense,
                "count": count,
                "avg_daily_expense": round(expense / date_diff, 2) if expense > 0 else 0,
                "avg_daily_income": round(income / date_diff, 2) if income > 0 else 0,
                "days": date_diff
            },
            "top_expense_category": expense_category_stats_sorted[0] if expense_category_stats_sorted else None,
            "top_income_category": income_category_stats_sorted[0] if income_category_stats_sorted else None,
            "category_stats": list(category_stats.values()),
            "income_category_stats": income_category_stats,
            "expense_category_stats": expense_category_stats,
            "daily_trend": [
                {"date": d, "income": v['income'], "expense": v['expense'], "count": v['count']}
                for d, v in sorted(daily_trend.items())
            ],
            "transactions": [
                {
                    "id": t.id,
                    "type": t.type,
                    "amount": float(t.amount) if t.amount else 0,
                    "category": t.category,
                    "date": t.date,
                    "time": t.time,
                    "remark": t.remark,
                    "account_id": t.account_id,
                    "currency": t.currency or 'CNY',
                    "original_amount": float(t.original_amount) if t.original_amount else None,
                    "exchange_rate": float(t.exchange_rate) if t.exchange_rate else None,
                    "reimbursement_status": t.reimbursement_status or 'none',
                    "reimbursed_amount": float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                    "write_off_id": t.write_off_id
                } for t in transactions[:50]  # 最多返回50条明细
            ]
        })
    except Exception as e:
        app.logger.error(f"获取高级报表失败: {e}")
        return jsonify({"success": False, "message": "获取报表失败"}), 500


@app.route('/api/reports/download', methods=['GET'])
@login_required
def download_report():
    """从报表页面导出统计汇总数据（非明细流水）"""
    try:
        period = request.args.get('period', 'month')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        file_format = request.args.get('format', 'csv').lower()

        if file_format not in ('csv', 'xlsx'):
            return jsonify({"success": False, "message": "不支持的导出格式"}), 400

        now = datetime.now()

        # 构建日期范围（与 advanced report 一致）
        if period == 'week':
            start = (now - timedelta(days=7)).strftime('%Y-%m-%d')
            end = now.strftime('%Y-%m-%d')
        elif period == 'month':
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')
        elif period == 'quarter':
            current_quarter = (now.month - 1) // 3 + 1
            start_month = (current_quarter - 1) * 3 + 1
            start = f'{now.year}-{start_month:02d}-01'
            last_day = calendar.monthrange(now.year, start_month + 2)[1]
            end = f'{now.year}-{start_month + 2:02d}-{last_day}'
        elif period == 'year':
            start = f'{now.year}-01-01'
            end = now.strftime('%Y-%m-%d')
        elif period == 'custom':
            start = start_date if start_date else (now - timedelta(days=30)).strftime('%Y-%m-%d')
            end = end_date if end_date else now.strftime('%Y-%m-%d')
        else:
            start = now.strftime('%Y-%m-01')
            end = now.strftime('%Y-%m-%d')

        # 查询数据（带权限过滤）
        query = Transaction.query.filter(
            Transaction.date >= start,
            Transaction.date <= end
        )

        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()
        if current_ledger_id:
            query = query.filter(Transaction.ledger_id == current_ledger_id)
        elif user_id and not is_admin or (is_admin and self_view):
            query = query.filter(Transaction.user_id == user_id)

        transactions = query.order_by(Transaction.date.asc(), Transaction.id.asc()).all()

        if not transactions:
            return jsonify({"success": False, "message": "该时间段内没有交易记录"}), 404

        # ---- 计算统计汇总数据 ----
        income = float(sum(t.amount for t in transactions if t.type == 'income'))
        expense = float(sum(t.amount for t in transactions if t.type == 'expense'))
        count = len(transactions)
        date_diff = max(1, (datetime.strptime(end, '%Y-%m-%d') - datetime.strptime(start, '%Y-%m-%d')).days + 1)

        # 每日趋势
        daily_trend = {}
        for t in transactions:
            if t.date not in daily_trend:
                daily_trend[t.date] = {'income': 0, 'expense': 0, 'count': 0}
            if t.type == 'income':
                daily_trend[t.date]['income'] += float(t.amount)
            else:
                daily_trend[t.date]['expense'] += float(t.amount)
            daily_trend[t.date]['count'] += 1

        daily_rows = [
            {'date': d, 'income': v['income'], 'expense': v['expense'],
             'net': round(v['income'] - v['expense'], 2), 'count': v['count']}
            for d, v in sorted(daily_trend.items())
        ]

        # 分类统计
        category_stats = {}
        for t in transactions:
            key = f"{t.type}:{t.category}"
            if key not in category_stats:
                category_stats[key] = {'category': t.category, 'type': t.type, 'amount': 0, 'count': 0}
            category_stats[key]['amount'] += float(t.amount)
            category_stats[key]['count'] += 1

        category_rows = sorted(category_stats.values(), key=lambda x: x['amount'], reverse=True)

        timestamp = now.strftime('%Y%m%d_%H%M%S')

        if file_format == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)

            # 第1段：汇总统计
            writer.writerow(['=== 汇总统计 ==='])
            writer.writerow(['指标', '值'])
            writer.writerow(['统计周期', f'{start} ~ {end}'])
            writer.writerow(['统计天数', date_diff])
            writer.writerow(['总收入', income])
            writer.writerow(['总支出', expense])
            writer.writerow(['净收支', round(income - expense, 2)])
            writer.writerow(['交易总笔数', count])
            writer.writerow(['日均收入', round(income / date_diff, 2)])
            writer.writerow(['日均支出', round(expense / date_diff, 2)])
            writer.writerow([])

            # 第2段：每日趋势
            writer.writerow(['=== 每日收支趋势 ==='])
            writer.writerow(['日期', '收入', '支出', '净收支', '笔数'])
            for row in daily_rows:
                writer.writerow([row['date'], row['income'], row['expense'], row['net'], row['count']])
            writer.writerow([])

            # 第3段：分类统计
            writer.writerow(['=== 分类统计 ==='])
            writer.writerow(['分类', '类型', '金额', '笔数'])
            for row in category_rows:
                type_label = '收入' if row['type'] == 'income' else '支出'
                writer.writerow([row['category'], type_label, round(row['amount'], 2), row['count']])

            csv_bytes = output.getvalue().encode('utf-8-sig')
            return send_file(
                io.BytesIO(csv_bytes),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'报表汇总_{start}_{end}.csv'
            )
        else:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                return jsonify({"success": False, "message": "导出 XLSX 需要安装 openpyxl"}), 500

            wb = openpyxl.Workbook()
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            label_font = Font(bold=True, size=11)

            def style_header(ws, headers):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

            def auto_width(ws):
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

            # Sheet 1: 汇总
            ws1 = wb.active
            ws1.title = '汇总'
            ws1.cell(row=1, column=1, value='指标').font = label_font
            ws1.cell(row=1, column=2, value='值').font = label_font
            ws1.cell(row=1, column=1).border = thin_border
            ws1.cell(row=1, column=2).border = thin_border
            summary_data = [
                ('统计周期', f'{start} ~ {end}'),
                ('统计天数', date_diff),
                ('总收入', income),
                ('总支出', expense),
                ('净收支', round(income - expense, 2)),
                ('交易总笔数', count),
                ('日均收入', round(income / date_diff, 2)),
                ('日均支出', round(expense / date_diff, 2)),
            ]
            for i, (k, v) in enumerate(summary_data, 2):
                ws1.cell(row=i, column=1, value=k).border = thin_border
                ws1.cell(row=i, column=2, value=v).border = thin_border
            ws1.column_dimensions['A'].width = 14
            ws1.column_dimensions['B'].width = 30

            # Sheet 2: 每日趋势
            ws2 = wb.create_sheet('每日趋势')
            style_header(ws2, ['日期', '收入', '支出', '净收支', '笔数'])
            for i, row in enumerate(daily_rows, 2):
                for j, val in enumerate([row['date'], row['income'], row['expense'], row['net'], row['count']], 1):
                    cell = ws2.cell(row=i, column=j, value=val)
                    cell.border = thin_border
            auto_width(ws2)
            ws2.freeze_panes = 'A2'

            # Sheet 3: 分类统计
            ws3 = wb.create_sheet('分类统计')
            style_header(ws3, ['分类', '类型', '金额', '笔数'])
            for i, row in enumerate(category_rows, 2):
                type_label = '收入' if row['type'] == 'income' else '支出'
                for j, val in enumerate([row['category'], type_label, round(row['amount'], 2), row['count']], 1):
                    cell = ws3.cell(row=i, column=j, value=val)
                    cell.border = thin_border
            auto_width(ws3)
            ws3.freeze_panes = 'A2'

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(tmp.name)
            tmp.close()

            return send_file(
                tmp.name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'报表汇总_{start}_{end}.xlsx'
            )
    except Exception as e:
        app.logger.error(f"导出报表失败: {e}")
        return jsonify({"success": False, "message": "导出失败"}), 500


# ==================== 账本管理 API ====================
import secrets


@app.route('/api/ledgers', methods=['GET'])
@login_required
def list_ledgers():
    """列出当前用户有权限的所有账本"""
    try:
        user_id = session.get('user_id')
        owned = Ledger.query.filter_by(owner_id=user_id, is_active=True).all()
        member_ledger_ids = [m.ledger_id for m in LedgerMember.query.filter_by(user_id=user_id).all()]
        member_ledgers = Ledger.query.filter(Ledger.id.in_(member_ledger_ids), Ledger.is_active == True).all() if member_ledger_ids else []
        seen = set()
        result = []
        for ledger in owned + member_ledgers:
            if ledger.id not in seen:
                seen.add(ledger.id)
                role = get_user_ledger_role(ledger.id, user_id)
                member_count = LedgerMember.query.filter_by(ledger_id=ledger.id).count()
                result.append({
                    'id': ledger.id,
                    'name': ledger.name,
                    'description': ledger.description,
                    'currency': ledger.currency,
                    'owner_id': ledger.owner_id,
                    'owner_name': ledger.owner.username if ledger.owner else None,
                    'role': role,
                    'member_count': member_count + 1,  # +1 for owner
                    'created_at': ledger.created_at.isoformat() if ledger.created_at else None,
                })
        return jsonify({'success': True, 'ledgers': result})
    except Exception as e:
        app.logger.error(f"获取账本列表失败: {e}")
        return jsonify({'success': False, 'message': '获取账本列表失败'}), 500


@app.route('/api/ledgers', methods=['POST'])
@login_required
def create_ledger():
    """创建新账本"""
    try:
        if session.get('is_admin') and not session.get('self_view'):
            return jsonify({"success": False, "message": "管理员在此模式下不能创建账本"}), 403
        data = request.get_json()
        name = data.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'message': '账本名称不能为空'}), 400
        user_id = session.get('user_id')
        ledger = Ledger(name=name, description=data.get('description', ''), owner_id=user_id, currency=data.get('currency', 'CNY'))
        db.session.add(ledger)
        db.session.flush()
        member = LedgerMember(ledger_id=ledger.id, user_id=user_id, role='manager')
        db.session.add(member)
        db.session.commit()
        return jsonify({'success': True, 'message': '账本创建成功', 'ledger': {
            'id': ledger.id, 'name': ledger.name, 'description': ledger.description,
            'currency': ledger.currency, 'owner_id': ledger.owner_id, 'role': 'manager'
        }}), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建账本失败: {e}")
        return jsonify({'success': False, 'message': '创建账本失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>', methods=['GET'])
@login_required
def get_ledger(ledger_id):
    """获取账本详情"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'viewer')
        if not has_access:
            return error
        ledger = Ledger.query.get_or_404(ledger_id)
        member_count = LedgerMember.query.filter_by(ledger_id=ledger.id).count()
        return jsonify({'success': True, 'ledger': {
            'id': ledger.id, 'name': ledger.name, 'description': ledger.description,
            'currency': ledger.currency, 'owner_id': ledger.owner_id,
            'owner_name': ledger.owner.username if ledger.owner else None,
            'role': role, 'member_count': member_count + 1,
            'created_at': ledger.created_at.isoformat() if ledger.created_at else None,
        }})
    except Exception as e:
        app.logger.error(f"获取账本详情失败: {e}")
        return jsonify({'success': False, 'message': '获取账本详情失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>', methods=['PUT'])
@login_required
def update_ledger(ledger_id):
    """更新账本信息"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'manager')
        if not has_access:
            return error
        ledger = Ledger.query.get_or_404(ledger_id)
        data = request.get_json()
        if 'name' in data:
            ledger.name = data['name'].strip()
        if 'description' in data:
            ledger.description = data['description']
        if 'currency' in data:
            ledger.currency = data['currency']
        db.session.commit()
        return jsonify({'success': True, 'message': '账本更新成功'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新账本失败: {e}")
        return jsonify({'success': False, 'message': '更新账本失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>', methods=['DELETE'])
@login_required
def delete_ledger(ledger_id):
    """软删除账本"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'manager')
        if not has_access:
            return error
        ledger = Ledger.query.get_or_404(ledger_id)
        if ledger.owner_id != session.get('user_id'):
            return jsonify({'success': False, 'message': '只有账本所有者才能删除'}), 403
        ledger.is_active = False
        db.session.commit()
        return jsonify({'success': True, 'message': '账本已删除'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除账本失败: {e}")
        return jsonify({'success': False, 'message': '删除账本失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>/switch', methods=['POST'])
@login_required
def switch_ledger(ledger_id):
    """切换当前活动账本"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'viewer')
        if not has_access:
            return error
        session['active_ledger_id'] = ledger_id
        return jsonify({'success': True, 'message': '已切换账本', 'role': role})
    except Exception as e:
        app.logger.error(f"切换账本失败: {e}")
        return jsonify({'success': False, 'message': '切换账本失败'}), 500


# ==================== 账本成员管理 API ====================

@app.route('/api/ledgers/<int:ledger_id>/members', methods=['GET'])
@login_required
def list_members(ledger_id):
    """列出账本所有成员"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'viewer')
        if not has_access:
            return error
        ledger = Ledger.query.get_or_404(ledger_id)
        owner_info = {'user_id': ledger.owner_id, 'username': ledger.owner.username if ledger.owner else '未知', 'role': 'owner'}
        members = LedgerMember.query.filter_by(ledger_id=ledger_id).all()
        member_list = [{
            'id': m.id, 'user_id': m.user_id, 'username': m.user.username if m.user else '未知',
            'role': m.role, 'joined_at': m.joined_at.isoformat() if m.joined_at else None
        } for m in members]
        return jsonify({'success': True, 'owner': owner_info, 'members': member_list})
    except Exception as e:
        app.logger.error(f"获取成员列表失败: {e}")
        return jsonify({'success': False, 'message': '获取成员列表失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>/members', methods=['POST'])
@login_required
def add_member(ledger_id):
    """添加成员到账本（需 manager 权限）"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'manager')
        if not has_access:
            return error
        data = request.get_json()
        username = data.get('username', '').strip()
        new_role = data.get('role', 'viewer')
        if not username:
            return jsonify({'success': False, 'message': '用户名不能为空'}), 400
        if new_role not in ('viewer', 'editor', 'manager'):
            return jsonify({'success': False, 'message': '角色无效，可选: viewer, editor, manager'}), 400
        user = User.query.filter_by(username=username).first()
        if not user:
            return jsonify({'success': False, 'message': '用户不存在'}), 404
        existing = LedgerMember.query.filter_by(ledger_id=ledger_id, user_id=user.id).first()
        if existing:
            return jsonify({'success': False, 'message': '该用户已经是成员'}), 400
        member = LedgerMember(ledger_id=ledger_id, user_id=user.id, role=new_role)
        db.session.add(member)
        db.session.commit()
        return jsonify({'success': True, 'message': f'成员 {username} 添加成功'}), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"添加成员失败: {e}")
        return jsonify({'success': False, 'message': '添加成员失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>/members/<int:user_id>', methods=['PUT'])
@login_required
def update_member_role(ledger_id, user_id):
    """修改成员角色"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'manager')
        if not has_access:
            return error
        ledger = Ledger.query.get_or_404(ledger_id)
        if ledger.owner_id == user_id:
            return jsonify({'success': False, 'message': '不能修改所有者的角色'}), 400
        member = LedgerMember.query.filter_by(ledger_id=ledger_id, user_id=user_id).first_or_404()
        data = request.get_json()
        new_role = data.get('role', 'viewer')
        if new_role not in ('viewer', 'editor', 'manager'):
            return jsonify({'success': False, 'message': '角色无效'}), 400
        member.role = new_role
        db.session.commit()
        return jsonify({'success': True, 'message': '角色更新成功'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新成员角色失败: {e}")
        return jsonify({'success': False, 'message': '更新角色失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>/members/<int:user_id>', methods=['DELETE'])
@login_required
def remove_member(ledger_id, user_id):
    """移除成员"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'manager')
        if not has_access:
            return error
        ledger = Ledger.query.get_or_404(ledger_id)
        if ledger.owner_id == user_id:
            return jsonify({'success': False, 'message': '不能移除所有者'}), 400
        member = LedgerMember.query.filter_by(ledger_id=ledger_id, user_id=user_id).first_or_404()
        db.session.delete(member)
        db.session.commit()
        return jsonify({'success': True, 'message': '成员已移除'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"移除成员失败: {e}")
        return jsonify({'success': False, 'message': '移除成员失败'}), 500


# ==================== 邀请码管理 API ====================

@app.route('/api/ledgers/<int:ledger_id>/invite-codes', methods=['GET'])
@login_required
def list_invite_codes(ledger_id):
    """列出账本的邀请码"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'viewer')
        if not has_access:
            return error
        codes = InviteCode.query.filter_by(ledger_id=ledger_id).order_by(InviteCode.created_at.desc()).all()
        return jsonify({'success': True, 'invite_codes': [{
            'id': c.id, 'code': c.code, 'created_by': c.created_by,
            'creator_name': c.creator.username if c.creator else None,
            'max_uses': c.max_uses, 'used_count': c.used_count,
            'expires_at': c.expires_at.isoformat() if c.expires_at else None,
            'is_active': c.is_active, 'created_at': c.created_at.isoformat() if c.created_at else None
        } for c in codes]})
    except Exception as e:
        app.logger.error(f"获取邀请码列表失败: {e}")
        return jsonify({'success': False, 'message': '获取邀请码列表失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>/invite-codes', methods=['POST'])
@login_required
def create_invite_code(ledger_id):
    """生成邀请码（需 manager 权限）"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'manager')
        if not has_access:
            return error
        data = request.get_json()
        code = secrets.token_urlsafe(16)
        max_uses = data.get('max_uses', 0)
        expires_in_hours = data.get('expires_in_hours', 0)
        expire_at = None
        if expires_in_hours > 0:
            expire_at = datetime.now() + timedelta(hours=expires_in_hours)
        invite = InviteCode(ledger_id=ledger_id, code=code, created_by=session.get('user_id'), max_uses=max_uses, expires_at=expire_at)
        db.session.add(invite)
        db.session.commit()
        return jsonify({'success': True, 'message': '邀请码创建成功', 'invite_code': {
            'id': invite.id, 'code': invite.code, 'max_uses': invite.max_uses,
            'expires_at': expire_at.isoformat() if expire_at else None
        }}), 201
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"创建邀请码失败: {e}")
        return jsonify({'success': False, 'message': '创建邀请码失败'}), 500


@app.route('/api/ledgers/<int:ledger_id>/invite-codes/<int:code_id>', methods=['DELETE'])
@login_required
def revoke_invite_code(ledger_id, code_id):
    """撤销邀请码"""
    try:
        has_access, role, error = require_ledger_access(ledger_id, 'manager')
        if not has_access:
            return error
        code = InviteCode.query.filter_by(id=code_id, ledger_id=ledger_id).first_or_404()
        code.is_active = False
        db.session.commit()
        return jsonify({'success': True, 'message': '邀请码已撤销'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"撤销邀请码失败: {e}")
        return jsonify({'success': False, 'message': '撤销邀请码失败'}), 500


@app.route('/api/ledgers/validate-code', methods=['POST'])
@login_required
def validate_invite_code():
    """校验邀请码是否有效（不加入账本）"""
    try:
        data = request.get_json()
        code_str = data.get('code', '').strip()
        if not code_str:
            return jsonify({'valid': False, 'message': '邀请码不能为空'})
        invite = InviteCode.query.filter_by(code=code_str, is_active=True).first()
        if not invite:
            return jsonify({'valid': False, 'message': '邀请码无效或已失效'})
        if invite.expires_at and invite.expires_at < datetime.now():
            invite.is_active = False
            db.session.commit()
            return jsonify({'valid': False, 'message': '邀请码已过期'})
        if invite.max_uses > 0 and invite.used_count >= invite.max_uses:
            return jsonify({'valid': False, 'message': '邀请码已达到使用上限'})
        ledger = Ledger.query.get(invite.ledger_id)
        if not ledger or not ledger.is_active:
            return jsonify({'valid': False, 'message': '账本不存在或已停用'})
        return jsonify({'valid': True, 'message': '邀请码有效', 'ledger_name': ledger.name, 'ledger_id': ledger.id})
    except Exception as e:
        app.logger.error(f"校验邀请码失败: {e}")
        return jsonify({'valid': False, 'message': '校验失败'}), 500


@app.route('/api/ledgers/join', methods=['POST'])
@login_required
def join_ledger():
    """通过邀请码加入账本"""
    try:
        data = request.get_json()
        code_str = data.get('code', '').strip()
        if not code_str:
            return jsonify({'success': False, 'message': '邀请码不能为空'}), 400
        invite = InviteCode.query.filter_by(code=code_str, is_active=True).first()
        if not invite:
            return jsonify({'success': False, 'message': '邀请码无效或已失效'}), 404
        if invite.expires_at and invite.expires_at < datetime.now():
            invite.is_active = False
            db.session.commit()
            return jsonify({'success': False, 'message': '邀请码已过期'}), 400
        if invite.max_uses > 0 and invite.used_count >= invite.max_uses:
            return jsonify({'success': False, 'message': '邀请码已达到使用上限'}), 400
        ledger = Ledger.query.get(invite.ledger_id)
        if not ledger or not ledger.is_active:
            return jsonify({'success': False, 'message': '账本不存在或已停用'}), 404
        user_id = session.get('user_id')
        existing = LedgerMember.query.filter_by(ledger_id=invite.ledger_id, user_id=user_id).first()
        if existing:
            return jsonify({'success': False, 'message': '您已经是该账本的成员'}), 400
        member = LedgerMember(ledger_id=invite.ledger_id, user_id=user_id, role='viewer')
        db.session.add(member)
        invite.used_count += 1
        db.session.commit()
        session['active_ledger_id'] = invite.ledger_id
        return jsonify({'success': True, 'message': f'已加入账本「{ledger.name}」', 'ledger': {
            'id': ledger.id, 'name': ledger.name, 'role': 'viewer'
        }})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"加入账本失败: {e}")
        return jsonify({'success': False, 'message': '加入账本失败'}), 500


@app.route('/data.json')
@login_required
def get_json_data():
    try:
        # 根据权限过滤数据
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)
        current_ledger_id = get_current_ledger_id()

        if current_ledger_id:
            transactions = Transaction.query.filter(Transaction.ledger_id == current_ledger_id).order_by(Transaction.id.desc()).all()
        elif user_id and is_admin and not self_view:
            transactions = Transaction.query.order_by(Transaction.id.desc()).all()
        elif user_id:
            transactions = Transaction.query.filter(Transaction.user_id == user_id).order_by(Transaction.id.desc()).all()
        else:
            transactions = Transaction.query.order_by(Transaction.id.desc()).all()
        categories = [c.name for c in Category.query.all()]
        return jsonify({
            "transactions": [
                {
                    "id": t.id,
                    "type": t.type,
                    "amount": t.amount,
                    "category": t.category,
                    "date": t.date,
                    "time": t.time,
                    "remark": t.remark,
                    "currency": t.currency or 'CNY',
                    "original_amount": float(t.original_amount) if t.original_amount else None,
                    "exchange_rate": float(t.exchange_rate) if t.exchange_rate else None,
                    "created_at": t.created_at,
                    "updated_at": t.updated_at,
                    "reimbursement_status": t.reimbursement_status or 'none',
                    "reimbursed_amount": float(t.reimbursed_amount) if t.reimbursed_amount else 0,
                    "write_off_id": t.write_off_id,
                    "user_id": t.user_id,
                    "username": t.user.username if t.user else None
                } for t in transactions
            ],
            "categories": categories,
            "balance": get_balance()
        })
    except Exception as e:
        return jsonify({"error": str(e), "status": 500}), 500


# ==================== 辅助函数 ====================
def get_balance():
    user_id = session.get('user_id')
    is_admin = session.get('is_admin')
    self_view = session.get('self_view', False)

    if not user_id:
        return 0

    # 如果 self_view 模式，按当前账本计算余额
    current_ledger_id = get_current_ledger_id()
    if current_ledger_id:
        income = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.ledger_id == current_ledger_id, Transaction.type == 'income'
        ).scalar() or 0
        expense = db.session.query(db.func.sum(Transaction.amount)).filter(
            Transaction.ledger_id == current_ledger_id, Transaction.type == 'expense'
        ).scalar() or 0
    elif is_admin and not self_view:
        income = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='income').scalar() or 0
        expense = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='expense').scalar() or 0
    else:
        income = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='income', user_id=user_id).scalar() or 0
        expense = db.session.query(db.func.sum(Transaction.amount)).filter_by(type='expense', user_id=user_id).scalar() or 0

    return float(income - expense)

def filter_transactions_by_period_orm(period=None, date=None):
    query = Transaction.query.options(db.joinedload(Transaction.user))

    now = datetime.now()

    # 如果存在当前账本，按账本过滤
    current_ledger_id = get_current_ledger_id()
    if current_ledger_id:
        query = query.filter(Transaction.ledger_id == current_ledger_id)
    else:
        # 管理员可以看到所有用户的交易数据，普通用户只能看到自己的数据
        # self_view 模式下管理员也只看自己的数据（从个人首页访问时）
        user_id = session.get('user_id')
        is_admin = session.get('is_admin')
        self_view = session.get('self_view', False)

        if user_id and (not is_admin or self_view):
            query = query.filter(Transaction.user_id == user_id)

    if date:
        query = query.filter(Transaction.date == date)
    elif period == 'day':
        today = now.strftime('%Y-%m-%d')
        query = query.filter(Transaction.date == today)
    elif period == 'week':
        query = query.order_by(Transaction.date.desc()).limit(50)
    elif period == 'month':
        current_month = now.strftime('%Y-%m')
        query = query.filter(Transaction.date.startswith(current_month))
    elif period == 'year':
        current_year = now.strftime('%Y')
        query = query.filter(Transaction.date.startswith(current_year))
    return query.order_by(Transaction.id.desc())


# ==================== 启动应用 ====================
# 注册蓝图
from blueprints.import_export import import_export_bp
app.register_blueprint(import_export_bp)
from blueprints.smart_bookkeeping import smart_bp
app.register_blueprint(smart_bp)

if __name__ == '__main__':
    # 从环境变量读取端口，默认8080
    port = int(os.getenv('PORT', config.PORT))
    
    app.logger.info("=" * 50)
    app.logger.info(f"启动 {config.APP_NAME}")
    app.logger.info(f"环境: {os.getenv('FLASK_ENV', 'development')}")
    app.logger.info(f"主机: {config.HOST}:{port}")
    app.logger.info("=" * 50)
    
    # 生产环境使用 gunicorn，这里仅用于开发
    app.run(
        host=config.HOST,
        port=port,
        debug=app.config.get('DEBUG', False)
    )

# 保证表已存在
def initialize_db():
    with app.app_context():
        db.create_all()

        # ---- 数据库迁移：为 users 表添加后续新增的字段（必须在任何 ORM 查询前执行） ----
        for col_def in [
            ('avatar', 'ALTER TABLE users ADD COLUMN avatar VARCHAR(255) DEFAULT "default_avatar.png"'),
            ('email', 'ALTER TABLE users ADD COLUMN email VARCHAR(120) DEFAULT NULL'),
            ('phone', 'ALTER TABLE users ADD COLUMN phone VARCHAR(20) DEFAULT NULL'),
            ('nickname', 'ALTER TABLE users ADD COLUMN nickname VARCHAR(64) DEFAULT NULL'),
            ('created_at', 'ALTER TABLE users ADD COLUMN created_at DATETIME DEFAULT CURRENT_TIMESTAMP'),
            ('last_login', 'ALTER TABLE users ADD COLUMN last_login DATETIME DEFAULT NULL'),
        ]:
            col_name, alter_sql = col_def
            try:
                db.session.execute(db.text(f'SELECT {col_name} FROM users LIMIT 0'))
            except Exception:
                try:
                    db.session.execute(db.text(alter_sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        # ---- 数据库迁移：为 Category 表添加 type 字段 ----
        try:
            db.session.execute(db.text('SELECT type FROM categories LIMIT 0'))
        except Exception:
            try:
                db.session.execute(db.text('ALTER TABLE categories ADD COLUMN type VARCHAR(10) NOT NULL DEFAULT "expense"'))
                db.session.commit()
            except Exception:
                db.session.rollback()

        # ---- 数据库迁移：为 transactions 表添加 account_id / user_id 字段 ----
        for col_def in [
            ('account_id', 'ALTER TABLE transactions ADD COLUMN account_id INT DEFAULT NULL'),
            ('user_id', 'ALTER TABLE transactions ADD COLUMN user_id INT DEFAULT NULL'),
        ]:
            col_name, alter_sql = col_def
            try:
                db.session.execute(db.text(f'SELECT {col_name} FROM transactions LIMIT 0'))
            except Exception:
                try:
                    db.session.execute(db.text(alter_sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        # ---- 数据库迁移：为 transactions 表添加 currency / original_amount / exchange_rate 字段 ----
        for col_def in [
            ('currency', 'ALTER TABLE transactions ADD COLUMN currency VARCHAR(10) DEFAULT "CNY"'),
            ('original_amount', 'ALTER TABLE transactions ADD COLUMN original_amount DECIMAL(10,2) DEFAULT NULL'),
            ('exchange_rate', 'ALTER TABLE transactions ADD COLUMN exchange_rate DECIMAL(10,6) DEFAULT NULL'),
        ]:
            col_name, alter_sql = col_def
            try:
                db.session.execute(db.text(f'SELECT {col_name} FROM transactions LIMIT 0'))
            except Exception:
                try:
                    db.session.execute(db.text(alter_sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        # ---- 数据库迁移：为 transactions 表添加 reimbursement 相关字段 ----
        for col_def in [
            ('reimbursement_status', 'ALTER TABLE transactions ADD COLUMN reimbursement_status VARCHAR(20) DEFAULT "none"'),
            ('reimbursed_amount', 'ALTER TABLE transactions ADD COLUMN reimbursed_amount DECIMAL(10,2) DEFAULT 0'),
            ('write_off_id', 'ALTER TABLE transactions ADD COLUMN write_off_id INT DEFAULT NULL'),
        ]:
            col_name, alter_sql = col_def
            try:
                db.session.execute(db.text(f'SELECT {col_name} FROM transactions LIMIT 0'))
            except Exception:
                try:
                    db.session.execute(db.text(alter_sql))
                    db.session.commit()
                    # Add FK constraint for write_off_id
                    if col_name == 'write_off_id':
                        try:
                            db.session.execute(db.text(
                                'ALTER TABLE transactions ADD CONSTRAINT fk_write_off FOREIGN KEY (write_off_id) REFERENCES transactions(id) ON DELETE SET NULL'
                            ))
                            db.session.commit()
                        except Exception:
                            db.session.rollback()
                except Exception:
                    db.session.rollback()

        # ---- 数据库迁移：为 budgets / budget_category_items 建表 ----
        try:
            db.session.execute(db.text('SELECT 1 FROM budgets LIMIT 0'))
        except Exception:
            try:
                db.session.execute(db.text('''
                    CREATE TABLE IF NOT EXISTS budgets (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        user_id INT NOT NULL,
                        account_id INT DEFAULT NULL,
                        month VARCHAR(7) NOT NULL,
                        total_amount DECIMAL(10,2) NOT NULL,
                        remark VARCHAR(255),
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id),
                        FOREIGN KEY (account_id) REFERENCES accounts(id)
                    )
                '''))
                db.session.execute(db.text('''
                    CREATE TABLE IF NOT EXISTS budget_category_items (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        budget_id INT NOT NULL,
                        category_id INT NOT NULL,
                        amount DECIMAL(10,2) NOT NULL,
                        FOREIGN KEY (budget_id) REFERENCES budgets(id),
                        FOREIGN KEY (category_id) REFERENCES categories(id)
                    )
                '''))
                db.session.commit()
            except Exception:
                db.session.rollback()

        # ---- 数据库迁移：为 budgets 表添加 account_id 字段 ----
        try:
            db.session.execute(db.text('SELECT account_id FROM budgets LIMIT 0'))
        except Exception:
            try:
                db.session.execute(db.text('ALTER TABLE budgets ADD COLUMN account_id INT DEFAULT NULL, ADD FOREIGN KEY (account_id) REFERENCES accounts(id)'))
                db.session.commit()
            except Exception:
                db.session.rollback()

        # ---- 数据库迁移：export_tasks / file_uploads 表 ----
        for table_name, create_sql in [
            ('export_tasks', '''
                CREATE TABLE IF NOT EXISTS export_tasks (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    status VARCHAR(20) DEFAULT 'pending',
                    progress INT DEFAULT 0,
                    file_format VARCHAR(10) NOT NULL,
                    file_path VARCHAR(500),
                    file_size INT,
                    filters TEXT,
                    total_records INT DEFAULT 0,
                    error_message TEXT,
                    email_to VARCHAR(200),
                    email_sent TINYINT(1) DEFAULT 0,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    completed_at DATETIME,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            '''),
            ('file_uploads', '''
                CREATE TABLE IF NOT EXISTS file_uploads (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    upload_id VARCHAR(36) UNIQUE NOT NULL,
                    user_id INT NOT NULL,
                    original_filename VARCHAR(255) NOT NULL,
                    file_path VARCHAR(500) NOT NULL,
                    file_format VARCHAR(10) NOT NULL,
                    total_rows INT DEFAULT 0,
                    columns TEXT,
                    preview_data TEXT,
                    status VARCHAR(20) DEFAULT 'uploaded',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            '''),
        ]:
            try:
                db.session.execute(db.text(f'SELECT 1 FROM {table_name} LIMIT 0'))
            except Exception:
                try:
                    db.session.execute(db.text(create_sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        # ---- 默认分类 ----
        if Category.query.count() == 0:
            default_income = ['工资', '奖金', '投资收益', '兼职', '红包', '报销收入', '其他收入']
            default_expense = ['餐饮', '交通', '购物', '娱乐', '医疗', '住房', '教育', '通讯', '其他支出']
            for cname in default_income:
                db.session.add(Category(name=cname, type='income'))
            for cname in default_expense:
                db.session.add(Category(name=cname, type='expense'))
        elif not Category.query.filter_by(name='报销收入', type='income').first():
            db.session.add(Category(name='报销收入', type='income'))

        # ---- 默认用户 ----
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', is_admin=True)
            admin.set_password('admin123')
            db.session.add(admin)
        if not User.query.filter_by(username='user').first():
            user = User(username='user', is_admin=False)
            user.set_password('user123')
            db.session.add(user)
        db.session.commit()

        # ---- 数据库迁移：创建账本相关表 ----
        for table_name, create_sql in [
            ('ledgers', '''
                CREATE TABLE IF NOT EXISTS ledgers (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    name VARCHAR(100) NOT NULL,
                    description VARCHAR(255),
                    owner_id INT NOT NULL,
                    currency VARCHAR(10) DEFAULT 'CNY',
                    is_active TINYINT(1) DEFAULT 1,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (owner_id) REFERENCES users(id)
                )
            '''),
            ('ledger_members', '''
                CREATE TABLE IF NOT EXISTS ledger_members (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    ledger_id INT NOT NULL,
                    user_id INT NOT NULL,
                    role VARCHAR(20) DEFAULT 'viewer',
                    joined_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (ledger_id) REFERENCES ledgers(id),
                    FOREIGN KEY (user_id) REFERENCES users(id),
                    UNIQUE KEY uq_ledger_user (ledger_id, user_id)
                )
            '''),
            ('invite_codes', '''
                CREATE TABLE IF NOT EXISTS invite_codes (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    ledger_id INT NOT NULL,
                    code VARCHAR(64) UNIQUE NOT NULL,
                    created_by INT NOT NULL,
                    max_uses INT DEFAULT 0,
                    used_count INT DEFAULT 0,
                    expires_at DATETIME,
                    is_active TINYINT(1) DEFAULT 1,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (ledger_id) REFERENCES ledgers(id),
                    FOREIGN KEY (created_by) REFERENCES users(id)
                )
            '''),
        ]:
            try:
                db.session.execute(db.text(f'SELECT 1 FROM {table_name} LIMIT 0'))
            except Exception:
                try:
                    db.session.execute(db.text(create_sql))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        # ---- 数据库迁移：为 accounts / transactions / budgets 添加 ledger_id 字段 ----
        # 先为 accounts 添加，后续创建默认账户时 ORM 不会报错
        for table_name in ['accounts', 'transactions', 'budgets']:
            try:
                db.session.execute(db.text(f'SELECT ledger_id FROM {table_name} LIMIT 0'))
            except Exception:
                try:
                    db.session.execute(db.text(f'ALTER TABLE {table_name} ADD COLUMN ledger_id INT DEFAULT NULL'))
                    db.session.commit()
                except Exception:
                    db.session.rollback()

        # ---- 为没有账户的用户创建默认账户（带 ledger_id） ----
        for u in User.query.all():
            if Account.query.filter_by(user_id=u.id).count() == 0:
                db.session.add(Account(name='默认账户', balance=0, account_type='cash', user_id=u.id))
        db.session.commit()

        # ---- 为已有用户创建个人默认账本 ----
        for u in User.query.all():
            if Ledger.query.filter_by(owner_id=u.id).count() == 0:
                ledger = Ledger(name=f"{u.username}的个人账本", owner_id=u.id)
                db.session.add(ledger)
                db.session.flush()
                member = LedgerMember(ledger_id=ledger.id, user_id=u.id, role='manager')
                db.session.add(member)
                db.session.flush()
                Transaction.query.filter_by(user_id=u.id, ledger_id=None).update({'ledger_id': ledger.id})
                Account.query.filter_by(user_id=u.id, ledger_id=None).update({'ledger_id': ledger.id})
                Budget.query.filter_by(user_id=u.id, ledger_id=None).update({'ledger_id': ledger.id})
        db.session.commit()

        # ---- 数据库迁移：创建借贷表 ----
        try:
            db.session.execute(db.text('SELECT 1 FROM loans LIMIT 0'))
        except Exception:
            try:
                db.session.execute(db.text('''
                    CREATE TABLE IF NOT EXISTS loans (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        user_id INT NOT NULL,
                        ledger_id INT DEFAULT NULL,
                        type VARCHAR(10) NOT NULL,
                        counterparty VARCHAR(100) NOT NULL,
                        amount DECIMAL(10,2) NOT NULL,
                        repaid_amount DECIMAL(10,2) DEFAULT 0,
                        date VARCHAR(20) NOT NULL,
                        due_date VARCHAR(20) DEFAULT NULL,
                        status VARCHAR(20) DEFAULT 'active',
                        remark VARCHAR(255) DEFAULT NULL,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users(id),
                        FOREIGN KEY (ledger_id) REFERENCES ledgers(id)
                    )
                '''))
                db.session.commit()
            except Exception:
                db.session.rollback()

initialize_db()
